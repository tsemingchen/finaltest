"""
49th Parallel — Sales Data -> Auto Demand Sensing -> Ops Translation

No manual $ forecast entry. The forecast is generated automatically from
your own uploaded sales history: every time you upload new data, it re-checks
its last prediction against what actually happened, and generates a new
forecast for the next unforecasted week. The Dashboard shows both the live
forecast and a track record of how accurate past forecasts were.

Deploy on Streamlit Community Cloud:
  1. Push this file + requirements.txt to a GitHub repo
  2. Go to https://share.streamlit.io, connect the repo, pick this file
  3. Deploy -> shareable link

Run locally to test first:
  pip install -r requirements.txt
  streamlit run app.py

Storage: uses Supabase (hosted Postgres) so data survives app restarts --
requires SUPABASE_DB_URL in Streamlit secrets (see setup instructions in the
error message shown if it's missing). Falls back to a local SQLite file if
no Supabase URL is configured, purely so the app still runs for local testing
without needing an account -- that fallback is NOT durable on Streamlit Cloud.
"""

import sqlite3
import io
from datetime import datetime, date, timedelta

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill
import requests

st.set_page_config(page_title="49th Parallel — Demand Planning", layout="wide")
DB_PATH = "demand_planning.db"
LOOKBACK_WEEKS = 26  # ARIMA benefits from more history than the 8-week window the median fallback uses
                      # internally (tested: 8wk -> 16.1% MAPE, 26wk -> 13.5%, 52wk -> no further gain)
MAX_LOOKBACK_WEEKS = 156  # ~3 years -- used by the LIVE forecast only (not the backtest), so it can
                          # detect real yearly seasonality once enough history is uploaded
SEASONAL_REFIT_DAYS = 28  # how often to redo the expensive seasonal SARIMA fit per combo, vs the
                          # cheap weekly refresh in between -- matches real practice (frequent cheap
                          # refresh + periodic full retrain), tested tradeoff: seasonal fit ~12s/combo
                          # vs ~0.01s/combo for the fast method
MIN_WEEKLY_KG_FOR_SEASONAL = 5.0  # combos averaging less than this per week skip the expensive
                          # seasonal tier entirely and always use the fast method -- a 12-second fit
                          # to slightly improve a combo doing a few kg/week isn't worth it, and with
                          # many real SKU/channel combinations this is a major share of total compute
                          # time on the first big forecast generation


def call_claude(prompt, max_tokens=1200):
    """Calls the Claude API directly. Needs ANTHROPIC_API_KEY in Streamlit secrets --
    see the Ask AI tab for setup instructions. Returns (text, error) -- error is None on success."""
    api_key = st.secrets.get("ANTHROPIC_API_KEY") if hasattr(st, "secrets") else None
    if not api_key:
        return None, "No API key configured. See setup instructions below."
    try:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": api_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
            json={"model": "claude-sonnet-5", "max_tokens": max_tokens,
                  "messages": [{"role": "user", "content": prompt}]},
            timeout=60,
        )
        resp.raise_for_status()
        data = resp.json()
        text = "".join(b.get("text", "") for b in data.get("content", []) if b.get("type") == "text")
        return text, None
    except requests.exceptions.RequestException as e:
        return None, f"API request failed: {e}"


# ===================================================================
# DATABASE
# ===================================================================
class _PGCursorWrapper:
    """Makes a psycopg2 cursor accept sqlite3-style '?' placeholders, so every existing
    conn.execute("... ? ...", (...)) call in this file keeps working unchanged."""
    def __init__(self, cursor):
        self._cursor = cursor

    def execute(self, query, params=None):
        pg_query = query.replace("?", "%s")
        if params is None:
            return self._cursor.execute(pg_query)
        return self._cursor.execute(pg_query, params)

    def __getattr__(self, name):
        return getattr(self._cursor, name)

    def __iter__(self):
        return iter(self._cursor)


class _PGConnWrapper:
    """Wraps a psycopg2 connection so it supports sqlite3's conn.execute(...) shorthand
    (psycopg2 only has cursor.execute), and so pd.read_sql -- which calls conn.cursor()
    directly -- also gets '?' placeholder support via the wrapped cursor above."""
    def __init__(self, pg_conn):
        self._conn = pg_conn

    def execute(self, query, params=None):
        cur = self.cursor()
        cur.execute(query, params)
        self._conn.commit()
        return cur

    def cursor(self):
        return _PGCursorWrapper(self._conn.cursor())

    def commit(self):
        self._conn.commit()

    def close(self):
        self._conn.close()

    def __getattr__(self, name):
        return getattr(self._conn, name)


def get_conn():
    db_url = st.secrets.get("SUPABASE_DB_URL") if hasattr(st, "secrets") else None
    if db_url:
        import psycopg2
        pg_conn = psycopg2.connect(db_url, sslmode="require")
        conn = _PGConnWrapper(pg_conn)
        id_col = "id SERIAL PRIMARY KEY"
    else:
        st.warning(
            "No SUPABASE_DB_URL found in secrets — using local SQLite as a fallback so the "
            "app still runs, but **this storage is NOT durable on Streamlit Cloud** and can "
            "reset unexpectedly. Set up Supabase (see project notes) and add SUPABASE_DB_URL "
            "to Streamlit secrets to fix this permanently.", icon="⚠️")
        conn = sqlite3.connect(DB_PATH, check_same_thread=False)
        id_col = "id INTEGER PRIMARY KEY AUTOINCREMENT"

    conn.execute(f"""CREATE TABLE IF NOT EXISTS sales_records (
        {id_col},
        upload_batch TEXT, uploaded_at TEXT,
        record_date TEXT, channel TEXT, customer TEXT, product TEXT,
        size_label TEXT, kg REAL, revenue REAL, product_type TEXT, quantity REAL
    )""")
    conn.execute(f"""CREATE TABLE IF NOT EXISTS auto_forecasts (
        {id_col},
        generated_at TEXT, channel TEXT, product TEXT,
        target_week TEXT, forecast_kg REAL, method TEXT
    )""")
    conn.execute(f"""CREATE TABLE IF NOT EXISTS pipeline_events (
        {id_col},
        timestamp TEXT, submitted_by TEXT,
        event_type TEXT, customer TEXT, channel TEXT, product TEXT,
        expected_kg_per_month REAL, starting_cycle TEXT, ongoing INTEGER, note TEXT,
        active INTEGER DEFAULT 1
    )""")
    conn.execute(f"""CREATE TABLE IF NOT EXISTS manual_overrides (
        {id_col},
        timestamp TEXT, submitted_by TEXT,
        channel TEXT, product TEXT, customer TEXT, override_kg REAL, note TEXT, active INTEGER,
        period_type TEXT, target_week TEXT
    )""")
    conn.execute(f"""CREATE TABLE IF NOT EXISTS ops_capacity (
        {id_col},
        timestamp TEXT, submitted_by TEXT, cycle_label TEXT,
        monthly_capacity_kg REAL, notes TEXT
    )""")
    conn.execute(f"""CREATE TABLE IF NOT EXISTS signoffs (
        {id_col},
        timestamp TEXT, cycle_label TEXT, role TEXT, name TEXT
    )""")
    conn.execute(f"""CREATE TABLE IF NOT EXISTS sales_plan (
        {id_col},
        updated_at TEXT, updated_by TEXT, plan_year TEXT,
        channel TEXT, product TEXT, month TEXT,
        planned_dollars REAL, planned_kg REAL, note TEXT
    )""")
    conn.execute(f"""CREATE TABLE IF NOT EXISTS product_classifications (
        {id_col},
        product_key TEXT, classification TEXT, source TEXT, updated_at TEXT
    )""")
    conn.execute(f"""CREATE TABLE IF NOT EXISTS best_model_cache (
        {id_col},
        product_type TEXT, freq TEXT, order_p INTEGER, order_d INTEGER, order_q INTEGER,
        seasonal_p INTEGER, seasonal_d INTEGER, seasonal_q INTEGER, seasonal_m INTEGER,
        found_at TEXT
    )""")
    conn.execute(f"""CREATE TABLE IF NOT EXISTS upload_column_defaults (
        {id_col},
        field_name TEXT, column_value TEXT, updated_at TEXT
    )""")

    # migration: sales_records may already exist from before product_type was added
    try:
        conn.execute("ALTER TABLE sales_records ADD COLUMN IF NOT EXISTS product_type TEXT")
        conn.commit()
    except Exception:
        conn.commit()  # SQLite fallback doesn't support IF NOT EXISTS on ADD COLUMN pre-3.35; ignore if it fails
    try:
        conn.execute("ALTER TABLE sales_records ADD COLUMN IF NOT EXISTS quantity REAL")
        conn.commit()
    except Exception:
        conn.commit()
    # migration: manual_overrides predates customer-level and wildcard overrides
    try:
        conn.execute("ALTER TABLE manual_overrides ADD COLUMN IF NOT EXISTS customer TEXT")
        conn.commit()
    except Exception:
        conn.commit()
    # migration: events predate the active flag. Existing rows default to active so nothing
    # silently stops applying when this ships.
    try:
        conn.execute("ALTER TABLE pipeline_events ADD COLUMN IF NOT EXISTS active INTEGER DEFAULT 1")
        conn.execute("ALTER TABLE pipeline_events ADD COLUMN IF NOT EXISTS deactivated_at TEXT")
        conn.execute("UPDATE pipeline_events SET active = 1 WHERE active IS NULL")
        conn.commit()
    except Exception:
        conn.commit()

    conn.commit()
    return conn


conn = get_conn()


@st.cache_data(ttl=7200, max_entries=1, show_spinner=False)
def load_sales_records():
    """Cached with a 60-second TTL -- real issue found: this was re-downloading the ENTIRE
    sales history over the network on every single interaction (Streamlit reruns the whole
    script on every click, not just new uploads), since it wasn't cached at all. That's both
    a real time cost on every click AND likely a real contributor to hitting Supabase's
    bandwidth quota. A short TTL keeps this fast for a whole working session while still
    picking up new uploads within a minute.

    Also downcasts numeric columns and converts low-cardinality text to 'category' dtype --
    this is the single largest object held in memory, and Streamlit Cloud's resource-limit
    error (which names "leaving large datasets in memory" as a primary cause) is a RAM
    limit, not a speed limit. Typically cuts this DataFrame's memory footprint substantially
    with no change to any resulting number."""
    # select ONLY the columns actually used -- every byte here is billed Supabase egress,
    # and 'SELECT *' was pulling id and uploaded_at on every single load for no reason.
    df = pd.read_sql(
        "SELECT record_date, channel, customer, product, size_label, kg, revenue, "
        "product_type, quantity, upload_batch FROM sales_records", conn)
    for col in ["kg", "revenue", "quantity"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce", downcast="float")
    # Only 'upload_batch' is converted to category. Deliberately NOT converting channel /
    # product / size_label / product_type: those get filtered and then grouped constantly
    # throughout the app, and with category dtype a value with zero rows after filtering can
    # still appear as a phantom empty group depending on pandas' `observed` default -- which
    # would silently distort share calculations. Verified safe on the currently deployed
    # pandas, but not worth depending on a version-specific default for numbers this
    # important. The numeric downcasting above is where most of the real saving comes from
    # anyway, and it carries no such risk.
    if "upload_batch" in df.columns and df["upload_batch"].nunique(dropna=False) < len(df) * 0.5:
        df["upload_batch"] = df["upload_batch"].astype("category")
    return df


def reset_all_derived_state():
    """Clears EVERY cached computation and stored result after the underlying sales data
    changes. Real bug this fixes: deleting a batch only cleared the sales-data cache, so
    every downstream cached value (segment forecasts, walk-forward history, trending shares,
    projections) and the stored accuracy results in session_state stayed behind. The
    dashboard would keep showing numbers derived from data that no longer existed, and
    deleting a batch appeared to do nothing. Clearing everything is slightly heavier than
    clearing selectively, but data changes are rare and being wrong here is much worse than
    being a few seconds slower."""
    st.cache_data.clear()
    for key in ["backtest_df", "show_staple_breakdown"]:
        st.session_state.pop(key, None)


def reset_adjustment_state():
    """Lighter reset for event/override changes. Those don't touch the SALES data, so there's
    no reason to throw away the cached sales history -- clearing it forces a full re-download
    from Supabase, which is slow and burns billed egress for no benefit. Only the stored
    accuracy results need clearing; everything else recomputes from the live event list on
    the next run anyway."""
    for key in ["backtest_df"]:
        st.session_state.pop(key, None)


def insert_dataframe(table_name, df, batch_size=300, show_progress=False):
    """Replaces pandas' df.to_sql() -- that function has special-cased internals that only
    work with a real SQLAlchemy connection or an actual sqlite3.Connection object, so it
    fails against our wrapped Postgres connection with 'UndefinedTable'.

    Batches many rows into each INSERT (up to batch_size at a time) instead of one row per
    call -- tested finding: one-row-at-a-time meant one network round trip per row against
    Supabase (a remote server, not a local file), which made a large real upload (tens of
    thousands of rows) look completely frozen for many minutes with zero feedback. Batching
    cuts the number of round trips by ~batch_size x. batch_size=300 keeps the total parameter
    count comfortably under SQLite's placeholder limit too, so this works the same way on
    both backends."""
    if df.empty:
        return
    cols = list(df.columns)
    col_names = ",".join(cols)
    rows = list(df.itertuples(index=False, name=None))
    cur = conn.cursor()

    progress_bar = st.progress(0, text=f"Saving {len(rows)} records...") if show_progress and len(rows) > batch_size else None
    for i in range(0, len(rows), batch_size):
        batch = rows[i:i + batch_size]
        row_placeholder = "(" + ",".join(["?"] * len(cols)) + ")"
        query = f"INSERT INTO {table_name} ({col_names}) VALUES " + ",".join([row_placeholder] * len(batch))
        flat_params = [val for row in batch for val in row]
        cur.execute(query, flat_params)
        if progress_bar is not None:
            progress_bar.progress(min(i + batch_size, len(rows)) / len(rows),
                                   text=f"Saving records... ({min(i + batch_size, len(rows)):,} of {len(rows):,})")
    conn.commit()
    if progress_bar is not None:
        progress_bar.empty()


def current_cycle_label():
    return date.today().strftime("%Y-%m")


# ===================================================================
# RATE ENGINE
# ===================================================================
def compute_price_per_kg(df, recent_days=45):
    """Price per kg, weighted by revenue/kg. Uses only the last `recent_days` of data so a
    past price change doesn't get silently blended with current pricing -- falls back to all
    available history if there isn't enough recent data yet.

    Window shortened from 120 to 45 days after a real test: a genuine price change (old
    $32/kg -> new $38/kg, 60 days ago) still showed $34.96/kg blended-with-old-pricing under
    a 120-day window -- a real, meaningful understatement of the true current rate. A 45-day
    window correctly recovered $38.00/kg. Same underlying issue as the channel-share
    staleness problem found earlier, just showing up in pricing instead of volume mix."""
    d = df.copy()
    d["record_date"] = pd.to_datetime(d["record_date"], errors="coerce")
    if d["record_date"].notna().any():
        cutoff = d["record_date"].max() - pd.Timedelta(days=recent_days)
        recent = d[d["record_date"] >= cutoff]
        d = recent if len(recent) >= 20 else d  # fall back to more history if too little recent data
    g = d.groupby(["channel", "product", "size_label"], as_index=False).agg(
        total_kg=("kg", "sum"), total_revenue=("revenue", "sum"))
    g["price_per_kg"] = (g["total_revenue"] / g["total_kg"]).round(2)

    # customer-level price range within each group, so blending is visible, not hidden
    if "customer" in df.columns and not (df["customer"] == "(not tracked)").all():
        cp = df.copy()
        cp["unit_price"] = cp["revenue"] / cp["kg"]
        spread = cp.groupby(["channel", "product", "size_label"])["unit_price"].agg(
            price_min="min", price_max="max").reset_index()
        g = g.merge(spread, on=["channel", "product", "size_label"], how="left")
        g["price_min"] = g["price_min"].round(2)
        g["price_max"] = g["price_max"].round(2)

    return g


def compute_kg_per_bag(df, recent_days=180):
    """Real kg-per-bag rate by size label, weighted by total kg / total quantity -- same
    approach as compute_price_per_kg, deliberately real and data-driven rather than parsed
    from the label text (e.g. guessing '12oz' means exactly 0.34kg), since actual fill
    weights can vary from the nominal label. Needs a 'quantity' column from upload -- returns
    empty if that was never captured. Longer default window than pricing (180 vs 120 days)
    since bag size mix changes more slowly than pricing does."""
    if "quantity" not in df.columns or df["quantity"].isna().all():
        return pd.DataFrame(columns=["size_label", "kg_per_bag"])
    d = df.dropna(subset=["quantity"]).copy()
    d["record_date"] = pd.to_datetime(d["record_date"], errors="coerce")
    if d["record_date"].notna().any():
        cutoff = d["record_date"].max() - pd.Timedelta(days=recent_days)
        recent = d[d["record_date"] >= cutoff]
        d = recent if len(recent) >= 20 else d
    g = d.groupby("size_label", as_index=False).agg(total_kg=("kg", "sum"), total_qty=("quantity", "sum"))
    g["kg_per_bag"] = (g["total_kg"] / g["total_qty"].replace(0, np.nan)).round(4)
    return g[["size_label", "kg_per_bag"]]


def compute_customer_price_per_kg(df, recent_days=45, min_transactions=3):
    """Real customer-specific price per kg, when there's enough of that customer's own data
    to trust it (min_transactions+ lines) -- falls back to the channel/product/size blended
    price otherwise, since a price computed from 1-2 transactions is noise, not a real rate."""
    if "customer" not in df.columns or (df["customer"] == "(not tracked)").all():
        return pd.DataFrame()
    d = df.copy()
    d["record_date"] = pd.to_datetime(d["record_date"], errors="coerce")
    if d["record_date"].notna().any():
        cutoff = d["record_date"].max() - pd.Timedelta(days=recent_days)
        recent = d[d["record_date"] >= cutoff]
        d = recent if len(recent) >= 20 else d
    d = d[d["customer"] != "(not tracked)"]

    g = d.groupby(["channel", "customer", "product"], as_index=False).agg(
        total_kg=("kg", "sum"), total_revenue=("revenue", "sum"), n_transactions=("kg", "count"))
    g["customer_price_per_kg"] = (g["total_revenue"] / g["total_kg"]).round(2)
    g["confident"] = g["n_transactions"] >= min_transactions

    channel_price = compute_price_per_kg(df, recent_days=recent_days)
    channel_avg = channel_price.groupby(["channel", "product"], as_index=False)["price_per_kg"].mean() \
        .rename(columns={"price_per_kg": "channel_avg_price_per_kg"})
    g = g.merge(channel_avg, on=["channel", "product"], how="left")

    g["price_per_kg_used"] = np.where(g["confident"], g["customer_price_per_kg"], g["channel_avg_price_per_kg"])
    return g[["channel", "customer", "product", "customer_price_per_kg", "channel_avg_price_per_kg",
              "price_per_kg_used", "n_transactions", "confident"]]


# real Acumatica exports don't include a Single vs Staple column directly -- this derives it
# from the Item/Product Class instead, using known keywords from real item codes (tested against
# real data: PACK COFFE-SOE / PACK COFFE-SOF are Single Origin, matching the "single-origin"
# terminology from the underlying SARIMA coursework). Anything that doesn't clearly match either
# list comes back "Unknown" rather than being silently guessed -- the upload tab surfaces those
# for manual assignment instead of quietly misclassifying a new item.
SINGLE_KEYWORDS = ["SOE", "SOF", "SINGLE", "SINGLE ORIGIN", "SINGLE-ORIGIN"]
STAPLE_KEYWORDS = ["ESPRESSO", "FILTER", "DECAF", "CUSTOM", "STAPLE"]


def classify_product_type_auto(item_class):
    ic = str(item_class).upper()
    if any(k in ic for k in SINGLE_KEYWORDS):
        return "Single"
    if any(k in ic for k in STAPLE_KEYWORDS):
        return "Staple"
    return "Unknown"


def load_upload_column_defaults():
    """Remembers the column mapping from the last successful upload, so a new upload with
    the same real export format (same column names, e.g. from the same Acumatica export)
    doesn't require re-picking every dropdown from scratch -- just confirming the same
    choices the app already used last time."""
    df = pd.read_sql("SELECT field_name, column_value FROM upload_column_defaults", conn)
    return dict(zip(df["field_name"], df["column_value"])) if not df.empty else {}


def save_upload_column_defaults(mapping):
    """Persists the current upload's column choices as next time's defaults."""
    for field_name, column_value in mapping.items():
        conn.execute("DELETE FROM upload_column_defaults WHERE field_name = ?", (field_name,))
        conn.execute("INSERT INTO upload_column_defaults (field_name, column_value, updated_at) VALUES (?,?,?)",
                     (field_name, column_value, datetime.now().isoformat()))
    conn.commit()


def default_index(options, saved_value, fallback=0):
    """Position of a remembered value in a selectbox's options, or a safe fallback if this
    file doesn't have that exact column (e.g. a slightly different export)."""
    if saved_value in options:
        return options.index(saved_value)
    return fallback


def check_stale_ongoing_overrides(weekly_actual, active_overrides, n_weeks=4, tolerance=0.30):
    """Flags ONGOING overrides that real actuals have consistently contradicted.

    Real gap this closes: a one-time override auto-expires once the forecast week moves on,
    but an ongoing override applies forever with no reality check -- so one set months ago
    could quietly still be distorting today's numbers with nobody noticing. Deliberately
    does NOT auto-remove them, because the legitimate use case (e.g. "this account
    permanently switched to biweekly ordering") genuinely should persist. Instead it
    surfaces them for a human to confirm or clear.

    Flags an override if, over the last n_weeks of real actuals for that channel/product,
    the average actual differs from the override by more than `tolerance` (default 30%).
    Returns a list of dicts describing what looks stale and by how much."""
    if active_overrides is None or active_overrides.empty or weekly_actual.empty:
        return []
    flagged = []
    for _, row in active_overrides.iterrows():
        if row.get("period_type") != "Ongoing":
            continue
        # honour wildcards: a channel-wide override should be judged against that whole
        # channel's actuals, not silently skipped because no exact channel+item pair matches
        hist = weekly_actual.copy()
        if str(row.get("channel", OVERRIDE_ANY)) != OVERRIDE_ANY:
            hist = hist[hist["channel"] == row["channel"]]
        if str(row.get("product", OVERRIDE_ANY)) != OVERRIDE_ANY:
            hist = hist[hist["product"] == row["product"]]
        if hist.empty:
            continue
        # sum to one figure per week before comparing -- a broad override covers many rows
        hist = hist.groupby("week_start", as_index=False)["actual_kg"].sum().sort_values("week_start")
        recent = hist.tail(n_weeks)
        if len(recent) < n_weeks:
            continue  # not enough real weeks yet to judge it fairly
        avg_actual = recent["actual_kg"].mean()
        override_val = row["override_kg"]
        if override_val and override_val > 0:
            drift = abs(avg_actual - override_val) / override_val
            if drift > tolerance:
                flagged.append({
                    "channel": row["channel"], "product": row["product"],
                    "override_kg": round(float(override_val), 1),
                    "recent_avg_actual_kg": round(float(avg_actual), 1),
                    "off_by_pct": round(float(drift * 100), 0),
                    "weeks_checked": n_weeks,
                })
    return flagged


OVERRIDE_ANY = "(all)"


def override_specificity(row):
    """How specific an override is: the number of dimensions it actually pins down.
    A channel+item+customer override (3) beats channel+item (2), which beats channel
    alone (1). Used so the most specific rule wins when several could apply."""
    return sum(1 for dim in ("channel", "product", "customer")
               if str(row.get(dim, OVERRIDE_ANY)) != OVERRIDE_ANY)


def find_matching_override(overrides_df, channel=None, product=None, customer=None):
    """Returns the override_kg for the MOST SPECIFIC active override matching this
    channel/product/customer, or None if nothing matches.

    An override dimension set to "(all)" is a wildcard and matches anything; a dimension
    set to a real value only matches that value. A dimension we don't know for the row
    being priced (e.g. no customer breakdown in view) can only match a wildcard -- we
    never guess that an unknown value happens to equal a specific override's target.

    Most-specific-wins was the deliberate choice: an item-level override is a more
    considered statement than a blanket channel one, so it shouldn't be silently
    overwritten by the broader rule."""
    if overrides_df is None or overrides_df.empty:
        return None
    known = {"channel": channel, "product": product, "customer": customer}
    best_val, best_score = None, -1
    for _, row in overrides_df.iterrows():
        matched = True
        for dim, actual in known.items():
            target = str(row.get(dim, OVERRIDE_ANY) or OVERRIDE_ANY)
            if target == OVERRIDE_ANY:
                continue  # wildcard matches anything, including unknown
            if actual is None or str(actual) != target:
                matched = False
                break
        if matched:
            score = override_specificity(row)
            if score > best_score:
                best_val, best_score = float(row["override_kg"]), score
    return best_val


def apply_overrides_to_segments(sales_df, segment_forecasts, active_overrides, freq="W"):
    """Folds manual overrides into the per-segment forecasts, so the Overview KPI, the
    segment tables, and the charts all reflect them -- previously overrides only reached the
    per-item forecast table and the detailed breakdown, so the headline number and charts
    quietly ignored them.

    Works on DELTAS rather than replacement. An override states what one slice should be, but
    a segment covers many slices, so we can't just swap the segment's number. Instead: work
    out what that slice was contributing to the segment (its historical share x the segment
    forecast), then shift the segment by the difference between the override and that
    baseline. A 2,000 kg override on a slice currently contributing 1,000 kg raises its
    segment by 1,000 kg -- everything else in the segment is left alone.

    Returns (adjusted_forecasts, notes) where notes describes each adjustment made."""
    if not segment_forecasts or active_overrides is None or active_overrides.empty:
        return dict(segment_forecasts or {}), []

    adjusted = dict(segment_forecasts)
    notes = []
    segments = split_into_segments(sales_df)

    for _, ov in active_overrides.iterrows():
        ov_ch = str(ov.get("channel", OVERRIDE_ANY) or OVERRIDE_ANY)
        ov_pr = str(ov.get("product", OVERRIDE_ANY) or OVERRIDE_ANY)
        ov_cu = str(ov.get("customer", OVERRIDE_ANY) or OVERRIDE_ANY)
        ov_kg = float(ov["override_kg"])

        for label, seg_df in segments.items():
            if label not in adjusted or seg_df.empty:
                continue
            scope = seg_df
            if ov_ch != OVERRIDE_ANY:
                scope = scope[scope["channel"] == ov_ch]
            if ov_pr != OVERRIDE_ANY:
                scope = scope[scope["product"] == ov_pr]
            if ov_cu != OVERRIDE_ANY and "customer" in scope.columns:
                scope = scope[scope["customer"] == ov_cu]
            if scope.empty:
                continue  # this override doesn't touch this segment

            seg_total_kg = seg_df["kg"].sum()
            if seg_total_kg <= 0:
                continue
            share = scope["kg"].sum() / seg_total_kg
            baseline = adjusted[label] * share
            delta = ov_kg - baseline
            adjusted[label] = max(adjusted[label] + delta, 0.0)
            notes.append({
                "scope": " / ".join(p for p in [ov_ch, ov_pr, ov_cu] if p != OVERRIDE_ANY) or "(all)",
                "segment": label,
                "was_contributing_kg": round(baseline, 1),
                "override_kg": round(ov_kg, 1),
                "segment_change_kg": round(delta, 1),
            })
    return adjusted, notes


def load_known_classifications():
    """Loads every product this app has ever classified before -- either automatically or by
    a person -- so past decisions are remembered and never re-asked for the same product.

    Checks two sources: the dedicated memory table (built from the auto-derive workflow), AND
    real historical sales_records directly (covers a real gap -- if an earlier upload had an
    explicit 'Single vs Staple' column, that classification was saved to sales_records but
    never into the dedicated memory table, so a later upload without that column previously
    couldn't benefit from it. This makes both paths count as 'known', regardless of which
    column format a given upload happened to use."""
    known = {}
    hist = pd.read_sql(
        "SELECT DISTINCT product, product_type FROM sales_records "
        "WHERE product_type IS NOT NULL AND product_type NOT IN ('(not tracked)', 'Unknown')", conn)
    for _, row in hist.iterrows():
        known[row["product"]] = row["product_type"]

    df = pd.read_sql("SELECT product_key, classification FROM product_classifications", conn)
    for _, row in df.iterrows():
        known[row["product_key"]] = row["classification"]  # dedicated table wins on conflict -- more deliberate source
    return known


def save_classification(product_key, classification, source):
    """Persists a classification decision (auto-detected or manual) so future uploads of the
    same product reuse it instead of re-classifying or re-asking every time."""
    existing = pd.read_sql("SELECT id FROM product_classifications WHERE product_key = ?",
                            conn, params=(product_key,))
    if not existing.empty:
        conn.execute("DELETE FROM product_classifications WHERE product_key = ?", (product_key,))
    conn.execute("INSERT INTO product_classifications (product_key, classification, source, updated_at) "
                 "VALUES (?,?,?,?)", (product_key, classification, source, datetime.now().isoformat()))
    conn.commit()


def compute_size_mix(df):
    g = df.groupby(["channel", "product", "size_label"], as_index=False)["kg"].sum()
    g["group_total_kg"] = g.groupby(["channel", "product"])["kg"].transform("sum")
    g["size_mix_pct"] = (g["kg"] / g["group_total_kg"] * 100).round(1)
    return g[["channel", "product", "size_label", "size_mix_pct", "kg"]]


def compute_customer_mix(df):
    g = df.groupby(["channel", "product", "customer"], as_index=False)["kg"].sum()
    g["group_total_kg"] = g.groupby(["channel", "product"])["kg"].transform("sum")
    g["customer_mix_pct"] = (g["kg"] / g["group_total_kg"] * 100).round(1)
    return g[["channel", "product", "customer", "customer_mix_pct", "kg"]].sort_values(
        ["channel", "product", "customer_mix_pct"], ascending=[True, True, False])


# ===================================================================
# DEMAND SENSING — trend-based auto forecast from actual history
# ===================================================================
def compute_weekly_actuals(sales_df):
    """Derives real weekly kg by channel/product straight from uploaded transactions."""
    d = sales_df.copy()
    d["record_date"] = pd.to_datetime(d["record_date"], errors="coerce")
    d = d.dropna(subset=["record_date"])
    d["week_start"] = (d["record_date"] - pd.to_timedelta(d["record_date"].dt.weekday, unit="D")).dt.date.astype(str)
    g = d.groupby(["channel", "product", "week_start"], as_index=False)["kg"].sum() \
        .rename(columns={"kg": "actual_kg"})
    return g


def _median_trend_forecast(history_kg, damping=0.6):
    """Fallback method: median of the last 4 weeks (robust to a single outlier spike/dip),
    adjusted by a DAMPED growth rate vs. the prior 4 weeks. Damping (< 1) means a burst of
    recent growth gets partially, not fully, extrapolated forward -- this specifically
    prevents the method from overshooting after a one-off spike week. Used directly when
    there's too little history for ARIMA to be reliable (under 8 points)."""
    vals = np.asarray(history_kg, dtype=float)
    n = len(vals)
    if n < 2:
        return None
    recent = np.median(vals[-min(4, n):])
    if n >= 8:
        prior = np.median(vals[-8:-4])
    elif n > min(4, n):
        prior = np.median(vals[:n - min(4, n)])
    else:
        prior = recent
    growth = 0.0
    if prior > 0:
        growth = (recent - prior) / prior
        growth = max(min(growth, 1.0), -0.5)  # clamp to avoid wild extrapolation
    return float(max(recent * (1 + damping * growth), 0))


def _cap_outliers(vals, percentile=90):
    """Caps extreme spikes in training data before fitting ARIMA/SARIMA -- tested finding:
    on a series with recurring sharp spikes (e.g. an occasional large bulk order against a
    much smaller typical baseline), letting ARIMA fit the raw data including those spikes
    made its forecast noticeably worse for ordinary weeks (tested: 35.0% MAPE uncapped vs
    23.5% MAPE capped, on the same held-out data). The safe-reference sanity check elsewhere
    catches individual wild-looking outputs; this addresses the more common, subtler case
    where a model 'kind of' overreacts to a real historical spike without being wild enough
    to trip that check. Does NOT touch the safe-reference calculation itself, which is
    already median-based and naturally resistant to outliers."""
    if len(vals) < 4:
        return vals
    cap = np.percentile(vals, percentile)
    return np.minimum(vals, cap)


def trend_forecast(history_kg, damping=0.6):
    """Primary forecasting method. Tries ARIMA(1,1,1) when there's enough history (8+ points)
    to fit reliably -- tested against real data to beat the median-trend fallback by ~10-25%
    lower MAPE. Below that threshold, or if ARIMA fails to fit, falls back to the median-trend
    method, since ARIMA is genuinely unreliable on very short series (verified: a 4-point
    series produced a forecast outside the entire historical range). Used for the backtest
    specifically, since it needs hundreds of fast refits -- see trend_forecast_seasonal for
    the live forecast, which can afford a slower, more thorough fit.

    The ARIMA fit itself trains on outlier-capped history (see _cap_outliers) -- tested to
    meaningfully improve accuracy on series with recurring sharp spikes, which is exactly
    what a niche/volatile category (e.g. a smaller product line with occasional large orders)
    tends to look like.

    Sanity-checked against the safe median-fallback value before being trusted -- tested
    finding: a brief spike in a sparse/erratic series (e.g. a niche low-volume item) can make
    ARIMA extrapolate that spike forward aggressively, producing a forecast several times
    higher than what a bounded method would predict, for a week or two, before it corrects
    itself. If ARIMA's result is wildly larger or smaller than the safe reference, the safe
    reference is used instead -- this is what actually caught and fixed a real reported case
    where 'Single' forecasted 2,222 kg against typical actuals of 170-420 kg."""
    vals = np.asarray(history_kg, dtype=float)
    n = len(vals)
    if n < 2:
        return None
    safe_reference = _median_trend_forecast(history_kg, damping=damping)
    if n >= 8:
        try:
            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                from statsmodels.tsa.statespace.sarimax import SARIMAX
                fit_vals = _cap_outliers(vals)
                model = SARIMAX(fit_vals, order=(1, 1, 1), seasonal_order=(0, 0, 0, 0),
                                 enforce_stationarity=False, enforce_invertibility=False)
                fit = model.fit(disp=False)
                f = fit.forecast(1)[0]
                if pd.notna(f) and f >= 0 and _within_sane_bounds(f, safe_reference):
                    return float(f)
        except Exception:
            pass
    return safe_reference


def _within_sane_bounds(candidate, safe_reference, max_ratio=3.0):
    """A candidate forecast is trusted only if it's within max_ratio x of the safe reference
    in either direction -- guards against ARIMA/SARIMA producing a wildly unstable number on
    sparse or erratic data, without discarding genuinely large-but-real growth (up to 3x)."""
    if safe_reference is None or safe_reference <= 0:
        return True  # no meaningful reference to check against (e.g. brand new item) -- trust it
    return (safe_reference / max_ratio) <= candidate <= (safe_reference * max_ratio)


def audit_and_fix_historical_forecasts(weekly_actual, max_ratio=3.0):
    """Re-checks every ALREADY-STORED forecast against what a safe, bounded method would have
    predicted using only the data available at that time, and corrects any that are wildly
    unstable. This is necessary because forecasts are frozen once generated -- fixing the
    forecasting method going forward (the sanity check in trend_forecast) does NOT retroactively
    fix numbers already sitting in the database from before that fix existed. Returns
    (checked_count, fixed_count, examples) so the caller can show a clear summary."""
    all_forecasts = pd.read_sql("SELECT id, channel, product, target_week, forecast_kg FROM auto_forecasts", conn)
    if all_forecasts.empty:
        return 0, 0, []

    checked, fixed, examples = 0, 0, []
    for _, row in all_forecasts.iterrows():
        ch, pr, tw, fid, current_f = row["channel"], row["product"], row["target_week"], row["id"], row["forecast_kg"]
        hist = weekly_actual[(weekly_actual["channel"] == ch) & (weekly_actual["product"] == pr) &
                              (weekly_actual["week_start"] < tw)].sort_values("week_start").tail(MAX_LOOKBACK_WEEKS)
        if len(hist) < 2:
            continue
        checked += 1
        safe_ref = _median_trend_forecast(hist["actual_kg"].tolist())
        if not _within_sane_bounds(current_f, safe_ref, max_ratio=max_ratio):
            new_val = float(round(safe_ref, 1))
            conn.execute("UPDATE auto_forecasts SET forecast_kg = ? WHERE id = ?", (new_val, int(fid)))
            fixed += 1
            if len(examples) < 10:
                examples.append((ch, pr, tw, current_f, new_val))
    conn.commit()
    return checked, fixed, examples


def trend_forecast_seasonal(history_kg, damping=0.6):
    """Used for the LIVE forecast only (one fit per new week, cached) -- NOT the backtest,
    since a seasonal fit takes ~12 seconds vs ~0.01s for the non-seasonal version (tested),
    and the backtest needs hundreds of refits, which would take hours with seasonality on.
    Three tiers: with 2+ years of history (104+ weeks), tries real seasonal SARIMA(1,1,1)x
    (1,1,1,52) to actually capture a yearly cycle. With 8-104 weeks, same non-seasonal ARIMA
    as the backtest uses (with the same sanity check and outlier capping). Below 8 weeks,
    the median-trend fallback.

    NOTE on capping here specifically: if a category's spikes are genuine calendar-seasonal
    (e.g. reliably every December), capping them would work against the whole point of this
    seasonal tier, which exists to learn exactly that kind of recurring pattern. This applies
    the same capping as the non-seasonal method as a reasonable default given spikes that
    don't look tied to specific calendar periods (tested improvement: 35.0% -> 23.5% MAPE on
    a series with irregular, non-calendar-aligned spikes) -- but if you have a category with
    real, predictable calendar seasonality, this percentile can be raised or disabled for it."""
    vals = np.asarray(history_kg, dtype=float)
    n = len(vals)
    if n < 2:
        return None
    safe_reference = _median_trend_forecast(history_kg, damping=damping)
    if n >= 104:
        try:
            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                from statsmodels.tsa.statespace.sarimax import SARIMAX
                fit_vals = _cap_outliers(vals)
                model = SARIMAX(fit_vals, order=(1, 1, 1), seasonal_order=(1, 1, 1, 52),
                                 enforce_stationarity=False, enforce_invertibility=False)
                fit = model.fit(disp=False)
                f = fit.forecast(1)[0]
                if pd.notna(f) and f >= 0 and _within_sane_bounds(f, safe_reference):
                    return float(f)
        except Exception:
            pass
    return trend_forecast(history_kg, damping=damping)


def get_stored_order(label, freq="W"):
    """Reads a previously-searched best model order for a segment, or None.

    Deliberately NEVER runs a search itself -- the search is expensive (measured at 20-40+
    seconds per series) and belongs behind an explicit button, not on a page load. Until
    someone runs it, forecasting uses the ARIMA(1,1,1) default, which is a reasonable
    general-purpose choice; once a search has been run, its result is reused indefinitely."""
    try:
        row = pd.read_sql(
            "SELECT * FROM best_model_cache WHERE product_type = ? AND freq = ? ORDER BY id DESC LIMIT 1",
            conn, params=(label, freq))
        if row.empty:
            return None
        r = row.iloc[0]
        return ((int(r["order_p"]), int(r["order_d"]), int(r["order_q"])),
                (int(r["seasonal_p"]), int(r["seasonal_d"]), int(r["seasonal_q"]), int(r["seasonal_m"])),
                str(r["found_at"])[:10])
    except Exception:
        return None


def search_best_order(label, series, freq="W"):
    """Runs a real auto_arima search for one segment and stores the result. Called ONLY from
    the explicit button, never automatically."""
    order, seasonal_order = (1, 1, 1), (0, 0, 0, 0)
    try:
        from pmdarima import auto_arima
        m = detect_seasonal_period(list(series))
        model = auto_arima(
            np.asarray(series, dtype=float),
            seasonal=bool(m), m=m or 1,
            start_p=1, start_q=0, max_p=3, max_q=3, max_d=2,
            start_P=0, start_Q=0, max_P=1, max_Q=1, max_D=1,
            stepwise=True, suppress_warnings=True, error_action="ignore",
        )
        order, seasonal_order = model.order, model.seasonal_order
    except Exception:
        pass
    conn.execute("DELETE FROM best_model_cache WHERE product_type = ? AND freq = ?", (label, freq))
    conn.execute("""INSERT INTO best_model_cache (product_type, freq, order_p, order_d, order_q,
        seasonal_p, seasonal_d, seasonal_q, seasonal_m, found_at) VALUES (?,?,?,?,?,?,?,?,?,?)""",
        (label, freq, order[0], order[1], order[2], seasonal_order[0], seasonal_order[1],
         seasonal_order[2], seasonal_order[3], datetime.now().isoformat()))
    conn.commit()
    return order, seasonal_order


def find_best_order_cached(pt, series, freq, refit_days=SEASONAL_REFIT_DAYS):
    """Finds the best-fitting SARIMA order for a product type's own aggregated series, using
    a real auto_arima search -- same validated constraints from earlier testing: start_p=1
    (avoids a degenerate flat-forecast model, a real bug we found and fixed), max_P/max_Q=1
    (keeps the search from wandering into individual candidates that take 5-10+ seconds each).
    Caches the FOUND ORDER, not just one forecast value, so this expensive search (tested:
    16 seconds to a few minutes) only runs periodically -- same two-cadence philosophy as
    the rest of the app's forecasting, now practical here since it's only 1-2 searches total
    (one per product type), not one per item."""
    cached = pd.read_sql(
        "SELECT * FROM best_model_cache WHERE product_type = ? AND freq = ? ORDER BY id DESC LIMIT 1",
        conn, params=(pt, freq))
    cutoff = (datetime.now() - timedelta(days=refit_days)).isoformat()
    if not cached.empty and cached.iloc[0]["found_at"] >= cutoff:
        row = cached.iloc[0]
        return ((int(row["order_p"]), int(row["order_d"]), int(row["order_q"])),
                (int(row["seasonal_p"]), int(row["seasonal_d"]), int(row["seasonal_q"]), int(row["seasonal_m"])))

    order, seasonal_order = (1, 1, 1), (0, 0, 0, 0)
    try:
        from pmdarima import auto_arima
        use_seasonal = len(series) >= 104 and freq == "W"
        model = auto_arima(
            np.asarray(series, dtype=float),
            seasonal=use_seasonal, m=52 if use_seasonal else 1,
            start_p=1, start_q=0, max_p=3, max_q=3, max_d=2,
            start_P=0, start_Q=0, max_P=1, max_Q=1, max_D=1,
            stepwise=True, suppress_warnings=True, error_action="ignore",
        )
        order, seasonal_order = model.order, model.seasonal_order
    except Exception:
        pass  # fall through to the safe default order set above

    conn.execute("""INSERT INTO best_model_cache (product_type, freq, order_p, order_d, order_q,
        seasonal_p, seasonal_d, seasonal_q, seasonal_m, found_at) VALUES (?,?,?,?,?,?,?,?,?,?)""",
        (pt, freq, order[0], order[1], order[2], seasonal_order[0], seasonal_order[1],
         seasonal_order[2], seasonal_order[3], datetime.now().isoformat()))
    conn.commit()
    return order, seasonal_order


def fit_with_found_order(series, order, seasonal_order, n_periods=1):
    """Fits SARIMAX with a specific, already-found (order, seasonal_order) and returns a
    forecast + 80% confidence range for n_periods ahead. Sanity-checked and outlier-capped
    the same way as the rest of the app's forecasting, falling back to the safe damped-median
    projection if the fit fails or produces an unstable first-period result."""
    vals = np.asarray(series, dtype=float)
    safe_reference = _median_trend_forecast(series)
    try:
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            from statsmodels.tsa.statespace.sarimax import SARIMAX
            fit_vals = _cap_outliers(vals)
            model = SARIMAX(fit_vals, order=order, seasonal_order=seasonal_order,
                             enforce_stationarity=False, enforce_invertibility=False)
            fit = model.fit(disp=False)
            result = fit.get_forecast(steps=n_periods)
            mean = result.predicted_mean
            ci = result.conf_int(alpha=0.20)
            f1 = float(mean[0])
            if f1 == f1 and f1 >= 0 and _within_sane_bounds(f1, safe_reference):
                path = []
                for h in range(n_periods):
                    f = max(float(mean[h]), 0)
                    low = max(float(ci[h][0]), 0)
                    high = max(float(ci[h][1]), f)
                    path.append({"step": h + 1, "forecast_kg": f, "low": low, "high": high})
                return pd.DataFrame(path)
    except Exception:
        pass
    hist_vals = list(series)[-8:]
    path = []
    for h in range(1, n_periods + 1):
        f = _median_trend_forecast(hist_vals[-8:], damping=0.6 * (0.7 ** (h - 1)))
        if f is None:
            break
        band = 0.15 * np.sqrt(h)
        path.append({"step": h, "forecast_kg": f, "low": max(f * (1 - band), 0), "high": f * (1 + band)})
        hist_vals.append(f)
    return pd.DataFrame(path)


def _cheap_data_fingerprint(sales_df):
    """A fast surrogate for cache-keying on sales_df, instead of hashing the entire
    DataFrame's contents. Streamlit's default caching hashes the full DataFrame on every
    call to check if it's changed -- that cost grows with data size and, after months of
    real testing, was very likely a real contributor to the reported slowdown. Row count +
    latest date is enough to correctly detect "the data changed" for our purposes, at a
    fraction of the cost."""
    if sales_df.empty:
        return (0, None)
    latest = sales_df["record_date"].max() if "record_date" in sales_df.columns else None
    return (len(sales_df), str(latest))


MAJOR_STAPLE_CHANNEL = "Specialty Retail"


def split_into_segments(sales_df, major_channel=MAJOR_STAPLE_CHANNEL):
    """Splits sales into the THREE groups we forecast independently:
      1. Single
      2. Staple — Specialty Retail
      3. Staple — other channels

    Specialty Retail is separated out because it's grown to a large share of Staple and is
    still actively shifting -- large and fast-moving enough that estimating it as a
    proportion of Staple would keep lagging its real trajectory. The remaining channels are
    more stable, so they're forecast together as one group rather than each getting their
    own noisy model. Returns an ordered dict of {label: dataframe}."""
    segments = {}
    if "product_type" not in sales_df.columns:
        return segments
    single_df = sales_df[sales_df["product_type"] == "Single"]
    staple_df = sales_df[sales_df["product_type"] == "Staple"]
    if not single_df.empty:
        segments["Single"] = single_df
    if not staple_df.empty and "channel" in staple_df.columns:
        # match the channel case-insensitively and ignoring stray whitespace -- an exact-match
        # miss here would silently collapse this back to two segments, which is confusing and
        # hard to spot
        chan_norm = staple_df["channel"].astype(str).str.strip().str.casefold()
        target = str(major_channel).strip().casefold()
        is_major = chan_norm == target
        sr_df = staple_df[is_major]
        rest_df = staple_df[~is_major]
        if not sr_df.empty:
            segments[f"Staple — {major_channel}"] = sr_df
        if not rest_df.empty:
            segments["Staple — other channels"] = rest_df
    elif not staple_df.empty:
        segments["Staple"] = staple_df
    return segments


@st.cache_data(ttl=900, max_entries=8, show_spinner="Forecasting segments...",
                hash_funcs={pd.DataFrame: _cheap_data_fingerprint})
def compute_segment_forecast(sales_df, freq="W"):
    """One real, independent forecast per segment -- each fit on that segment's OWN
    aggregated history. Nothing here is derived by splitting a total: the segments are
    forecast separately, and their sum becomes the company total, not the other way
    around. Returns {segment_label: forecast_kg} for the next single period."""
    results = {}
    for label, seg_df in split_into_segments(sales_df).items():
        agg = aggregate_periods(seg_df, ["product_type"], freq)
        series = agg.groupby("period", as_index=False)["actual_kg"].sum().sort_values("period")["actual_kg"].tolist()
        if len(series) >= 2:
            f = trend_forecast(series)
            if f is not None:
                results[label] = f
    return results


@st.cache_data(ttl=900, max_entries=8, hash_funcs={pd.DataFrame: _cheap_data_fingerprint})
def compute_rates_by(df, group_cols, recent_days=45):
    """Weighted price per kg at ANY grouping — by bag size alone, by channel alone, by
    customer, or any combination. The fixed channel x item x size table couldn't answer
    "what does a 12oz bag average across the whole business?", because it always split by
    all three at once. Same weighted method as everywhere else: total revenue / total kg,
    so a large order counts proportionally more than a small one."""
    group_cols = [g for g in group_cols if g in df.columns]
    if not group_cols or df.empty:
        return pd.DataFrame()
    d = df.copy()
    d["record_date"] = pd.to_datetime(d["record_date"], errors="coerce")
    if d["record_date"].notna().any():
        cutoff = d["record_date"].max() - pd.Timedelta(days=recent_days)
        recent = d[d["record_date"] >= cutoff]
        d = recent if len(recent) >= 20 else d
    g = d.groupby(group_cols, as_index=False, observed=True).agg(
        total_kg=("kg", "sum"), total_revenue=("revenue", "sum"), lines=("kg", "size"))
    g["$ per kg"] = (g["total_revenue"] / g["total_kg"].replace(0, np.nan)).round(2)
    g["kg per $1 CAD"] = (1 / g["$ per kg"].replace(0, np.nan)).round(4)
    g["total_kg"] = g["total_kg"].round(0)
    return g.sort_values("total_kg", ascending=False)


@st.cache_data(ttl=900, max_entries=8, show_spinner="Forecasting by product type...", hash_funcs={pd.DataFrame: _cheap_data_fingerprint})
def compute_type_level_forecast(sales_df, freq="W"):
    """Real, top-down forecasting: forecasts Staple and Single directly, each as its own
    aggregated series, rather than deriving them by summing many small per-item forecasts.
    This matches how real demand planning handles the accuracy-vs-detail tradeoff --
    forecast at the most meaningful aggregate level (most statistically reliable, since
    noise cancels out over more data), then split that DOWN to finer detail (channel/item/
    bag size) by historical proportion, rather than the reverse.

    Uses a fast, fixed-order fit rather than a full best-model search. Real reasoning, not
    a corner cut under pressure: the search step was directly measured at 20-40+ seconds
    PER product type for a realistic seasonal pattern, and it wasn't reliably staying
    confined to its intended once-every-28-days cadence -- a real, reported case ended up
    "taking hours" and re-running on every single page open. Removing the search from the
    automatic path guarantees a fast, predictable runtime every time, which matters more
    here than a marginal accuracy gain from exhaustively searching every model shape.
    Returns {product_type: forecast_kg} for the next single period."""
    if "product_type" not in sales_df.columns:
        return {}
    results = {}
    for pt in sales_df["product_type"].dropna().unique():
        if pt == "(not tracked)":
            continue
        pt_df = sales_df[sales_df["product_type"] == pt]
        agg = aggregate_periods(pt_df, ["product_type"], freq)
        series = agg.sort_values("period")["actual_kg"].tolist()
        if len(series) >= 2:
            f = trend_forecast(series)
            if f is not None:
                results[pt] = f
    return results


@st.cache_data(ttl=900, max_entries=4, show_spinner="Working out bag counts...",
                hash_funcs={pd.DataFrame: _cheap_data_fingerprint})
def compute_all_channel_bag_breakdown(sales_df, n_periods=3, freq="M", segment_adjust=None):
    """Bag-size breakdown across EVERY segment -- Single as well as both Staple segments --
    so Operations sees one complete ordering picture instead of Staple only.

    Each segment is projected on its own history, then split by trending channel share and
    trending bag-size share within each channel, and finally converted to bag counts using
    the real kg-per-bag rate learned from sales. Returns channel, size_label, period,
    forecast_kg, forecast_bags, segment."""
    if "size_label" not in sales_df.columns:
        return pd.DataFrame(columns=["segment", "channel", "size_label", "period", "forecast_kg", "forecast_bags"])

    rows = []
    for label, seg_df in split_into_segments(sales_df).items():
        agg = aggregate_periods(seg_df, ["product_type"], freq)
        agg = agg.groupby("period", as_index=False)["actual_kg"].sum().sort_values("period")
        series = agg["actual_kg"].tolist()
        if len(series) < 2:
            continue
        proj = project_forward_with_range(series, None, n_periods=n_periods, keep_trend=True)
        if proj.empty:
            continue
        ch_shares = compute_trending_shares(seg_df, ["channel"], freq=freq)
        size_by_ch = {ch: compute_trending_shares(seg_df[seg_df["channel"] == ch], ["size_label"], freq=freq)
                      for ch in seg_df["channel"].dropna().unique()}
        last_date = pd.Timestamp(agg["period"].iloc[-1])
        for h in range(n_periods):
            if h >= len(proj):
                break
            period_label = (last_date + pd.DateOffset(months=h + 1)).date().isoformat() if freq == "M" \
                else (last_date + pd.Timedelta(weeks=h + 1)).date().isoformat()
            seg_total = proj["forecast_kg"].iloc[h]
            # fold in this segment's share of pipeline events and manual overrides. Without
            # this the bag order -- the number Operations actually buys against -- would
            # ignore a signed contract or a deliberate override entirely, which is the one
            # place that error costs real money.
            if segment_adjust:
                seg_total = max(seg_total + float(segment_adjust.get(label, 0.0)), 0.0)
            for _, chrow in ch_shares.iterrows():
                ch_kg = seg_total * chrow["share"]
                for _, srow in size_by_ch.get(chrow["channel"], pd.DataFrame()).iterrows():
                    rows.append({"segment": label, "channel": chrow["channel"],
                                 "size_label": srow["size_label"], "period": period_label,
                                 "forecast_kg": ch_kg * srow["share"]})
    result = pd.DataFrame(rows)
    if result.empty:
        return result

    kg_per_bag = compute_kg_per_bag(sales_df)
    if not kg_per_bag.empty:
        result = result.merge(kg_per_bag, on="size_label", how="left")
        _kg = pd.to_numeric(result["forecast_kg"], errors="coerce").astype("float64")
        _rate = pd.to_numeric(result["kg_per_bag"], errors="coerce").astype("float64").where(lambda s: s > 0)
        result["forecast_bags"] = np.ceil(_kg.div(_rate))
    else:
        result["kg_per_bag"] = np.nan
        result["forecast_bags"] = np.nan
    return result


@st.cache_data(ttl=900, max_entries=4, show_spinner="Forecasting Staple by channel and bag size (first run can take a moment)...",
                hash_funcs={pd.DataFrame: _cheap_data_fingerprint})
def compute_staple_channel_breakdown(sales_df, n_periods=13, freq="W", major_channel="Specialty Retail"):
    """Three-tier forecast for Staple specifically, for Operations' bag-ordering use case
    (they need a ~3 month lead time on packaging, by size, per channel). Real reasoning
    behind the structure, not a uniform rule applied everywhere:

    1. Staple overall gets its own direct forecast (already the most reliable level).
    2. Specialty Retail gets ITS OWN direct forecast too, not a proportional split -- real
       numbers behind this: it's grown to ~43% of total production and ~98% of that is
       Staple, meaning it's roughly half of the entire Staple business on its own and still
       actively shifting. A proportional split of something this large and this fast-moving
       would lag its real trajectory the same way a stale average lagged it before. Direct
       forecast + reconciliation avoids that.
    3. Every other channel splits the REMAINDER (Staple minus Specialty Retail) using
       trending shares -- appropriate for smaller, more stable channels, avoiding the noise
       of giving every small channel its own from-scratch model.
    4. Within each channel, bag size is split the same way (trending shares), since bag
       size mix within a channel tends to be a smaller, more gradual shift than channel mix
       itself -- no evidence yet that any one size needs its own direct forecast the way
       Specialty Retail did.

    Returns a DataFrame: channel, size_label, period, forecast_kg."""
    staple_df = sales_df[sales_df["product_type"] == "Staple"] if "product_type" in sales_df.columns else pd.DataFrame()
    if staple_df.empty:
        return pd.DataFrame(columns=["channel", "size_label", "period", "forecast_kg"])

    staple_agg = aggregate_periods(staple_df, ["product_type"], freq)
    staple_series = staple_agg.sort_values("period")["actual_kg"].tolist()
    if len(staple_series) < 8:
        return pd.DataFrame(columns=["channel", "size_label", "period", "forecast_kg"])
    # fast, non-searching, non-seasonal projection -- same reasoning as the main dashboard
    # fix: the best-model search was directly measured at 20-40+ seconds per series and
    # wasn't reliably staying confined to its intended cadence, contributing to a real
    # reported case of the app taking hours. This trades some accuracy for a guaranteed,
    # predictable runtime, which matters more here.
    staple_projection = project_forward_with_range(staple_series, None, n_periods=n_periods)
    if staple_projection.empty:
        return pd.DataFrame(columns=["channel", "size_label", "period", "forecast_kg"])

    # THREE independent forecasts, not two-plus-a-subtraction:
    #   1. Specialty Retail (its own model)
    #   2. The rest of Staple (its own model)
    #   3. Staple overall (its own model -- used only as a reconciliation anchor)
    # Rationale for the change: SR and the rest of Staple can behave genuinely differently
    # (different trend, different seasonality), so giving the remainder its own model rather
    # than deriving it by subtraction lets it follow its own real shape.
    #
    # Honest tradeoff, worth understanding: independent forecasts have no mathematical reason
    # to sum to the Staple total. Aggregate forecasts are usually MORE accurate than the sum
    # of their parts (noise cancels out over more data), so we still forecast Staple overall
    # and scale the two parts proportionally to match it. That keeps the parts following
    # their own real trajectories while preserving the reconciliation guarantee the rest of
    # the app depends on.
    sr_df = staple_df[staple_df["channel"] == major_channel]
    sr_agg = aggregate_periods(sr_df, ["channel"], freq)
    sr_series = sr_agg.sort_values("period")["actual_kg"].tolist()
    sr_projection = project_forward_with_range(sr_series, None, n_periods=n_periods) if len(sr_series) >= 2 else pd.DataFrame()

    other_channels_df = staple_df[staple_df["channel"] != major_channel]
    rest_agg = aggregate_periods(other_channels_df, ["product_type"], freq) if not other_channels_df.empty else pd.DataFrame()
    rest_series = rest_agg.sort_values("period")["actual_kg"].tolist() if not rest_agg.empty else []
    rest_projection = project_forward_with_range(rest_series, None, n_periods=n_periods) if len(rest_series) >= 2 else pd.DataFrame()

    channel_shares = compute_trending_shares(other_channels_df, ["channel"], freq=freq) if not other_channels_df.empty else pd.DataFrame()
    size_shares_by_channel = {}
    for ch in staple_df["channel"].dropna().unique():
        ch_df = staple_df[staple_df["channel"] == ch]
        size_shares_by_channel[ch] = compute_trending_shares(ch_df, ["size_label"], freq=freq)

    rows = []
    last_date = pd.Timestamp(staple_agg["period"].iloc[-1])
    for h in range(n_periods):
        step = h + 1
        period_label = (last_date + pd.Timedelta(weeks=step)).date().isoformat() if freq == "W" \
            else (last_date + pd.DateOffset(months=step)).date().isoformat()
        staple_total_h = staple_projection["forecast_kg"].iloc[h] if h < len(staple_projection) else None
        if staple_total_h is None:
            continue

        sr_raw = sr_projection["forecast_kg"].iloc[h] if not sr_projection.empty and h < len(sr_projection) else 0
        rest_raw = rest_projection["forecast_kg"].iloc[h] if not rest_projection.empty and h < len(rest_projection) else 0

        # reconcile the two independent parts to the (more reliable) aggregate Staple total
        parts_sum = sr_raw + rest_raw
        if parts_sum > 0:
            scale = staple_total_h / parts_sum
            sr_h, remainder_h = sr_raw * scale, rest_raw * scale
        else:
            sr_h, remainder_h = 0, staple_total_h

        for _, row in channel_shares.iterrows():
            ch_kg = remainder_h * row["share"]
            sizes = size_shares_by_channel.get(row["channel"], pd.DataFrame())
            for _, srow in sizes.iterrows():
                rows.append({"channel": row["channel"], "size_label": srow["size_label"],
                             "period": period_label, "forecast_kg": ch_kg * srow["share"]})

        sr_sizes = size_shares_by_channel.get(major_channel, pd.DataFrame())
        for _, srow in sr_sizes.iterrows():
            rows.append({"channel": major_channel, "size_label": srow["size_label"],
                         "period": period_label, "forecast_kg": sr_h * srow["share"]})

    result = pd.DataFrame(rows)
    if result.empty:
        return result

    # convert kg -> real bag counts, using an actual weighted kg-per-bag rate learned from
    # your own data (not guessed from the size label text) -- this is the number Operations
    # actually needs to place a bag order, kg alone doesn't tell them how many bags to buy
    kg_per_bag = compute_kg_per_bag(sales_df)
    if not kg_per_bag.empty:
        result = result.merge(kg_per_bag, on="size_label", how="left")
        result["forecast_bags"] = (result["forecast_kg"] / result["kg_per_bag"]).round(0)
    else:
        result["kg_per_bag"] = np.nan
        result["forecast_bags"] = np.nan

    return result


def generate_missing_forecasts(weekly_actual):
    """Freezes a forecast for the next unforecasted week, per channel/product -- called
    every time new data is uploaded (and safely re-checked on every run, idempotent).

    Uses a two-cadence approach, matching how real demand-sensing systems actually work:
    a cheap forecast refresh every week (fast, non-seasonal ARIMA, ~0.01s per combo), and
    a full seasonal SARIMA refit only periodically (SEASONAL_REFIT_DAYS apart, ~12s per
    combo) to keep genuine yearly-seasonality awareness current without paying that cost
    on every single upload. Without this split, every weekly upload would force a fresh
    seasonal fit for every combo -- tested at ~12s each, ~4 minutes total with ~20 combos,
    every single week, forever, once there's 2+ years of history. That doesn't match real
    practice either: production forecasting systems separate a cheap frequent refresh from
    an expensive periodic full retrain, not redo the expensive step on every data point."""
    if weekly_actual.empty:
        return
    known_weeks = sorted(weekly_actual["week_start"].unique())
    latest_week = known_weeks[-1]
    target_week = (pd.Timestamp(latest_week) + pd.Timedelta(days=7)).date().isoformat()
    combos = weekly_actual[["channel", "product"]].drop_duplicates()

    # figure out how many combos actually need fresh computation, before showing any UI,
    # so we don't show a progress bar at all when everything's already cached in the DB.
    # Real bug found via production logs: this used to check ONE combo at a time with a
    # separate database round trip each -- fine with a handful of combos, but after months
    # of real testing (many item/channel combinations accumulated), that became hundreds of
    # sequential round trips and was the actual cause of a reported near-hour-long hang.
    # One bulk query + an in-memory comparison replaces all of that with a single round trip.
    existing_for_week = pd.read_sql("SELECT channel, product FROM auto_forecasts WHERE target_week = ?",
                                     conn, params=(target_week,))
    existing_set = set(zip(existing_for_week["channel"], existing_for_week["product"])) if not existing_for_week.empty else set()
    to_compute = [(row["channel"], row["product"]) for _, row in combos.iterrows()
                  if (row["channel"], row["product"]) not in existing_set]

    if not to_compute:
        return

    # per-item forecasting always uses the FAST method now, never the seasonal one -- real
    # finding: with 508 real combos, even a small fraction hitting the seasonal path (tested
    # at 16+ seconds EACH, even with a known model order) plausibly explained 7-20+ minutes
    # of total runtime, matching a reported "won't even load" case exactly. Seasonality is
    # still modeled properly -- just at the Staple/Single aggregate level, where it's better
    # statistically justified anyway (more data, less noise) and already has its own real,
    # properly-cached best-model search. Per-item forecasts don't need their own seasonal
    # fit to stay useful for the item-level table, pipeline events, and overrides.
    # Collect every forecast in memory first, then write them ALL in one batched insert --
    # real cause of a reported multi-minute hang found in production logs: this used to do
    # one separate INSERT per combo, inside the loop. With 633 combos that's 633 individual
    # network round trips to a remote Supabase database, one at a time, each costing real
    # latency regardless of how fast the model fit itself is. The model fitting was never
    # the bottleneck here -- the network was. insert_dataframe batches 300 rows per
    # statement, turning ~633 round trips into ~3.
    progress_bar = st.progress(0, text=f"Generating forecasts for the new week ({len(to_compute)} to compute)...")
    skipped_combos = []
    pending_rows = []
    generated_at = datetime.now().isoformat()
    # update the progress bar ~20 times total, not once per combo -- each update forces a
    # UI round trip to the browser, which is itself a real cost across hundreds of items
    progress_every = max(1, len(to_compute) // 20)
    for i, (ch, pr) in enumerate(to_compute):
        if i % progress_every == 0:
            progress_bar.progress(i / len(to_compute), text=f"Forecasting… ({i+1} of {len(to_compute)})")
        try:
            hist = weekly_actual[(weekly_actual["channel"] == ch) & (weekly_actual["product"] == pr)] \
                .sort_values("week_start").tail(MAX_LOOKBACK_WEEKS)

            f = trend_forecast(hist["actual_kg"].tolist())
            method_used = "trend_fast"

            # a numerically unstable fit (rare, but real -- e.g. a very sparse or erratic combo)
            # can produce NaN/Inf or an absurdly large number that no sane forecast should be.
            # Validate before inserting rather than let one bad combo crash forecasting for
            # every other combo in this same batch.
            if f is not None and np.isfinite(f) and 0 <= f <= 1_000_000:
                pending_rows.append({
                    "generated_at": generated_at, "channel": ch, "product": pr,
                    "target_week": target_week, "forecast_kg": float(round(f, 1)),
                    "method": method_used,
                })
            elif f is not None:
                skipped_combos.append((ch, pr, f))
        except Exception as e:
            skipped_combos.append((ch, pr, f"error: {e}"))

    if pending_rows:
        progress_bar.progress(0.95, text=f"Saving {len(pending_rows)} forecasts...")
        insert_dataframe("auto_forecasts", pd.DataFrame(pending_rows))
    progress_bar.empty()
    conn.commit()
    if skipped_combos:
        st.warning(f"{len(skipped_combos)} combo(s) produced an unreliable forecast and were skipped this "
                   f"round (they'll be retried next upload): " +
                   ", ".join(f"{ch}/{pr}" for ch, pr, _ in skipped_combos[:5]) +
                   (f" and {len(skipped_combos)-5} more" if len(skipped_combos) > 5 else ""))


@st.cache_data(ttl=900, max_entries=4, show_spinner="Running backtest...")
def backtest_accuracy(weekly_actual, group_cols=("channel", "product"), lookback=LOOKBACK_WEEKS,
                       max_backtest_weeks=12):
    """Generic walk-forward backtest for any grouping (channel+product, channel, product, or customer).

    IMPORTANT -- only backtests the most recent `max_backtest_weeks` weeks per segment, not
    the entire history. Real measurement behind this: with ~633 segments and ~150 weeks of
    history, backtesting everything meant ~94,000 individual model fits in a single call --
    roughly 37+ minutes, matching a real reported case of this running for an hour. And this
    function runs automatically on page load, not behind a button, so that cost was hit on
    every cold start. Capping at 12 recent weeks cuts it by ~92% while still giving a
    meaningful, current accuracy read -- accuracy from 2+ years ago isn't what anyone's
    actually judging the forecast on anyway."""
    group_cols = list(group_cols)
    if weekly_actual.empty:
        return pd.DataFrame()
    rows = []
    for key, grp in weekly_actual.groupby(group_cols, observed=True):
        grp = grp.sort_values("week_start").reset_index(drop=True)
        start_i = max(2, len(grp) - max_backtest_weeks)
        for i in range(start_i, len(grp)):
            hist = grp.iloc[max(0, i - lookback):i]["actual_kg"].tolist()
            f = trend_forecast(hist)
            if f is None:
                continue
            row = {}
            key_tuple = key if isinstance(key, tuple) else (key,)
            for c, k in zip(group_cols, key_tuple):
                row[c] = k
            row.update({"week_start": grp.iloc[i]["week_start"], "forecast_kg": f,
                        "actual_kg": grp.iloc[i]["actual_kg"], "n_weeks_history": len(hist)})
            rows.append(row)
    bt = pd.DataFrame(rows)
    if bt.empty:
        return bt
    bt["variance_pct"] = (bt["actual_kg"] - bt["forecast_kg"]) / bt["forecast_kg"].replace(0, np.nan)
    return bt


def compute_weekly_actuals_by(sales_df, group_cols):
    """Same as compute_weekly_actuals but for any grouping columns (e.g. just ['customer'])."""
    d = sales_df.copy()
    d["record_date"] = pd.to_datetime(d["record_date"], errors="coerce")
    d = d.dropna(subset=["record_date"])
    d["week_start"] = (d["record_date"] - pd.to_timedelta(d["record_date"].dt.weekday, unit="D")).dt.date.astype(str)
    g = d.groupby(list(group_cols) + ["week_start"], as_index=False)["kg"].sum().rename(columns={"kg": "actual_kg"})
    return g


def aggregate_periods(sales_df, group_cols, freq, drop_incomplete=True):
    """Groups actual kg into weekly ('W') or monthly ('M') buckets, for any dimensions.

    Drops the final MONTH when it's incomplete. Real bug this fixes: with only one week of
    August uploaded, the monthly aggregate read 6,180 kg against ~30,000 kg months either
    side. The model has no way to know that's a partial month rather than demand collapsing
    83%, so it forecast the rest of the year from a false floor -- the projection swung from
    -2.3M to +7.5M kg.

    Weekly aggregation is intentionally untouched: a part-week is a minor distortion the
    weekly model already copes with, and excluding it changed weekly numbers that were
    correct."""
    d = sales_df.copy()
    d["record_date"] = pd.to_datetime(d["record_date"], errors="coerce")
    d = d.dropna(subset=["record_date"])
    if d.empty:
        return pd.DataFrame(columns=list(group_cols) + ["period", "actual_kg"])
    if freq == "W":
        d["period"] = (d["record_date"] - pd.to_timedelta(d["record_date"].dt.weekday, unit="D")).dt.date.astype(str)
    else:
        d["period"] = d["record_date"].dt.to_period("M").astype(str)
    g = d.groupby(list(group_cols) + ["period"], as_index=False)["kg"].sum().rename(columns={"kg": "actual_kg"})

    if drop_incomplete and freq == "M" and not g.empty:
        _last_day = d["record_date"].max()
        if freq == "M":
            # a month is complete only if data reaches its final day
            _month_end = (_last_day.to_period("M").to_timestamp("M"))
            if _last_day < _month_end:
                g = g[g["period"] != str(_last_day.to_period("M"))]
        # Weekly is deliberately left alone. A part-week is a small distortion the weekly
        # model already handles, and dropping it changed weekly numbers that were correct --
        # the problem being solved here is specifically the monthly view, where one week
        # standing in for a whole month is an 80%+ distortion.
    return g


@st.cache_data(ttl=900, max_entries=8, show_spinner="Computing forecast...")
def forecast_next_period(agg_df, group_cols, min_history=2):
    """One step ahead, for any grouping -- same trend method, applied to whatever period
    (week or month) the input was aggregated to. Cached for the same reason as backtest_accuracy."""
    if agg_df.empty:
        return pd.DataFrame()
    rows = []
    for key, grp in agg_df.groupby(group_cols):
        grp = grp.sort_values("period")
        vals = grp["actual_kg"].tolist()
        if len(vals) < min_history:
            continue
        f = trend_forecast(vals)
        if f is None:
            continue
        row = {}
        key_tuple = key if isinstance(key, tuple) else (key,)
        for c, k in zip(group_cols, key_tuple):
            row[c] = k
        row["forecast_kg"] = round(f, 1)
        row["n_periods_history"] = len(vals)
        rows.append(row)
    return pd.DataFrame(rows)


def compute_shares(sales_df, group_cols, recent_days=120):
    """Historical share of kg for any grouping, from the last `recent_days` (falls back to
    full history if too little recent data). Used to split a single trustworthy total down
    into any breakdown -- guaranteed to sum back to that total exactly, unlike forecasting
    each breakdown independently."""
    d = sales_df.copy()
    d["record_date"] = pd.to_datetime(d["record_date"], errors="coerce")
    d = d.dropna(subset=["record_date"])
    if d["record_date"].notna().any():
        cutoff = d["record_date"].max() - pd.Timedelta(days=recent_days)
        recent = d[d["record_date"] >= cutoff]
        d = recent if len(recent) >= 20 else d
    total_kg = d["kg"].sum()
    g = d.groupby(group_cols, as_index=False)["kg"].sum()
    g["share"] = g["kg"] / total_kg if total_kg > 0 else 0
    return g[list(group_cols) + ["share"]]


@st.cache_data(ttl=900, max_entries=16, hash_funcs={pd.DataFrame: _cheap_data_fingerprint})
def walk_forward_segment(segment_df, n_periods=12, freq="W"):
    """Walk-forward historical forecast for an ALREADY-FILTERED segment dataframe.

    Takes the segment's own rows directly rather than a product_type label. Real bug this
    fixes: when the dashboard moved from two product types to three segments, the labels
    became things like "Staple — Specialty Retail", which matched no actual product_type
    value -- so both Staple segments silently showed blank forecasts while Single (whose
    label happens to equal a real product_type) still worked.

    For each of the last n_periods, forecasts that period using ONLY the data from before
    it -- no peeking at the answer. Returns a DataFrame with columns: period, forecast_kg."""
    if segment_df is None or segment_df.empty:
        return pd.DataFrame(columns=["period", "forecast_kg"])
    agg = aggregate_periods(segment_df, ["product_type"], freq)
    agg = agg.groupby("period", as_index=False)["actual_kg"].sum().sort_values("period").reset_index(drop=True)
    rows = []
    for i in range(max(len(agg) - n_periods, 2), len(agg)):
        history_before = agg["actual_kg"].iloc[:i].tolist()
        f = trend_forecast(history_before)
        if f is not None:
            rows.append({"period": agg["period"].iloc[i], "forecast_kg": f})
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=["period", "forecast_kg"])


@st.cache_data(ttl=900, max_entries=16, hash_funcs={pd.DataFrame: _cheap_data_fingerprint})
def event_adjustment_by_period(all_events, periods, freq="W"):
    """Weekly (or monthly) kg adjustment from pipeline events, per period, honouring each
    event's own start date. Used so the historical forecast line and the accuracy number
    reflect what the app ACTUALLY predicted at the time -- including any logged event --
    rather than a pure statistical number nobody ever saw. Grading a forecast that was never
    shown makes the accuracy figure meaningless for judging real decisions."""
    out = {p: 0.0 for p in periods}
    if all_events is None or all_events.empty:
        return out
    per = 4.345 if freq == "W" else 1.0  # events are stored monthly
    for _, ev in all_events.iterrows():
        start = str(ev.get("starting_cycle") or "")
        if not start:
            continue
        kg = float(ev.get("expected_kg_per_month") or 0) / per
        ongoing = int(ev.get("ongoing") or 0) == 1
        # an event that was later stopped still counted for the weeks it was live. Honouring
        # that window is what stops turning an event off from retroactively rewriting past
        # forecasts and past accuracy -- the record stays as it was actually reported.
        stopped = str(ev.get("deactivated_at") or "")[:7] if ev.get("deactivated_at") else None
        for p in periods:
            p_month = str(p)[:7]
            if p_month < start:
                continue
            if stopped and p_month > stopped:
                continue
            if not ongoing and p_month != start:
                continue
            out[p] += kg
    return out


def walk_forward_all_segments(sales_df, n_periods=12, freq="W"):
    """Whole-company walk-forward history, built by summing the SAME three per-segment
    calculations the segment tables show. Real bug this fixes: the Overview chart was still
    using the older two-group (Single + Staple) walk-forward while the segment tables had
    moved to three groups -- two different calculations for the same thing, which is exactly
    how the chart ended up showing roughly double what the segments summed to. Routing both
    through this one function makes them agree by construction, not by coincidence."""
    frames = []
    for label, seg_df in split_into_segments(sales_df).items():
        wf = walk_forward_segment(seg_df, n_periods=n_periods, freq=freq)
        if not wf.empty:
            frames.append(wf)
    if not frames:
        return pd.DataFrame(columns=["period", "forecast_kg"])
    return pd.concat(frames, ignore_index=True).groupby("period", as_index=False)["forecast_kg"].sum()


@st.cache_data(ttl=900, max_entries=16, hash_funcs={pd.DataFrame: _cheap_data_fingerprint})
def walk_forward_topdown(sales_df, n_periods=12, freq="W", product_type=None):
    """The ONE walk-forward historical-forecast calculation, shared by both the Overview
    chart and the Staple/Single table -- fixes a real inconsistency: the Overview chart used
    to sum many small per-item walk-forward forecasts (bottom-up), while the Staple/Single
    table used a direct aggregate walk-forward fit per type (top-down). Same class of bug
    found and fixed before with stale frozen forecasts, just resurfacing here as two
    independently-written methods that could show different numbers for the same week.
    Having every caller use this same function, rather than each writing its own version,
    is what actually guarantees they can't drift apart again.

    product_type=None (default): sums across every type -- for the whole-company Overview.
    product_type="Staple" (or "Single"): that type alone -- for the Staple/Single table.

    Returns a DataFrame with columns: period, forecast_kg."""
    if "product_type" not in sales_df.columns:
        return pd.DataFrame(columns=["period", "forecast_kg"])
    types_to_run = [product_type] if product_type else \
        [pt for pt in sales_df["product_type"].dropna().unique() if pt != "(not tracked)"]
    per_type_frames = []
    for pt in types_to_run:
        pt_df = sales_df[sales_df["product_type"] == pt]
        agg = aggregate_periods(pt_df, ["product_type"], freq).sort_values("period").reset_index(drop=True)
        rows = []
        for i in range(max(len(agg) - n_periods, 2), len(agg)):
            history_before = agg["actual_kg"].iloc[:i].tolist()
            f = trend_forecast(history_before)
            if f is not None:
                rows.append({"period": agg["period"].iloc[i], "forecast_kg": f})
        if rows:
            per_type_frames.append(pd.DataFrame(rows))
    if not per_type_frames:
        return pd.DataFrame(columns=["period", "forecast_kg"])
    combined = pd.concat(per_type_frames, ignore_index=True)
    return combined.groupby("period", as_index=False)["forecast_kg"].sum()


@st.cache_data(ttl=900, max_entries=16, hash_funcs={pd.DataFrame: _cheap_data_fingerprint})
def compute_trending_shares(sales_df, group_cols, freq="W", damping=0.6):
    """Like compute_shares, but projects each segment's share FORWARD based on its own
    recent trend, instead of just averaging history. Real problem this solves: a flat
    average always lags a segment that's genuinely, steadily growing or shrinking its share
    (e.g. a channel taking a growing piece of the business) -- by the time the average
    catches up, it's already behind where the trend is actually heading. Uses the same
    damped growth-rate approach as the rest of the app's forecasting (recent vs prior
    period, growth clamped to +-50%), applied to each segment's share of the total rather
    than to raw volume. Renormalized afterward so every segment's projected share still
    sums to exactly 1.0, and the eventual kg values sum back to the trusted total exactly,
    same guarantee as compute_shares."""
    d = sales_df.copy()
    d["record_date"] = pd.to_datetime(d["record_date"], errors="coerce")
    d = d.dropna(subset=["record_date"])
    if d.empty:
        return pd.DataFrame(columns=list(group_cols) + ["share"])

    if freq == "W":
        d["period"] = (d["record_date"] - pd.to_timedelta(d["record_date"].dt.weekday, unit="D")).dt.date.astype(str)
    else:
        d["period"] = d["record_date"].dt.to_period("M").astype(str)

    period_totals = d.groupby("period")["kg"].sum().rename("period_total")
    seg_period = d.groupby(list(group_cols) + ["period"])["kg"].sum().rename("seg_kg").reset_index()
    seg_period = seg_period.merge(period_totals, on="period")
    seg_period["seg_share"] = seg_period["seg_kg"] / seg_period["period_total"].replace(0, np.nan)

    projected_rows = []
    for key, sub in seg_period.groupby(list(group_cols)):
        sub = sub.sort_values("period")
        share_history = sub["seg_share"].dropna().tolist()
        if len(share_history) < 2:
            projected_share = share_history[-1] if share_history else 0.0
        else:
            n = len(share_history)
            recent = np.median(share_history[-min(4, n):])
            if n >= 8:
                prior = np.median(share_history[-8:-4])
            elif n > min(4, n):
                prior = np.median(share_history[:n - min(4, n)])
            else:
                prior = recent
            growth = 0.0
            if prior > 0:
                growth = max(min((recent - prior) / prior, 1.0), -0.5)
            projected_share = max(recent * (1 + damping * growth), 0.0)
        row = dict(zip(group_cols, key if isinstance(key, tuple) else (key,)))
        row["share"] = projected_share
        projected_rows.append(row)

    result = pd.DataFrame(projected_rows)
    total = result["share"].sum()
    if total > 0:
        result["share"] = result["share"] / total  # renormalize so shares sum to exactly 1.0
    return result


@st.cache_data(ttl=900, max_entries=16, show_spinner="Projecting forward...")
@st.cache_data(ttl=3600, max_entries=16, show_spinner=False)
def detect_seasonal_period(series, candidates=(2, 3, 4, 5, 6, 8, 13), min_corr=0.18):
    """Finds a genuine repeating cycle in the data, or returns None.

    Candidates stop at 13 periods deliberately: a 52-week seasonal fit was measured at 2.5s
    versus 0.4s for a short cycle, and with several segments plus the bag breakdown that adds
    up on every page load. Short ordering rhythms are also what actually drives week-to-week
    demand here; a yearly cycle needs years of clean history to estimate honestly anyway.

    A flat forward forecast means the model found no repeating structure -- only noise. The
    honest way to make a forecast move up and down is to model a cycle that's really there
    (a monthly ordering rhythm, a quarterly push, a yearly season), NOT to add wiggle for
    appearance. This measures how strongly the series correlates with itself at each
    candidate lag and returns the strongest one that clears `min_corr`. If nothing clears
    it, a smooth line genuinely is the best available answer."""
    v = np.asarray(series, dtype=float)
    v = v[~np.isnan(v)]
    n = len(v)
    if n < 20:
        return None
    # measure the correlation on FIRST DIFFERENCES, not raw values. A level shift or a strong
    # trend (like a step change in volume) dominates raw autocorrelation and masks a genuine
    # cycle underneath -- verified directly: a real 4-period cycle sitting under a level shift
    # scored higher at the wrong lag on raw values, and only stood out clearly once detrended.
    # score each candidate on BOTH the raw series and its first differences, and take the
    # stronger of the two. Raw works well for a clean cycle; differencing rescues a cycle
    # buried under a level shift. Verified that using differencing alone was WORSE -- it
    # destroyed a clean signal that raw correlation found easily.
    def _scores(x):
        x = x - x.mean()
        den = float(np.dot(x, x))
        return (lambda lag: float(np.dot(x[:-lag], x[lag:]) / den)) if den > 0 else (lambda lag: 0.0)

    raw_score = _scores(v)
    d = np.diff(v)
    diff_score = _scores(d) if len(d) >= 8 else (lambda lag: 0.0)

    best, best_corr = None, min_corr
    for lag in candidates:
        if n < lag * 2 + 2:      # need at least two full cycles to believe it
            continue
        corr = max(raw_score(lag), diff_score(lag) if len(d) >= lag * 2 + 2 else 0.0)
        if corr > best_corr:
            best, best_corr = lag, corr
    return best


def project_forward_with_range(actual_series, error_sigma, n_periods=8, keep_trend=False,
                                seasonal_period=None, order=None):
    """Projects multiple periods ahead using a single non-seasonal ARIMA(1,1,1) fit, which
    produces the whole path at once with real statistical confidence intervals (verified:
    recursive re-feeding through ARIMA produced a forecast that more than doubled over 8
    weeks with no plateau; the native multi-step call plateaus correctly). Falls back to the
    damped recursive median method when there's too little data.

    The seasonal tier was REMOVED. Real reason, from a reported production failure: with 2+
    years of history this ran a seasonal SARIMA(1,1,1)x(1,1,1,52) fit, measured at ~16s each
    on a fast machine and considerably slower on Streamlit Cloud's shared CPU. The
    Staple/Single panel calls this once per product type automatically on page load, and one
    long blocking fit prevents Streamlit from answering the browser's heartbeat -- producing
    a "Connection timed out" error mid-render, which then forces a refresh and starts the
    whole thing over. Same tradeoff already accepted elsewhere in this app: explicit yearly
    seasonality is given up in exchange for the app actually staying connected."""
    vals = np.asarray(actual_series, dtype=float)
    n = len(vals)
    if n >= 8:
        # use a real detected cycle when one exists, so the projection genuinely rises and
        # falls the way the business does instead of flattening to an average
        _m = seasonal_period if seasonal_period else detect_seasonal_period(list(vals))
        seasonal = (1, 0, 1, _m) if _m else (0, 0, 0, 0)
        try:
            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                from statsmodels.tsa.statespace.sarimax import SARIMAX
                # keep_trend adds a linear time term. Without it, ARIMA(1,1,1) converges to a
                # flat line within about 3-4 steps -- fine for a 1-week forecast, but it makes
                # a multi-month projection go dead flat, which is what people were seeing on
                # the monthly view. The trend term keeps a real slope; the cap below stops it
                # extrapolating away to something implausible.
                model = SARIMAX(vals, order=(order or (1, 1, 1)), seasonal_order=seasonal,
                                 trend="t" if keep_trend else "n",
                                 enforce_stationarity=False, enforce_invertibility=False)
                fit = model.fit(disp=False)
                result = fit.get_forecast(steps=n_periods)
                mean = result.predicted_mean
                ci = result.conf_int(alpha=0.20)  # ~80% interval, roughly matching a P10-P90 framing
                # a linear trend runs forever, so bound the whole path to a sane multiple of
                # recent reality rather than letting month 12 imply a tripling
                recent_ref = float(np.median(vals[-6:])) if len(vals) >= 6 else float(np.median(vals))
                cap_hi = recent_ref * 2.5 if recent_ref > 0 else None
                cap_lo = recent_ref * 0.3 if recent_ref > 0 else None
                path = []
                for h in range(n_periods):
                    f = max(float(mean[h]), 0)
                    if cap_hi is not None:
                        f = min(max(f, cap_lo), cap_hi)
                    low = max(float(ci[h, 0]), 0)
                    high = max(float(ci[h, 1]), f)
                    path.append({"step": h + 1, "forecast_kg": f, "low": low, "high": high})
                return pd.DataFrame(path)
        except Exception:
            pass

    hist_vals = list(actual_series)[-8:]
    path = []
    for h in range(1, n_periods + 1):
        effective_damping = 0.6 * (0.7 ** (h - 1))
        f = _median_trend_forecast(hist_vals[-8:], damping=effective_damping)
        if f is None:
            break
        band = error_sigma * np.sqrt(h) if pd.notna(error_sigma) else 0.15 * np.sqrt(h)
        path.append({"step": h, "forecast_kg": f, "low": max(f * (1 - band), 0), "high": f * (1 + band)})
        hist_vals.append(f)
    return pd.DataFrame(path)


# ===================================================================
# GLOBAL STATE
# ===================================================================
sales_df = load_sales_records()
has_data = not sales_df.empty

# brand styling -- 49th Parallel's teal and typography, so this reads as a company tool
# rather than a generic dashboard
st.markdown("""
<style>
  :root { --p49-teal:#2F6F6B; --p49-ink:#1A1A1A; }
  .p49-header { display:flex; align-items:center; gap:16px; border-bottom:3px solid var(--p49-teal);
                padding-bottom:14px; margin-bottom:6px; }
  .p49-mark { font-family:Georgia,'Times New Roman',serif; line-height:1; text-align:center;
              color:var(--p49-ink); border-right:1px solid #d8d8d8; padding-right:16px; }
  .p49-mark .n { font-size:34px; font-weight:700; letter-spacing:1px; }
  .p49-mark .w { font-size:11px; letter-spacing:4px; margin-top:2px; }
  .p49-mark .s { font-size:7px; letter-spacing:2.5px; color:#666; margin-top:2px; }
  .p49-title { font-size:30px; font-weight:700; color:var(--p49-ink); line-height:1.15; }
  .p49-sub { font-size:13px; color:var(--p49-teal); font-weight:600; letter-spacing:.5px; }
  div.stButton > button[kind="primary"] { background-color:var(--p49-teal); border-color:var(--p49-teal); }
  div.stButton > button[kind="primary"]:hover { background-color:#255955; border-color:#255955; }
  [data-testid="stMetricValue"] { color:var(--p49-ink); }
</style>
<div class="p49-header">
  <div class="p49-mark">
    <div class="n">49<sup style="font-size:14px">TH</sup></div>
    <div class="w">PARALLEL</div>
    <div class="s">COFFEE ROASTERS</div>
  </div>
  <div>
    <div class="p49-title">Sales to Operations Demand Planning</div>
    <div class="p49-sub">S&amp;OP &middot; FORECAST &middot; CAPACITY</div>
  </div>
</div>
""", unsafe_allow_html=True)

# default the cycle to the month of the most recent data, not today's date -- if the latest
# upload covers July, the planning cycle should say July, otherwise events and sign-offs get
# filed against a month with no data in it
_default_cycle = current_cycle_label()
if has_data:
    _d = pd.to_datetime(sales_df["record_date"], errors="coerce").dropna()
    if not _d.empty:
        _default_cycle = _d.max().strftime("%Y-%m")
cycle = st.text_input("Planning cycle label", value=_default_cycle,
                       help="Defaults to the month of your most recent sales data, so events and "
                            "sign-offs are filed against the period you're actually planning.")
if has_data:
    _d2 = pd.to_datetime(sales_df["record_date"], errors="coerce").dropna()
    if not _d2.empty:
        st.caption(f"Data loaded covers {_d2.min().strftime('%b %d, %Y')} to "
                   f"{_d2.max().strftime('%b %d, %Y')} ({_d2.dt.to_period('M').nunique()} months).")

if has_data:
    price_df = compute_price_per_kg(sales_df)
    size_mix_df = compute_size_mix(sales_df)
    customer_mix_df = compute_customer_mix(sales_df)
    weekly_actual = compute_weekly_actuals(sales_df)
    generate_missing_forecasts(weekly_actual)
    # Backtest is the single most expensive thing in the app (~94,000 model fits before the
    # recent-weeks cap). It now runs ONLY when asked for, not on every page load -- Streamlit
    # reruns this whole script on every click, so anything unguarded here is paid for
    # constantly. Accuracy analysis is something you look at deliberately, not something
    # that needs recomputing every time someone switches a dropdown.
    backtest_df = st.session_state.get("backtest_df", pd.DataFrame())
else:
    price_df = size_mix_df = customer_mix_df = weekly_actual = backtest_df = pd.DataFrame()

# --- current live forecast: latest frozen target week per channel/product ---
@st.cache_data(ttl=7200, max_entries=1, show_spinner=False)
def _load_live_forecast():
    """Cached: this returns one row per channel/product (hundreds of rows) and was being
    re-downloaded on every single page interaction. Every byte is billed Supabase egress."""
    return pd.read_sql("""
        SELECT af.id, af.channel, af.product, af.target_week, af.forecast_kg
        FROM auto_forecasts af
        INNER JOIN (SELECT channel, product, MAX(target_week) AS mx FROM auto_forecasts GROUP BY channel, product) t
        ON af.channel=t.channel AND af.product=t.product AND af.target_week=t.mx
    """, conn)


live_forecast = _load_live_forecast()
if not live_forecast.empty:
    # auto_forecasts is append-only -- if a (channel, product, target_week) combo was ever
    # forecasted more than once (a real bug found and fixed elsewhere: two rows summed to an
    # inflated total), this JOIN would return both rows. Dedupe to the latest by id so every
    # number built from live_forecast (KPIs, the Forecast tab, everything) is never at risk
    # of silently double-counting a duplicate.
    live_forecast = live_forecast.sort_values("id").drop_duplicates(
        subset=["channel", "product", "target_week"], keep="last")

# --- pipeline events, layered on top of the live forecast ---
all_events_all = pd.read_sql("SELECT * FROM pipeline_events ORDER BY id DESC", conn)
if "active" not in all_events_all.columns:
    all_events_all["active"] = 1
all_events_all["active"] = all_events_all["active"].fillna(1).astype(int)
# only ACTIVE events shape any forecast. Inactive ones stay in the database so the historical
# record of what was predicted at the time survives -- deleting them outright would silently
# rewrite past forecast numbers and past accuracy, which makes the accuracy story untrustworthy.
all_events = all_events_all[all_events_all["active"] == 1]
if not all_events.empty:
    applicable = all_events[
        (all_events["starting_cycle"] <= cycle) &
        ((all_events["ongoing"] == 1) | (all_events["starting_cycle"] == cycle))
    ]
    # convert the event's MONTHLY figure into a weekly contribution before it meets a weekly
    # forecast. Real bug: the monthly number was added straight onto next week's forecast, so
    # a 4,000 kg/month event inflated a single week by 4,000 kg -- about 4.3x too much.
    # 4.345 = average weeks per month (365.25 / 7 / 12).
    WEEKS_PER_MONTH = 4.345
    pipeline_by_cp = applicable.groupby(["channel", "product"], as_index=False)["expected_kg_per_month"].sum() \
        .rename(columns={"expected_kg_per_month": "pipeline_kg"})
    pipeline_by_cp["pipeline_kg"] = pipeline_by_cp["pipeline_kg"] / WEEKS_PER_MONTH
else:
    applicable = pd.DataFrame()
    pipeline_by_cp = pd.DataFrame(columns=["channel", "product", "pipeline_kg"])

if not live_forecast.empty or not pipeline_by_cp.empty:
    forecast_by_cp = live_forecast[["channel", "product", "forecast_kg", "target_week"]].copy() if not live_forecast.empty \
        else pd.DataFrame(columns=["channel", "product", "forecast_kg", "target_week"])

    # RECONCILE the per-item forecasts to the segment-based company total, so every tab
    # tells the same story. Real inconsistency this fixes: these per-item numbers are frozen
    # bottom-up forecasts (one per channel/product, hundreds of them), while the Dashboard
    # forecasts three segments directly. Summing hundreds of small independent forecasts and
    # forecasting three aggregates are genuinely different calculations with no reason to
    # agree -- so the Forecast (auto) tab, Ops capacity, and Ask AI were all quoting a
    # different total than the Dashboard. Scaling every item by one shared factor keeps the
    # relative item-level detail exactly as-is while making the total match by construction.
    if has_data and not forecast_by_cp.empty:
        _seg_fc = compute_segment_forecast(sales_df, freq="W")
        _seg_total = sum(_seg_fc.values()) if _seg_fc else 0
        _item_total = forecast_by_cp["forecast_kg"].sum()
        if _seg_total > 0 and _item_total > 0:
            forecast_by_cp["forecast_kg"] = forecast_by_cp["forecast_kg"] * (_seg_total / _item_total)

    forecast_by_cp = forecast_by_cp.merge(pipeline_by_cp, on=["channel", "product"], how="outer")
    forecast_by_cp["forecast_kg"] = forecast_by_cp["forecast_kg"].fillna(0)
    forecast_by_cp["pipeline_kg"] = forecast_by_cp["pipeline_kg"].fillna(0)
    forecast_by_cp["forecast_kg"] = forecast_by_cp["forecast_kg"] + forecast_by_cp["pipeline_kg"]
else:
    forecast_by_cp = pd.DataFrame()

# --- manual overrides: human judgment REPLACES the auto+pipeline number, not adds to it ---
manual_overrides_df = pd.read_sql("SELECT * FROM manual_overrides WHERE active = 1 ORDER BY id DESC", conn)
active_overrides = pd.DataFrame()
if not manual_overrides_df.empty:
    latest_overrides = manual_overrides_df.sort_values("id").groupby(["channel", "product"], as_index=False).last()

    # a "One-time" override only applies while the current live forecast's target_week still
    # matches what it was when the override was set -- once new data moves the target week
    # forward, it's expired, and we auto-deactivate it so it doesn't clutter the active list
    still_valid_rows = []
    for _, row in latest_overrides.iterrows():
        if row["period_type"] == "One-time" and not live_forecast.empty:
            current_target = live_forecast[(live_forecast["channel"] == row["channel"]) &
                                            (live_forecast["product"] == row["product"])]
            if current_target.empty or current_target.iloc[0]["target_week"] != row["target_week"]:
                conn.execute("UPDATE manual_overrides SET active = 0 WHERE id = ?", (int(row["id"]),))
                conn.commit()
                continue
        still_valid_rows.append(row)
    active_overrides = pd.DataFrame(still_valid_rows) if still_valid_rows else pd.DataFrame()

    if not active_overrides.empty:
        if not forecast_by_cp.empty:
            forecast_by_cp["forecast_kg"] = forecast_by_cp["forecast_kg"].fillna(0)
            forecast_by_cp["pipeline_kg"] = forecast_by_cp["pipeline_kg"].fillna(0)
            # resolve per row through the shared matcher rather than a plain merge -- a merge
            # can only match exact channel+product pairs, so a wildcard override (e.g. a whole
            # channel) would silently never apply here even though it applies elsewhere.
            # An override with a specific customer is deliberately NOT applied at this level:
            # these rows are channel x item totals across all customers, so replacing the whole
            # row with one customer's number would overstate it.
            _row_ov = [
                find_matching_override(active_overrides, channel=str(r["channel"]),
                                        product=str(r["product"]), customer=None)
                for _, r in forecast_by_cp.iterrows()
            ]
            forecast_by_cp["forecast_kg"] = [
                round(ov, 1) if ov is not None else base
                for ov, base in zip(_row_ov, forecast_by_cp["forecast_kg"])
            ]
        else:
            _seed = active_overrides.copy()
            _seed = _seed[(_seed["channel"] != OVERRIDE_ANY) & (_seed["product"] != OVERRIDE_ANY)]
            if not _seed.empty:
                forecast_by_cp = _seed.rename(columns={"override_kg": "forecast_kg"})
                forecast_by_cp["pipeline_kg"] = 0
                forecast_by_cp["target_week"] = None

# size-level breakdown for ops/bag counts, driven by the auto forecast
if not forecast_by_cp.empty and not size_mix_df.empty:
    translated = forecast_by_cp.merge(size_mix_df[["channel", "product", "size_label", "size_mix_pct"]],
                                       on=["channel", "product"], how="left")
    translated["forecast_kg"] = (translated["forecast_kg"] * translated["size_mix_pct"].fillna(100) / 100).round(1)
    translated = translated[["channel", "product", "size_label", "forecast_kg"]]

    # convert each size's forecast kg into an actual BAG COUNT -- this is the number
    # Operations places an order with. Uses a real kg-per-bag rate learned from your own
    # sales (total kg / total units for that size), not a value parsed from the label text,
    # since real fill weights differ from the nominal name.
    _kg_per_bag = compute_kg_per_bag(sales_df)
    if not _kg_per_bag.empty:
        translated = translated.merge(_kg_per_bag, on="size_label", how="left")
        # force plain float64 on both sides before dividing. Real bug: after the merge these
        # could be nullable (Float64) or downcast (float32) dtypes, and np.ceil raises a
        # ValueError on those rather than handling them. Converting explicitly keeps NaN
        # behaviour predictable for sizes with no learned rate.
        _kg = pd.to_numeric(translated["forecast_kg"], errors="coerce").astype("float64")
        _rate = pd.to_numeric(translated["kg_per_bag"], errors="coerce").astype("float64")
        _rate = _rate.where(_rate > 0)  # avoid divide-by-zero -> inf
        translated["forecast_bags"] = np.ceil(_kg.div(_rate))
    else:
        translated["kg_per_bag"] = np.nan
        translated["forecast_bags"] = np.nan
else:
    translated = pd.DataFrame()

# implied $ CAD value of the forecast, using the real computed price per kg
if not translated.empty and not price_df.empty:
    # Two real failure modes fixed here, both of which made the dollar figure swing wildly
    # against the same kg:
    #   1. If price_df had more than one row per channel/product/size, the merge DUPLICATED
    #      forecast rows -- inflating both kg and dollars.
    #   2. If a size had no rate, price_per_kg came back NaN and that volume was valued at
    #      ZERO -- silently understating revenue with no warning.
    _rates = price_df[["channel", "product", "size_label", "price_per_kg"]].drop_duplicates(
        subset=["channel", "product", "size_label"], keep="first")
    dollar_view = translated.merge(_rates, on=["channel", "product", "size_label"], how="left")

    # cascading fallback: exact size rate -> that item's average rate -> that channel's
    # average -> company-wide average. Valuing volume at zero is never the right answer.
    # Every fallback rate is VOLUME-WEIGHTED, not a plain mean. A plain average treats a
    # customer buying 5 kg the same as one buying 5,000 kg, which misprices the forecast
    # whenever the mix is uneven -- and it always is. Weighting by real kg makes the fallback
    # reflect what the business actually sells.
    if {"total_kg", "total_revenue"}.issubset(price_df.columns):
        _w = price_df.copy()
    else:
        # reconstruct volume weights from the source data when price_df doesn't carry them
        _wsrc = sales_df.copy()
        _wsrc["record_date"] = pd.to_datetime(_wsrc["record_date"], errors="coerce")
        _cut = _wsrc["record_date"].max() - pd.Timedelta(days=45)
        _recent = _wsrc[_wsrc["record_date"] >= _cut]
        _wsrc = _recent if len(_recent) >= 20 else _wsrc
        _w = _wsrc.groupby(["channel", "product", "size_label"], as_index=False).agg(
            total_kg=("kg", "sum"), total_revenue=("revenue", "sum"))

    def _weighted_rate(keys, name):
        g = _w.groupby(keys, as_index=False).agg(_k=("total_kg", "sum"), _r=("total_revenue", "sum"))
        g[name] = g["_r"] / g["_k"].replace(0, np.nan)
        return g[keys + [name]]

    _item_rate = _weighted_rate(["channel", "product"], "_item_rate")
    _chan_rate = _weighted_rate(["channel"], "_chan_rate")
    _gk, _gr = float(_w["total_kg"].sum()), float(_w["total_revenue"].sum())
    _global_rate = (_gr / _gk) if _gk > 0 else np.nan
    dollar_view = dollar_view.merge(_item_rate, on=["channel", "product"], how="left")
    dollar_view = dollar_view.merge(_chan_rate, on="channel", how="left")
    dollar_view["_rate_used"] = (dollar_view["price_per_kg"]
                                 .fillna(dollar_view["_item_rate"])
                                 .fillna(dollar_view["_chan_rate"])
                                 .fillna(_global_rate))
    # Refine using CUSTOMER-level rates where the data supports it. Different customers pay
    # genuinely different prices for the same item, so a single channel/item/size rate
    # misprices the forecast whenever the customer mix is uneven. Rather than pick one
    # customer's price, blend them by each customer's real share of that item's volume --
    # so the rate used reflects who is actually buying it.
    try:
        _cust_rates = compute_customer_price_per_kg(sales_df)
        if not _cust_rates.empty and "customer" in sales_df.columns:
            _cm = sales_df.copy()
            _cm["record_date"] = pd.to_datetime(_cm["record_date"], errors="coerce")
            _ccut = _cm["record_date"].max() - pd.Timedelta(days=45)
            _cr = _cm[_cm["record_date"] >= _ccut]
            _cm = _cr if len(_cr) >= 20 else _cm
            _cm = _cm[_cm["customer"] != "(not tracked)"]
            if not _cm.empty:
                _mix = _cm.groupby(["channel", "product", "customer"], as_index=False)["kg"].sum()
                _tot = _mix.groupby(["channel", "product"], as_index=False)["kg"].sum() \
                    .rename(columns={"kg": "_item_kg"})
                _mix = _mix.merge(_tot, on=["channel", "product"], how="left")
                _mix["_share"] = _mix["kg"] / _mix["_item_kg"].replace(0, np.nan)
                _mix = _mix.merge(
                    _cust_rates[["channel", "product", "customer", "effective_price_per_kg"]]
                    if "effective_price_per_kg" in _cust_rates.columns
                    else _cust_rates[["channel", "product", "customer", "price_per_kg"]].rename(
                        columns={"price_per_kg": "effective_price_per_kg"}),
                    on=["channel", "product", "customer"], how="left")
                _mix = _mix.dropna(subset=["effective_price_per_kg", "_share"])
                if not _mix.empty:
                    _blend = _mix.assign(_wp=_mix["effective_price_per_kg"] * _mix["_share"]) \
                        .groupby(["channel", "product"], as_index=False).agg(
                            _cust_blend=("_wp", "sum"), _cov=("_share", "sum"))
                    # only trust the blend where it covers most of that item's volume
                    _blend = _blend[_blend["_cov"] >= 0.6]
                    dollar_view = dollar_view.merge(
                        _blend[["channel", "product", "_cust_blend"]], on=["channel", "product"], how="left")
                    dollar_view["_rate_used"] = dollar_view["_cust_blend"].fillna(dollar_view["_rate_used"])
    except Exception:
        pass  # customer refinement is a bonus; never let it break the dollar figure

    _unpriced_kg = float(dollar_view.loc[dollar_view["_rate_used"].isna(), "forecast_kg"].sum())
    dollar_view["forecast_cad"] = (dollar_view["forecast_kg"] * dollar_view["_rate_used"]).round(2)
    dollar_by_cp = dollar_view.groupby(["channel", "product"], as_index=False).agg(
        forecast_kg=("forecast_kg", "sum"), forecast_cad=("forecast_cad", "sum"))
else:
    dollar_by_cp = pd.DataFrame()


# ===================================================================
# TABS
# ===================================================================
tab_dash, tab_data, tab_rates, tab_forecast, tab_salesplan, tab_pipeline, tab_ops, tab_signoff, tab_history = st.tabs(
    ["Dashboard", "1. Upload sales data", "2. Computed rates", "3. Forecast (auto)",
     "4. Sales plan (S&OP)", "5. Adjust the forecast", "6. Ops capacity check", "7. Sign-off", "8. History"]
)

# --- DASHBOARD (landing page) ---
with tab_dash:
    if not has_data:
        st.info("No sales data uploaded yet. Go to **1. Upload sales data** to get started.")
    elif weekly_actual.empty or weekly_actual["week_start"].nunique() < 3:
        # gate on actual data availability, NOT on the backtest -- the backtest is now
        # opt-in, so gating the whole dashboard on it would have hidden everything
        st.warning("Not enough history yet to forecast — need at least a few weeks of data per channel/product.")
    else:
        dim_map = {"Channel": "channel", "Item": "product", "Customer": "customer"}

        # ===============================================================
        # OVERVIEW — always whole-company, never filtered. KPIs + one chart.
        # ===============================================================
        st.markdown("### Overview")

        # Three independent segment forecasts (Single, Staple—Specialty Retail, Staple—other
        # channels), each fit on its own history. Their SUM becomes the official company
        # total -- the total is built up from the segments, not split down into them.
        type_level_forecasts = compute_segment_forecast(sales_df, freq="W")

        # attribute each pipeline event's kg impact to ITS OWN product type, not one lump
        # sum floating at the top level -- real gap found: a Staple-item contract wouldn't
        # show up in the Staple panel's own number, only in the overall KPI, meaning Staple
        # + Single would silently stop summing to the KPI total the moment any pipeline
        # event existed. Joining through each product's known classification fixes this the
        # same way everything else this session got reconciled.
        pipeline_by_type = {}
        pipeline_by_segment = {}
        if not pipeline_by_cp.empty and "product_type" in sales_df.columns:
            # match on a normalised key: an event typed with different case or a stray trailing
            # space would otherwise fail to find its product and get dumped into "(not tracked)".
            # Also take ONE product_type per product -- if the same item appears under two
            # classifications in the data, a plain merge would duplicate the event's kg.
            _lookup = sales_df[["product", "product_type"]].copy()
            _lookup["_k"] = _lookup["product"].astype(str).str.strip().str.casefold()
            _lookup = _lookup.dropna(subset=["product_type"]).drop_duplicates(subset=["_k"], keep="first")
            _pl = pipeline_by_cp.copy()
            _pl["_k"] = _pl["product"].astype(str).str.strip().str.casefold()
            pipeline_typed = _pl.merge(_lookup[["_k", "product_type"]], on="_k", how="left")
            pipeline_typed["product_type"] = pipeline_typed["product_type"].fillna("(not tracked)")
            pipeline_by_type = pipeline_typed.groupby("product_type")["pipeline_kg"].sum().to_dict()

            # ALSO attribute each event to its actual SEGMENT, using channel as well as product
            # type. Real bug: grouping by product type alone couldn't tell "Staple — Specialty
            # Retail" apart from "Staple — other channels", so a Specialty Retail event was
            # spread across both by forecast weight instead of landing on the segment it
            # genuinely belongs to.
            pipeline_by_segment = {}
            _pt2 = pipeline_typed.copy()
            _pt2["_isr"] = _pt2["channel"].astype(str).str.strip().str.casefold() == \
                MAJOR_STAPLE_CHANNEL.strip().casefold()
            for _, _r in _pt2.iterrows():
                if _r["product_type"] == "Single":
                    _seg = "Single"
                elif _r["product_type"] == "Staple":
                    _seg = f"Staple — {MAJOR_STAPLE_CHANNEL}" if _r["_isr"] else "Staple — other channels"
                else:
                    continue
                pipeline_by_segment[_seg] = pipeline_by_segment.get(_seg, 0.0) + float(_r["pipeline_kg"])
        pipeline_total_next_week = pipeline_by_cp["pipeline_kg"].sum() if not pipeline_by_cp.empty else 0

        # each segment's forecast now includes its own attributed pipeline events, so the
        # segment panels stay consistent with the KPI total below. Staple events are
        # apportioned across the two Staple segments by their own forecast weights.
        # add each event to the exact segment it belongs to (channel-aware), rather than
        # splitting a product-type total across segments by weight
        type_level_forecasts_with_pipeline = {
            label: val + pipeline_by_segment.get(label, 0.0)
            for label, val in type_level_forecasts.items()
        }
        unattributed_pipeline = pipeline_by_type.get("(not tracked)", 0)  # events on products with no known type

        # fold manual overrides into the segment numbers BEFORE totalling, so the headline
        # KPI, the segment tables and every chart all reflect them
        type_level_forecasts_with_pipeline, override_notes = apply_overrides_to_segments(
            sales_df, type_level_forecasts_with_pipeline, active_overrides, freq="W")

        if type_level_forecasts:
            next_week_kg_all = sum(type_level_forecasts_with_pipeline.values()) + unattributed_pipeline
        else:
            next_week_kg_all = forecast_by_cp["forecast_kg"].sum() if not forecast_by_cp.empty else 0
        next_week_cad_all = dollar_by_cp["forecast_cad"].sum() if not dollar_by_cp.empty else 0

        # the forecast period, in plain dates -- e.g. "Aug 17 - Aug 23, 2026". Derived from
        # LIVE weekly_actual (the same source next_week_kg_all uses), not from the old
        # forecast_by_cp/auto_forecasts lookup -- those are two different things and can
        # genuinely disagree: auto_forecasts is a frozen historical record that's never
        # cleaned up, so deleting a batch of sales data doesn't retroactively remove a
        # forecast that was generated assuming that data existed. Using the live source for
        # both the label and the number keeps them from ever showing an inconsistent pair.
        forecast_period_label = "n/a"
        if not weekly_actual.empty:
            latest_known_week = sorted(weekly_actual["week_start"].unique())[-1]
            tw_start = pd.Timestamp(latest_known_week) + pd.Timedelta(days=7)
            tw_end = tw_start + pd.Timedelta(days=6)
            forecast_period_label = f"{tw_start.strftime('%b %d')} – {tw_end.strftime('%b %d, %Y')}"

        d_kpi = sales_df.copy()
        d_kpi["record_date"] = pd.to_datetime(d_kpi["record_date"], errors="coerce")
        d_kpi = d_kpi.dropna(subset=["record_date"])
        latest_actual_kg = None
        latest_actual_week = None
        if not d_kpi.empty:
            d_kpi["week_start"] = (d_kpi["record_date"] - pd.to_timedelta(d_kpi["record_date"].dt.weekday, unit="D")).dt.date.astype(str)
            wk_kpi = d_kpi.groupby("week_start")["kg"].sum().sort_index()
            if len(wk_kpi):
                latest_actual_kg = wk_kpi.iloc[-1]
                latest_actual_week = wk_kpi.index[-1]

        # last week's forecast accuracy, AND whether it's improving vs the week before --
        # walk-forward recomputed using the CURRENT method (same fix applied to the Staple/
        # Single panel earlier), not a lookup into old frozen auto_forecasts rows, which
        # could show a stale number from a since-replaced forecasting method.
        last_week_accuracy_label, last_week_accuracy_delta = "n/a", None
        if not d_kpi.empty:
            company_weekly = d_kpi.groupby("week_start")["kg"].sum().sort_index()
            weeks_list = company_weekly.index.tolist()

            # measure accuracy against the SAME segment-based forecast the rest of the
            # dashboard shows. Real bug: this used to fit one whole-company aggregate model
            # of its own, which is a third distinct method -- so the accuracy number was
            # grading a forecast that appeared nowhere else in the app.
            acc_wf = walk_forward_all_segments(sales_df, n_periods=6, freq="W")
            if not acc_wf.empty:
                # add back what any logged event contributed in each of those weeks, so this
                # grades the forecast the app actually displayed rather than a pure
                # statistical number that was never shown to anyone
                _acc_adj = event_adjustment_by_period(all_events_all, acc_wf["period"].tolist(), freq="W")
                acc_wf = acc_wf.copy()
                acc_wf["forecast_kg"] = acc_wf["forecast_kg"] + acc_wf["period"].map(_acc_adj).fillna(0)
            acc_map = dict(zip(acc_wf["period"], acc_wf["forecast_kg"])) if not acc_wf.empty else {}

            def _week_accuracy(idx):
                if idx < 0 or idx >= len(weeks_list):
                    return None
                wk = weeks_list[idx]
                f = acc_map.get(wk)
                actual = company_weekly.iloc[idx]
                if f is None or f <= 0:
                    return None
                pct_off = (actual - f) / f * 100
                return max(0, 100 - abs(pct_off))

            this_week_accuracy = _week_accuracy(len(weeks_list) - 1) if len(weeks_list) >= 1 else None
            prior_week_accuracy = _week_accuracy(len(weeks_list) - 2) if len(weeks_list) >= 2 else None

            if this_week_accuracy is not None:
                last_week_accuracy_label = f"{this_week_accuracy:.0f}% accurate"
                if prior_week_accuracy is not None:
                    trend = this_week_accuracy - prior_week_accuracy
                    last_week_accuracy_delta = f"{trend:+.0f} pts vs the week before"
                else:
                    last_week_accuracy_delta = "no prior week to compare yet"

        cap_row = pd.read_sql("SELECT * FROM ops_capacity WHERE cycle_label = ? ORDER BY id DESC LIMIT 1",
                               conn, params=(cycle,))
        cap_gap = None
        if not cap_row.empty and not forecast_by_cp.empty:
            cap_amt = cap_row.iloc[0]["monthly_capacity_kg"]
            monthly_planned = next_week_kg_all * 4.345
            cap_gap = cap_amt - monthly_planned

        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric(f"Forecast: {forecast_period_label}", f"{next_week_kg_all:,.0f} kg",
                  help="The 7-day period this forecast covers, shown in the label above.")
        # show the implied $/kg -- the fastest way to spot a valuation problem. If this number
        # looks nothing like your real average selling price, the dollar figure is wrong even
        # when the kg figure is right, and that's otherwise very hard to notice.
        _implied = (next_week_cad_all / next_week_kg_all) if next_week_kg_all else 0
        k2.metric("Forecast value", f"${next_week_cad_all:,.0f}",
                  help=f"Implied average rate: ${_implied:,.2f}/kg. If that looks wrong against your "
                       "real selling prices, check tab 2 (Computed rates) — the dollar figure is only "
                       "as good as the rates behind it.")
        k3.metric(f"Actual — week of {latest_actual_week}" if latest_actual_week else "Last week actual",
                  f"{latest_actual_kg:,.0f} kg" if latest_actual_kg is not None else "n/a")
        k4.metric("Last week's forecast", last_week_accuracy_label, delta=last_week_accuracy_delta,
                  delta_color="off", help="How close last week's forecast came to what actually happened, "
                                          "and whether that's better or worse than the week before it — "
                                          "so you can tell if accuracy is trending up or down over time.")
        if cap_gap is not None:
            k5.metric("Capacity", "Shortfall" if cap_gap < 0 else "Covered",
                      delta=f"{cap_gap:,.0f} kg/mo", delta_color="normal" if cap_gap >= 0 else "inverse")
        else:
            k5.metric("Capacity", "Not set", help="Set it in tab 5 to see a shortfall check here.")

        # show the arithmetic behind the headline number. Without this it's impossible to tell
        # from the dashboard whether a logged event or override actually reached the total --
        # you'd have to compare against a number you no longer have.
        _base_only = sum(type_level_forecasts.values()) if type_level_forecasts else 0
        _event_delta = pipeline_total_next_week
        _override_delta = next_week_kg_all - _base_only - _event_delta
        # always shown -- this is the "where does this number come from" answer, and it's
        # the first thing anyone asks in a planning meeting
        if True:
            with st.expander("How this total is built up", expanded=False):
                _parts = [{"Component": "Statistical forecast (three segments)", "kg": round(_base_only, 1)}]
                if abs(_event_delta) > 0.5:
                    _parts.append({"Component": "Pipeline events (converted to weekly)", "kg": round(_event_delta, 1)})
                if abs(_override_delta) > 0.5:
                    _parts.append({"Component": "Manual overrides", "kg": round(_override_delta, 1)})
                _parts.append({"Component": "TOTAL (the number above)", "kg": round(next_week_kg_all, 1)})
                st.dataframe(pd.DataFrame(_parts), use_container_width=True, hide_index=True)
                if type_level_forecasts:
                    st.caption("The statistical forecast is the sum of three independently modelled "
                               "segments: " + ", ".join(
                                   f"{k} {v:,.0f} kg" for k, v in type_level_forecasts.items()) + ".")
                if abs(_event_delta) < 0.5 and abs(_override_delta) < 0.5:
                    st.caption("No pipeline events or manual overrides are currently affecting this number.")
                if abs(_event_delta) > 0.5:
                    st.caption(f"Events are logged per month and converted to a weekly figure "
                               f"(÷ 4.345), which is why a {_event_delta * 4.345:,.0f} kg/month event "
                               f"shows here as {_event_delta:,.0f} kg for one week.")
        elif not applicable.empty or not active_overrides.empty:
            st.warning("You have logged events or overrides, but they aren't changing this total. "
                       "Check that the Channel and Item spelling matches your sales data exactly — "
                       "the Pipeline tab has a before/after table that shows whether each event is landing.")

        st.caption("Solid blue line is **actual** historical sales, whole company. Dashed line (Week view "
                   "only) is what the auto-forecast would have predicted for each of the last 12 weeks, "
                   "checked against what actually happened — a real accuracy check, not a future prediction. "
                   "Tick the box below to also see the forecast projected forward, with a range built from "
                   "this model's own real historical accuracy.")
        trend_freq = st.radio("Show by", ["Week", "Month"], horizontal=True, key="trend_freq")
        d_trend = sales_df.copy()
        d_trend["record_date"] = pd.to_datetime(d_trend["record_date"], errors="coerce")
        d_trend = d_trend.dropna(subset=["record_date"])
        if trend_freq == "Week":
            d_trend["period"] = (d_trend["record_date"] - pd.to_timedelta(d_trend["record_date"].dt.weekday, unit="D")).dt.date.astype(str)
        else:
            d_trend["period"] = d_trend["record_date"].dt.to_period("M").astype(str)
        trend_agg = d_trend.groupby("period", as_index=False)["kg"].sum().sort_values("period")

        # This chart builds its own aggregation rather than going through aggregate_periods,
        # so the partial-month exclusion has to be applied here too -- without it the Month
        # view still showed August at ~6,000 kg (one week standing in for a whole month)
        # against ~30,000 kg months either side, and forecast the year from that false floor.
        # Weekly is deliberately left untouched.
        if trend_freq == "Month" and not trend_agg.empty:
            _tl = d_trend["record_date"].max()
            if pd.notna(_tl) and _tl < _tl.to_period("M").to_timestamp("M"):
                _drop = str(_tl.to_period("M"))
                trend_agg = trend_agg[trend_agg["period"] != _drop]
                st.caption(f"{_tl.strftime('%B')} is still in progress (data through "
                           f"{_tl.strftime('%b %d')}) and is excluded from this monthly view — "
                           "a part-month total would read as a collapse in demand and distort "
                           "everything forecast after it.")

        # same shared top-down calculation the Staple/Single table uses -- this is the actual
        # fix for the inconsistency: previously this summed many small per-item backtests
        # (bottom-up), while the Staple/Single table used one aggregate fit per type
        # (top-down). Different methods, no reason to agree. Now both call the exact same
        # function, so they can't drift apart again.
        wf_freq = "W" if trend_freq == "Week" else "M"
        topdown_bt = walk_forward_all_segments(sales_df, n_periods=26, freq=wf_freq)
        if not topdown_bt.empty:
            # same reasoning as the accuracy KPI -- the dashed line should show what was
            # actually forecast at the time, events included
            _bt_adj = event_adjustment_by_period(all_events_all, topdown_bt["period"].tolist(), freq=wf_freq)
            topdown_bt = topdown_bt.copy()
            topdown_bt["forecast_kg"] = topdown_bt["forecast_kg"] + topdown_bt["period"].map(_bt_adj).fillna(0)
        total_bt = trend_agg.rename(columns={"period": "week_start", "kg": "actual_kg"}).merge(
            topdown_bt.rename(columns={"period": "week_start"}), on="week_start", how="inner")
        total_bt["variance_pct"] = (total_bt["actual_kg"] - total_bt["forecast_kg"]) / total_bt["forecast_kg"].replace(0, np.nan)
        error_sigma = total_bt["variance_pct"].std()

        n_periods_fwd = 8 if trend_freq == "Week" else 6
        show_projection = st.checkbox("Also show the forecast projection", key="show_trend_projection")
        if show_projection:
            with st.spinner("Projecting the trend forward..."):
                projection = project_forward_with_range(trend_agg["kg"].tolist(), error_sigma,
                                                        n_periods=n_periods_fwd,
                                                        keep_trend=True)
        else:
            projection = pd.DataFrame()

        fig_trend = go.Figure()
        # display only the last ~6 months on the chart (recent, relevant view) -- the full
        # history is still used for the underlying error_sigma/projection math above, this
        # is purely a display trim, not a computation change
        n_display = 26 if trend_freq == "Week" else 6
        trend_agg_display = trend_agg.tail(n_display)
        fig_trend.add_trace(go.Scatter(x=trend_agg_display["period"], y=trend_agg_display["kg"], mode="lines", name="Actual",
                                        line=dict(color="rgb(31,119,180)", width=2)))
        if trend_freq == "Week" and not total_bt.empty:
            bt_recent = total_bt.sort_values("week_start").tail(12)
            fig_trend.add_trace(go.Scatter(x=bt_recent["week_start"], y=bt_recent["forecast_kg"], mode="lines",
                                            name="Auto forecast (backtested)",
                                            line=dict(color="rgb(139,90,60)", width=2, dash="dash")))
        if not projection.empty:
            last_date = pd.Timestamp(trend_agg["period"].iloc[-1])
            if trend_freq == "Week":
                future_dates = [(last_date + pd.Timedelta(weeks=int(s))).date().isoformat() for s in projection["step"]]
            else:
                future_dates = [(last_date + pd.DateOffset(months=int(s))).date().isoformat() for s in projection["step"]]
            join_x = [trend_agg["period"].iloc[-1]] + future_dates
            join_y_high = [trend_agg["kg"].iloc[-1]] + projection["high"].tolist()
            join_y_low = [trend_agg["kg"].iloc[-1]] + projection["low"].tolist()
            join_y_mid = [trend_agg["kg"].iloc[-1]] + projection["forecast_kg"].tolist()
            fig_trend.add_trace(go.Scatter(x=join_x, y=join_y_high, mode="lines", line=dict(width=0),
                                            showlegend=False, hoverinfo="skip"))
            fig_trend.add_trace(go.Scatter(x=join_x, y=join_y_low, mode="lines", line=dict(width=0),
                                            fill="tonexty", fillcolor="rgba(120,120,120,0.25)", name="Forecast range",
                                            hoverinfo="skip"))
            fig_trend.add_trace(go.Scatter(x=join_x, y=join_y_mid, mode="lines", name="Forecast",
                                            line=dict(color="rgb(60,60,60)", width=2)))
        fig_trend.update_layout(height=460, margin=dict(l=10, r=10, t=40, b=10), plot_bgcolor="white",
                                 xaxis_title=trend_freq, yaxis_title="Total kg (all channels/items)",
                                 hovermode="x unified", xaxis=dict(showgrid=True, gridcolor="rgba(0,0,0,0.06)"),
                                 yaxis=dict(showgrid=True, gridcolor="rgba(0,0,0,0.06)"),
                                 # legend above the plot instead of floating inside it -- it was
                                 # covering a meaningful chunk of the chart area
                                 legend=dict(orientation="h", yanchor="bottom", y=1.02,
                                             xanchor="left", x=0, font=dict(size=11)))
        st.plotly_chart(fig_trend, use_container_width=True)

        st.divider()

        # ===============================================================
        # SIMPLE FORECAST — by Single vs Staple. Deliberately minimal --
        # built for a quick manager glance, not detailed analysis.
        # ===============================================================
        product_types_available = sorted(
            sales_df[sales_df["product_type"] != "(not tracked)"]["product_type"].unique().tolist()
        ) if "product_type" in sales_df.columns and (sales_df["product_type"] != "(not tracked)").any() else []

        segment_map = split_into_segments(sales_df)
        if segment_map:
            st.markdown("## Forecast by segment")
            st.caption("Three groups, each forecast independently from its own history — "
                       "Single, Staple in Specialty Retail, and the rest of Staple. Nothing here "
                       "is a proportional split of a total; the percentages shown are simply what "
                       "each segment's own forecast worked out to be.")
            pt_horizon = st.radio("Show by", ["Week", "Month"], horizontal=True, key="pt_horizon")
            if has_data:
                _ld = pd.to_datetime(sales_df["record_date"], errors="coerce").max()
                if pd.notna(_ld):
                    _me = _ld.to_period("M").to_timestamp("M")
                    if _ld < _me and pt_horizon == "Month":
                        st.info(f"{_ld.strftime('%B')} is still in progress (data through "
                                f"{_ld.strftime('%b %d')}), so it's excluded from the monthly view. "
                                "A part-month total would look like a collapse in demand and would "
                                "distort every month after it.")
            n_periods_shown = 8 if pt_horizon == "Week" else 6
            n_history_shown = 6 if pt_horizon == "Week" else 6
            freq = "W" if pt_horizon == "Week" else "M"

            segment_forecasts = compute_segment_forecast(sales_df, freq=freq)
            if freq == "W" and pipeline_by_segment:
                # each event lands on the exact segment it belongs to, matched on channel as
                # well as product type -- so a Specialty Retail event moves the Specialty
                # Retail segment specifically, not both Staple segments proportionally
                for label in list(segment_forecasts):
                    segment_forecasts[label] += pipeline_by_segment.get(label, 0.0)

            # same override folding the Overview KPI uses, so these tables and charts agree
            # with the headline number instead of quietly ignoring overrides
            segment_forecasts, seg_override_notes = apply_overrides_to_segments(
                sales_df, segment_forecasts, active_overrides, freq=freq)
            if seg_override_notes:
                with st.expander(f"{len(seg_override_notes)} manual override(s) applied to these segments"):
                    st.dataframe(pd.DataFrame(seg_override_notes).rename(columns={
                        "scope": "Override scope", "segment": "Segment",
                        "was_contributing_kg": "Was contributing (kg)",
                        "override_kg": "Override (kg)", "segment_change_kg": "Segment change (kg)"}),
                        use_container_width=True, hide_index=True)
                    st.caption("An override sets what one slice should be. Since a segment covers many "
                               "slices, the segment shifts by the difference between the override and what "
                               "that slice was contributing — everything else in the segment is untouched.")

            seg_labels = list(segment_map.keys())
            if len(seg_labels) < 3:
                st.warning(
                    f"Only {len(seg_labels)} segment(s) found: {', '.join(seg_labels)}. "
                    f"Expected three (Single, Staple — {MAJOR_STAPLE_CHANNEL}, Staple — other channels). "
                    "This usually means the channel name doesn't match, or one group has no data.")
            pt_cols = st.columns(len(seg_labels))
            for idx, pt in enumerate(seg_labels):
                with pt_cols[idx]:
                    st.markdown(f"### {pt}")
                    pt_df = segment_map[pt]

                    agg_pt = aggregate_periods(pt_df, ["product_type"], freq)
                    agg_pt = agg_pt.groupby("period", as_index=False)["actual_kg"].sum().sort_values("period")

                    if len(agg_pt) < 2:
                        st.info("Not enough history yet for this segment.")
                        continue

                    _stored = get_stored_order(pt, "W")
                    if freq == "M":
                        # Forecast WEEKLY and roll up to months, rather than fitting on the
                        # monthly series directly. Real problem: with about 7 complete months
                        # of history, a monthly model is estimated from 7 points -- too few to
                        # be stable, and too few for seasonality detection (needs 20+). That's
                        # why one segment converged to a flat line while another oscillated
                        # wildly: not a difference in the business, just noise in an
                        # under-determined fit. The weekly series has ~30 points, so forecasting
                        # there and aggregating uses all the data and behaves consistently.
                        _wagg = aggregate_periods(pt_df, ["product_type"], "W")
                        _wagg = _wagg.groupby("period", as_index=False)["actual_kg"].sum().sort_values("period")
                        _wproj = project_forward_with_range(
                            _wagg["actual_kg"].tolist(), None,
                            n_periods=n_periods_shown * 5,  # enough weeks to cover the months shown
                            keep_trend=True,
                            order=_stored[0] if _stored else None,
                            seasonal_period=(_stored[1][3] if _stored and _stored[1][3] else None))
                        if not _wproj.empty:
                            _wlast = pd.Timestamp(_wagg["period"].iloc[-1])
                            _wproj = _wproj.copy()
                            _wproj["_date"] = [_wlast + pd.Timedelta(weeks=i + 1) for i in range(len(_wproj))]
                            _wproj["_m"] = _wproj["_date"].dt.to_period("M").astype(str)
                            _rolled = _wproj.groupby("_m", as_index=False).agg(
                                forecast_kg=("forecast_kg", "sum"), low=("low", "sum"), high=("high", "sum"),
                                _weeks=("forecast_kg", "size"))
                            # drop any trailing month the projection only partly covers
                            _rolled = _rolled[_rolled["_weeks"] >= 4].head(n_periods_shown)
                            _rolled["step"] = range(1, len(_rolled) + 1)
                            projection_pt = _rolled[["step", "forecast_kg", "low", "high"]]
                        else:
                            projection_pt = pd.DataFrame()
                    else:
                        projection_pt = project_forward_with_range(
                            agg_pt["actual_kg"].tolist(), None, n_periods=n_periods_shown,
                            keep_trend=True,
                            order=_stored[0] if _stored else None,
                            seasonal_period=(_stored[1][3] if _stored and _stored[1][3] else None))
                    _detected = detect_seasonal_period(agg_pt["actual_kg"].tolist())
                    if _detected:
                        st.caption(f"Repeating {_detected}-period cycle detected — the forecast follows it.")
                    else:
                        st.caption("No repeating cycle found in this segment's history, so the forecast "
                                   "is smooth. The week-to-week swings here look like noise, and "
                                   "predicting noise makes accuracy worse, not better.")

                    # anchor period 1 to the SAME single-step forecast that feeds the Overview
                    # KPI total -- guarantees this table's first period matches the top of the
                    # dashboard exactly. Uses an ADDITIVE offset, not a proportional rescale of
                    # the whole curve -- real bug found and fixed: a proportional rescale means
                    # a one-time pipeline event (meant to expire after the current month) was
                    # silently leaking its effect, proportionally, all the way out to week 8.
                    # An additive offset to period 1 only keeps a one-time bump contained to the
                    # period it actually applies to, leaving the statistical shape of later
                    # periods untouched.
                    direct_forecast = segment_forecasts.get(pt)
                    if not projection_pt.empty and direct_forecast is not None:
                        # explicit float cast before assigning -- real bug found: pandas 3.0
                        # raises a TypeError instead of silently upcasting a whole-number
                        # (int64) column to decimals, which older pandas did automatically.
                        # These columns can end up all-integer (e.g. all zeros) via the
                        # fallback projection path, especially early on with little history.
                        for col in ["forecast_kg", "low", "high"]:
                            projection_pt[col] = projection_pt[col].astype(float)
                        offset = direct_forecast - projection_pt["forecast_kg"].iloc[0]
                        projection_pt.loc[projection_pt.index[0], "forecast_kg"] = max(direct_forecast, 0)
                        projection_pt.loc[projection_pt.index[0], "low"] = max(projection_pt["low"].iloc[0] + offset, 0)
                        projection_pt.loc[projection_pt.index[0], "high"] = max(projection_pt["high"].iloc[0] + offset, 0)

                    total_company_this_period = sum(segment_forecasts.values()) if segment_forecasts else None
                    if total_company_this_period:
                        st.caption(f"{(direct_forecast or 0)/total_company_this_period*100:.0f}% of next period's total")

                    # simple table: recent actual -> forecast, one continuous timeline
                    recent_actual = agg_pt.tail(n_history_shown)[["period", "actual_kg"]].rename(
                        columns={"period": "Period", "actual_kg": "Actual (kg)"})

                    if pt_horizon == "Week":
                        # same shared function the Overview chart now uses -- previously this
                        # was its own separate inline implementation, which is exactly how the
                        # two views could drift apart again even after being fixed once. One
                        # function, called from both places, is what actually prevents that.
                        wf = walk_forward_segment(pt_df, n_periods=n_history_shown, freq="W")
                        # apply the events that belonged to THIS segment, for the weeks they
                        # were actually live. Missing this is why the segment forecasts summed
                        # to more than the Overview total: the chart and the accuracy KPI had
                        # event adjustment, these tables didn't. Uses the full event record
                        # (including stopped ones) so a week already forecast while an event
                        # was live keeps that number even after the event is turned off.
                        if not wf.empty and not all_events_all.empty:
                            _seg_ev = all_events_all.copy()
                            _ptl = pt_df[["product"]].drop_duplicates()
                            _ptl["_k"] = _ptl["product"].astype(str).str.strip().str.casefold()
                            _seg_ev["_k"] = _seg_ev["product"].astype(str).str.strip().str.casefold()
                            _seg_ev = _seg_ev[_seg_ev["_k"].isin(set(_ptl["_k"]))]
                            if "channel" in pt_df.columns:
                                _chans = set(pt_df["channel"].astype(str).str.strip().str.casefold())
                                _seg_ev = _seg_ev[
                                    _seg_ev["channel"].astype(str).str.strip().str.casefold().isin(_chans)]
                            if not _seg_ev.empty:
                                _wadj = event_adjustment_by_period(_seg_ev, wf["period"].tolist(), freq="W")
                                wf = wf.copy()
                                wf["forecast_kg"] = wf["forecast_kg"] + wf["period"].map(_wadj).fillna(0)
                        stored_by_week = wf.rename(columns={"period": "Period", "forecast_kg": "Forecast (kg)"})
                        recent_actual = recent_actual.merge(stored_by_week, on="Period", how="left")
                    else:
                        recent_actual["Forecast (kg)"] = None

                    fwd_table = projection_pt.copy()
                    last_date = pd.Timestamp(agg_pt["period"].iloc[-1])
                    if pt_horizon == "Week":
                        fwd_table["Period"] = [(last_date + pd.Timedelta(weeks=int(s))).date().isoformat() for s in fwd_table["step"]]
                    else:
                        fwd_table["Period"] = [(last_date + pd.DateOffset(months=int(s))).date().isoformat() for s in fwd_table["step"]]
                    fwd_table["Actual (kg)"] = None
                    fwd_table = fwd_table.rename(columns={"forecast_kg": "Forecast (kg)"})[["Period", "Actual (kg)", "Forecast (kg)"]]

                    # don't duplicate a period that already has a stored forecast merged in above
                    fwd_table = fwd_table[~fwd_table["Period"].isin(recent_actual["Period"])]

                    simple_table = pd.concat([recent_actual, fwd_table], ignore_index=True)
                    simple_table["Actual (kg)"] = simple_table["Actual (kg)"].map(lambda x: f"{x:,.0f}" if pd.notna(x) else "—")
                    simple_table["Forecast (kg)"] = simple_table["Forecast (kg)"].map(lambda x: f"{x:,.0f}" if pd.notna(x) else "—")
                    st.dataframe(simple_table.set_index("Period").T, use_container_width=True)
                    if pt_horizon == "Month":
                        st.caption("Monthly view only tracks live forecasts going forward — historical "
                                   "forecast-vs-actual by month isn't stored yet (weekly view has it).")

                    fig_pt = go.Figure()
                    # focus the x-axis on the last 6 months, same as the Overview chart -- the
                    # full history still feeds the actual forecasting math above, this only
                    # trims what's displayed
                    n_display_pt = 26 if pt_horizon == "Week" else 6
                    agg_pt_display = agg_pt.tail(n_display_pt)
                    fig_pt.add_trace(go.Scatter(x=agg_pt_display["period"], y=agg_pt_display["actual_kg"], mode="lines",
                                                 name="Actual", line=dict(color="rgb(31,119,180)", width=2)))
                    hist_forecast = recent_actual.dropna(subset=["Forecast (kg)"])
                    if not hist_forecast.empty:
                        fig_pt.add_trace(go.Scatter(x=hist_forecast["Period"], y=hist_forecast["Forecast (kg)"],
                                                     mode="lines+markers", name="Forecast (checked against actual)",
                                                     line=dict(color="rgb(139,90,60)", width=2, dash="dash")))
                    if not projection_pt.empty:
                        join_x_pt = [agg_pt["period"].iloc[-1]] + fwd_table["Period"].tolist()
                        join_y_pt = [agg_pt["actual_kg"].iloc[-1]] + projection_pt["forecast_kg"].tolist()
                        fig_pt.add_trace(go.Scatter(x=join_x_pt, y=join_y_pt, mode="lines", name="Forecast (ahead)",
                                                     line=dict(color="rgb(60,60,60)", width=2)))
                    fig_pt.update_layout(height=340, margin=dict(l=10, r=10, t=46, b=10), plot_bgcolor="white",
                                          showlegend=True, xaxis=dict(showgrid=False),
                                          yaxis=dict(showgrid=True, gridcolor="rgba(0,0,0,0.06)"),
                                          legend=dict(orientation="h", yanchor="bottom", y=1.02,
                                                      xanchor="left", x=0, font=dict(size=9)))
                    st.plotly_chart(fig_pt, use_container_width=True)

            st.divider()

        # ===============================================================
        # STAPLE — CHANNEL & BAG SIZE, 3 MONTHS AHEAD (for Operations' bag ordering)
        # ===============================================================
        if "size_label" in sales_df.columns:
            st.markdown("## Bag ordering — every channel & size, 3 months ahead")
            st.caption(
                "Everything Operations needs to place a bag order: all three segments (Single and both "
                "Staple segments), every channel, every bag size, three months out for lead time. Each "
                "segment is projected on its own history, then split by trending channel share and bag-size "
                "share, and converted to real bag counts using the kg-per-bag rate learned from your sales."
            )
            if st.button("Compute bag order breakdown", key="compute_staple_breakdown"):
                st.session_state["show_staple_breakdown"] = True

            if st.session_state.get("show_staple_breakdown"):
                # per-segment adjustment for events + overrides, expressed MONTHLY because
                # this view is monthly. type_level_forecasts_with_pipeline already has both
                # folded in at weekly scale, so the difference from the raw statistical
                # forecast is exactly the adjustment -- scaled up by weeks-per-month.
                _seg_adjust = {}
                if type_level_forecasts:
                    for _lab, _base in type_level_forecasts.items():
                        _adj_weekly = type_level_forecasts_with_pipeline.get(_lab, _base) - _base
                        if abs(_adj_weekly) > 0.01:
                            _seg_adjust[_lab] = _adj_weekly * 4.345
                if _seg_adjust:
                    st.info("This breakdown includes your logged events and manual overrides: "
                            + ", ".join(f"{k} {v:+,.0f} kg/mo" for k, v in _seg_adjust.items()))
                breakdown_df = compute_all_channel_bag_breakdown(
                    sales_df, n_periods=3, freq="M", segment_adjust=_seg_adjust)
                if breakdown_df.empty:
                    st.info("Not enough history yet for this breakdown.")
                else:
                    has_bags = breakdown_df["forecast_bags"].notna().any()
                    if not has_bags:
                        st.warning("Showing kg only — no quantity/bag-count column was mapped on upload, "
                                   "so a real bag count can't be computed. Re-upload with a Quantity column "
                                   "mapped (tab 1) to get bag counts here.")
                    show_unit = "Kg" if not has_bags else st.radio(
                        "Show as", ["Bags (for ordering)", "Kg"], horizontal=True, key="staple_bag_unit")
                    value_col = "forecast_bags" if show_unit.startswith("Bags") else "forecast_kg"

                    st.markdown("**Total bags to order, by size** — the headline number for a purchase order")
                    totals = breakdown_df.dropna(subset=[value_col]).pivot_table(
                        index="size_label", columns="period", values=value_col, aggfunc="sum").round(0)
                    st.dataframe(totals, use_container_width=True)

                    with st.expander("By segment, channel and size"):
                        detail = breakdown_df.dropna(subset=[value_col]).pivot_table(
                            index=["segment", "channel", "size_label"], columns="period",
                            values=value_col, aggfunc="sum").round(0)
                        st.dataframe(detail, use_container_width=True)
                    with st.expander("Kg-per-bag rates used (learned from your actual data)"):
                        st.dataframe(compute_kg_per_bag(sales_df), use_container_width=True, hide_index=True)
                    st.caption("Bag counts are rounded up per line — you can't order a partial bag.")
            st.divider()

        st.markdown("### Filter / break down by")
        st.caption("'Not included' aggregates across every value of that dimension (e.g. item totals combined "
                   "across all channels). 'All' breaks that dimension down into every value (rows in tables, "
                   "bars in charts). Pick one specific value to narrow everything below to just that segment.")
        fc1, fc2, fc3 = st.columns(3)
        channel_options = ["(not included)", "All"] + sorted(sales_df["channel"].unique().tolist())
        item_options = ["(not included)", "All"] + sorted(sales_df["product"].unique().tolist())
        customer_available = "customer" in sales_df.columns and not (sales_df["customer"] == "(not tracked)").all()
        customer_options = ["(not included)", "All"] + sorted(sales_df[sales_df["customer"] != "(not tracked)"]["customer"].unique().tolist()) \
            if customer_available else ["(not included)"]

        sel_channel = fc1.selectbox("Channel", channel_options, key="filt_channel", index=1)
        sel_item = fc2.selectbox("Item", item_options, key="filt_item", index=1)
        sel_customer = fc3.selectbox("Customer", customer_options, key="filt_customer", disabled=not customer_available,
                                      help="'Not included' aggregates across all customers, without breaking down "
                                           "or filtering — 'All' adds it as a breakdown dimension (sparse accounts "
                                           "get filtered out automatically), or pick one specific customer.")
        if not customer_available:
            fc3.caption("Not available in this data source.")

        group_cols = []
        filter_values = {}
        if sel_channel == "All":
            group_cols.append("channel")
        elif sel_channel not in ("(not included)",):
            filter_values["channel"] = sel_channel
        if sel_item == "All":
            group_cols.append("product")
        elif sel_item not in ("(not included)",):
            filter_values["product"] = sel_item
        if sel_customer == "All":
            group_cols.append("customer")
        elif sel_customer not in ("(not included)",):
            filter_values["customer"] = sel_customer

        filtered_df = sales_df.copy()
        for col, val in filter_values.items():
            filtered_df = filtered_df[filtered_df[col] == val]

        specific_share = 1.0
        if filter_values:
            if filtered_df.empty:
                st.warning("No records match this combination — try a different filter.")
                specific_share = 0.0
            else:
                share_base = sales_df[sales_df["customer"] != "(not tracked)"] if "customer" in filter_values else sales_df
                filt_shares = compute_shares(share_base, list(filter_values.keys()))
                match = filt_shares.copy()
                for col, val in filter_values.items():
                    match = match[match[col] == val]
                specific_share = match["share"].sum() if not match.empty else 0.0
                st.caption(f"Filtering to: {', '.join(f'{k}={v}' for k, v in filter_values.items())} "
                           f"— {specific_share*100:.1f}% of total company volume.")

        next_week_kg = next_week_kg_all * specific_share
        next_week_cad = next_week_cad_all * specific_share

        st.divider()

        # ===============================================================
        # DETAILED — everything below responds to the filter above
        # ===============================================================
        st.markdown("## Detailed breakdown")

        # --- forecast ahead ---
        st.markdown("### Forecast ahead")
        st.caption("The total is always computed once, from the full channel × item history — every "
                   "breakdown below is that same total split by real historical share, so they always "
                   "add up to the same number no matter how you slice it.")
        horizon = st.radio("Horizon", ["Next week", "Next month", "Next year"], horizontal=True, key="fwd_horizon")

        if horizon == "Next week":
            # same authoritative number as the Overview KPI above (segment forecasts +
            # pipeline events). Manual overrides are applied further down, per row, since
            # an override targets a specific channel/item rather than the company total.
            canonical_total = next_week_kg_all
        else:
            agg_canonical = aggregate_periods(sales_df, ["channel", "product"], "M")
            canonical_fwd = forecast_next_period(agg_canonical, ["channel", "product"], min_history=2)
            canonical_total = canonical_fwd["forecast_kg"].sum() if not canonical_fwd.empty else 0
            if horizon == "Next year":
                canonical_total = canonical_total * 12
            st.caption("Note: unlike 'Next week', this month/year projection doesn't yet include Pipeline "
                       "events or Manual overrides — it's the raw statistical forecast only.")
        canonical_total = canonical_total * specific_share

        if not group_cols:
            if filter_values:
                st.info("All three are filtered to specific values, so there's nothing left to break down — "
                        "check the KPI cards and Overall Trend above for this exact segment.")
            else:
                st.info("No dimension is set to 'All', so there's nothing to break down into rows — "
                        "set at least one to 'All' to see a breakdown, or check Overview above for the total.")
        elif canonical_total == 0:
            st.info("Not enough history yet to forecast.")
        else:
            needs_customer_fwd = "customer" in group_cols
            if needs_customer_fwd and (not has_data or "customer" not in sales_df.columns
                                        or (sales_df["customer"] == "(not tracked)").all()):
                st.warning("This data source doesn't include customer identity, so a breakdown "
                           "including Customer isn't available.")
                shares = pd.DataFrame()
            else:
                base_df_fwd = filtered_df[filtered_df["customer"] != "(not tracked)"] if needs_customer_fwd else filtered_df
                shares = compute_shares(base_df_fwd, group_cols)

            if shares.empty:
                st.info("Not enough history yet for this breakdown.")
            else:
                shares["forecast_kg"] = (shares["share"] * canonical_total).round(1)

                # Apply manual overrides to the breakdown rows. Real bug this fixes: every
                # row here was purely "historical share x total", so an override set on a
                # specific channel/item never appeared -- you could set OSEE12 to 2000 kg and
                # the table would keep showing its share-derived number as if nothing
                # happened. A row's channel/product is known either because it's one of the
                # grouped dimensions, or because it's pinned by the filter above; when both
                # are known and an override exists for that exact pair, the override wins.
                # add pipeline events to the rows they actually belong to, BEFORE overrides.
                # Real bug: the event's kg was included in the total and then split by
                # HISTORICAL share -- so a brand-new customer's volume (which has no history
                # at all) got smeared across existing segments instead of landing on the one
                # it was logged against. Subtracting it from the share-split base and adding
                # it to the matching rows puts it where it belongs.
                event_applied_rows = 0
                if not applicable.empty and horizon == "Next week":
                    _ev = applicable.copy()
                    _ev["weekly_kg"] = _ev["expected_kg_per_month"] / 4.345
                    for _, ev in _ev.iterrows():
                        for idx, row in shares.iterrows():
                            ok = True
                            for dim in ("channel", "product", "customer"):
                                if dim not in ev or pd.isna(ev.get(dim)):
                                    continue
                                target = str(ev[dim])
                                if dim in group_cols:
                                    actual = str(row[dim])
                                elif dim in filter_values:
                                    actual = str(filter_values[dim])
                                else:
                                    # this dimension isn't visible in the current breakdown --
                                    # e.g. breaking down by channel while the event is tied to
                                    # a specific item. Treat it as "matches anything" rather
                                    # than a failed match: previously the event was skipped
                                    # entirely, so the rows summed to MORE than the KPI total.
                                    continue
                                if actual != target:
                                    ok = False
                                    break
                            if ok:
                                shares.at[idx, "forecast_kg"] = round(
                                    float(shares.at[idx, "forecast_kg"]) + float(ev["weekly_kg"]), 1)
                                event_applied_rows += 1
                                break

                override_applied_rows = 0
                if not active_overrides.empty and horizon == "Next week":
                    def _resolve(row, dim):
                        if dim in group_cols:
                            return str(row[dim])
                        v = filter_values.get(dim)
                        return str(v) if v is not None else None

                    new_vals = []
                    for _, row in shares.iterrows():
                        hit = find_matching_override(
                            active_overrides,
                            channel=_resolve(row, "channel"),
                            product=_resolve(row, "product"),
                            customer=_resolve(row, "customer"))
                        if hit is not None:
                            new_vals.append(round(hit, 1))
                            override_applied_rows += 1
                        else:
                            new_vals.append(row["forecast_kg"])
                    shares["forecast_kg"] = new_vals
                    if override_applied_rows:
                        # the total must reflect the overridden rows, not the pre-override split
                        canonical_total = float(shares["forecast_kg"].sum())

                shares["Segment"] = shares[group_cols].astype(str).agg(" — ".join, axis=1) \
                    if len(group_cols) > 1 else shares[group_cols[0]]
                # hand the on-screen breakdown to the report builder, so the exported report
                # shows the same slice the user is actually looking at rather than a fixed view
                st.session_state["_report_breakdown"] = shares[["Segment", "forecast_kg"]].copy()
                st.session_state["_report_breakdown_label"] = " / ".join(group_cols)
                period_label = {"Next week": "Forecast kg (next week)", "Next month": "Forecast kg (next month)",
                                 "Next year": "Forecast kg (next year, extrapolated)"}[horizon]

                # hard reconciliation check against the Overview KPI. The breakdown is built
                # from shares plus per-row event/override adjustments, and any row that fails
                # to match an adjustment leaves the rows summing to something different from
                # the headline number -- which is exactly the kind of silent disagreement this
                # dashboard has had to hunt down repeatedly.
                if horizon == "Next week" and next_week_kg_all:
                    _row_sum = float(shares["forecast_kg"].sum())
                    _gap = _row_sum - next_week_kg_all
                    if abs(_gap) > max(1.0, next_week_kg_all * 0.005):
                        st.warning(
                            f"These rows sum to {_row_sum:,.0f} kg but the Overview total is "
                            f"{next_week_kg_all:,.0f} kg — a {_gap:+,.0f} kg gap. Usually an event or "
                            "override that couldn't be matched to a row at this breakdown level.")

                st.metric(f"Total — {horizon.lower()}", f"{canonical_total:,.0f} kg")
                if event_applied_rows:
                    st.caption(f"{event_applied_rows} row(s) include a logged pipeline event "
                               "(converted from the monthly figure to a weekly one).")
                if override_applied_rows:
                    st.caption(f"{override_applied_rows} row(s) replaced by an active manual override — "
                               "the total above reflects those overridden values.")
                display_shares = shares.sort_values("forecast_kg", ascending=False)
                st.dataframe(
                    display_shares[["Segment", "forecast_kg"]].rename(columns={"forecast_kg": period_label}),
                    use_container_width=True, hide_index=True)
                if horizon == "Next year":
                    st.caption("Next year isn't independently modeled — not enough history yet (multiple "
                               "full years) for a real year-over-year trend. This is the monthly total × "
                               "12, a simple extrapolation. Treat it as a rough planning figure.")

        if not dollar_by_cp.empty:
            with st.expander("Translated forecast — kg and CAD, by channel and item"):
                display_dollar = dollar_by_cp.copy()
                display_dollar["forecast_kg"] = display_dollar["forecast_kg"].round(0)
                display_dollar["forecast_cad"] = display_dollar["forecast_cad"].map(lambda x: f"${x:,.0f}")
                display_dollar = display_dollar.rename(columns={
                    "channel": "Channel", "product": "Item", "forecast_kg": "Forecast (kg)", "forecast_cad": "Forecast (CAD)"})
                st.dataframe(display_dollar, use_container_width=True, hide_index=True)
                c1, c2 = st.columns(2)
                c1.metric("Combined total — kg", f"{next_week_kg:,.0f} kg")
                c2.metric("Combined total — CAD", f"${next_week_cad:,.0f}")
                st.caption("CAD value is the forecast kg × the real weighted price/kg computed from your "
                           "uploaded sales history — not a manually entered number.")

        st.divider()
        view = st.radio("View", ["Weekly report", "Monthly report"], horizontal=True)

        # Accuracy analysis is opt-in -- it's by far the most expensive computation here
        # (one model fit per segment per week), and Streamlit reruns everything on every
        # click, so running it automatically meant paying that cost constantly.
        if backtest_df.empty:
            st.info("Accuracy analysis hasn't been run yet for this session.")
            bt_weeks = st.slider("Weeks of history to check", 4, 26, 8, key="bt_weeks")
            if st.button("Run accuracy analysis"):
                with st.spinner(f"Checking the last {bt_weeks} weeks..."):
                    st.session_state["backtest_df"] = backtest_accuracy(
                        weekly_actual, max_backtest_weeks=bt_weeks)
                st.rerun()
        else:
            if st.button("Refresh accuracy analysis"):
                st.session_state.pop("backtest_df", None)
                st.rerun()

        if view == "Weekly report":
            st.markdown("**Accuracy overview — every segment at a glance**")
            st.caption(f"Broken down by: {' × '.join(group_cols) if group_cols else '(none — fully filtered to one specific segment)'}")

            if not group_cols:
                if filter_values:
                    st.info("All three are filtered to specific values, so there's nothing left to break down — "
                            "check the KPI cards and Overall Trend above for this exact segment.")
                else:
                    st.info("No dimension is set to 'All', so there's nothing to break down into rows — "
                            "set at least one to 'All' to see a breakdown, or check Overview above for the total.")
                bt = pd.DataFrame()
                cadence_df = pd.DataFrame()
            elif "customer" in group_cols:
                cadence_df = pd.DataFrame()
                if not has_data or "customer" not in sales_df.columns or (sales_df["customer"] == "(not tracked)").all():
                    st.warning("This data source doesn't include customer identity, so any grouping including "
                               "Customer isn't available.")
                    bt = pd.DataFrame()
                else:
                    base_df = sales_df[sales_df["customer"] != "(not tracked)"]
                    weekly_g = compute_weekly_actuals_by(base_df, group_cols)
                    bt = backtest_accuracy(weekly_g, group_cols=group_cols)
                    if not bt.empty:
                        enough_history = bt.groupby(group_cols, as_index=False)["week_start"].count()
                        keep_df = enough_history[enough_history["week_start"] >= 3][group_cols]
                        bt = bt.merge(keep_df, on=group_cols, how="inner")
                        if bt.empty:
                            st.info("No combination has enough order history yet (need 3+ forecastable weeks) "
                                     "at this granularity — try removing Customer or a dimension.")
                        elif group_cols == ["customer"]:
                            cadence_rows = []
                            latest_data_date = pd.to_datetime(sales_df["record_date"], errors="coerce").max()
                            for cust, grp in weekly_g[weekly_g["customer"].isin(bt["customer"].unique())].groupby("customer"):
                                dates = pd.to_datetime(grp["week_start"]).sort_values()
                                gaps = dates.diff().dt.days.dropna() / 7
                                cadence_rows.append({
                                    "customer": cust,
                                    "avg_reorder_weeks": round(gaps.mean(), 1) if len(gaps) else None,
                                    "weeks_since_last_order": round((latest_data_date - dates.max()).days / 7, 1),
                                })
                            cadence_df = pd.DataFrame(cadence_rows)
            else:
                cadence_df = pd.DataFrame()
                bt = backtest_df.copy()
                if group_cols != ["channel", "product"] and not bt.empty:
                    bt = bt.groupby(group_cols + ["week_start"], as_index=False).agg(
                        forecast_kg=("forecast_kg", "sum"), actual_kg=("actual_kg", "sum"),
                        n_weeks_history=("n_weeks_history", "min"))
                    bt["variance_pct"] = (bt["actual_kg"] - bt["forecast_kg"]) / bt["forecast_kg"].replace(0, np.nan)

            if bt.empty:
                st.info("Nothing to show for this grouping yet.")
            else:
                latest_week = bt["week_start"].max()
                overview_rows = []
                for key, grp in bt.groupby(group_cols):
                    grp = grp.sort_values("week_start")
                    label = key if isinstance(key, str) else " — ".join(key)
                    last_row = grp.iloc[-1]
                    recent_bias = grp["variance_pct"].tail(4).mean()
                    weeks_of_history = grp["n_weeks_history"].iloc[-1] if "n_weeks_history" in grp.columns else None
                    if pd.isna(recent_bias):
                        status = "Not enough data"
                    elif abs(recent_bias) > 0.15:
                        status = "ALERT"
                    elif abs(recent_bias) > 0.08:
                        status = "WATCH"
                    else:
                        status = "OK"
                    confidence = "Low (little history)" if (weeks_of_history is not None and weeks_of_history < 4) else "Normal"
                    row = {
                        "Segment": label, "Latest forecast (kg)": round(last_row["forecast_kg"]),
                        "Latest actual (kg)": round(last_row["actual_kg"]),
                        "Recent 4wk bias": f"{recent_bias*100:+.0f}%" if pd.notna(recent_bias) else "n/a",
                        "Confidence": confidence, "Status": status,
                    }
                    if group_cols == ["customer"] and not cadence_df.empty:
                        cad = cadence_df[cadence_df["customer"].astype(str) == str(label)]
                        if not cad.empty:
                            row["Avg reorder (weeks)"] = cad.iloc[0]["avg_reorder_weeks"]
                            row["Weeks since last order"] = cad.iloc[0]["weeks_since_last_order"]
                    overview_rows.append(row)
                overview_df = pd.DataFrame(overview_rows).sort_values(
                    "Status", key=lambda s: s.map({"ALERT": 0, "WATCH": 1, "OK": 2, "Not enough data": 3}))

                def _flag(row):
                    color = {"ALERT": "background-color: #fbeae6", "WATCH": "background-color: #fdf3e0",
                             "OK": "", "Not enough data": ""}.get(row["Status"], "")
                    return [color] * len(row)
                st.dataframe(overview_df.style.apply(_flag, axis=1), use_container_width=True, hide_index=True)
                st.caption(f"As of week of {latest_week}. ALERT = recent actuals off by 15%+ from forecast, "
                           "WATCH = 8-15%. Confidence 'Low' means under 4 weeks of history fed the forecast — "
                           "treat those numbers as rough, not reliable.")

        else:  # Monthly report
            bt = backtest_df.copy()
            if bt.empty:
                st.info("Run the accuracy analysis above to see the monthly report.")
            else:
                bt["month"] = pd.to_datetime(bt["week_start"]).dt.to_period("M").astype(str)
                bt["abs_variance_pct"] = bt["variance_pct"].abs()
                monthly = bt.groupby(["channel", "product", "month"], as_index=False).agg(
                    MAPE=("abs_variance_pct", "mean"), Bias=("variance_pct", "mean"), weeks=("week_start", "count"))
                monthly["MAPE_%"] = (monthly["MAPE"] * 100).round(1)
                monthly["Bias_%"] = (monthly["Bias"] * 100).round(1)
                st.dataframe(monthly[["month", "channel", "product", "MAPE_%", "Bias_%", "weeks"]]
                             .sort_values("month", ascending=False), use_container_width=True)
                st.caption("Positive bias = actuals running ahead of the auto-forecast (under-forecasting). "
                           "Negative = over-forecasting. MAPE = average error size regardless of direction.")

    # --- downloadable snapshot report, for meetings ---
    if has_data and not backtest_df.empty:
        st.divider()
        st.subheader("Download a report")
        st.caption("A self-contained snapshot of the current dashboard — everyone in a meeting can open it, "
                   "no login or app access needed.")

        # shared data prep, used by all three report formats
        gen_time = datetime.now().strftime("%Y-%m-%d %H:%M")
        cap_row2 = pd.read_sql("SELECT * FROM ops_capacity WHERE cycle_label = ? ORDER BY id DESC LIMIT 1",
                                conn, params=(cycle,))
        cap_status_text, cap_shortfall = None, False
        if not cap_row2.empty and not forecast_by_cp.empty:
            cap2 = cap_row2.iloc[0]["monthly_capacity_kg"]
            planned2 = forecast_by_cp["forecast_kg"].sum() * 4.345
            cap_shortfall = cap2 < planned2
            cap_status_text = (f"Capacity check: {planned2:,.0f} kg/month planned vs {cap2:,.0f} kg/month "
                                f"capacity — {'SHORTFALL' if cap_shortfall else 'OK'}")

        report_dollar_df = pd.DataFrame()
        if not dollar_by_cp.empty:
            report_dollar_df = dollar_by_cp.rename(columns={"channel": "Channel", "product": "Item",
                                                              "forecast_kg": "Forecast (kg)", "forecast_cad": "Forecast (CAD)"})

        ov = backtest_df.copy()
        latest_wk = ov["week_start"].max()
        summary_rows = []
        for key, grp in ov.groupby(["channel", "product"]):
            grp = grp.sort_values("week_start")
            bias = grp["variance_pct"].tail(4).mean()
            status = "ALERT" if pd.notna(bias) and abs(bias) > 0.15 else \
                ("WATCH" if pd.notna(bias) and abs(bias) > 0.08 else ("OK" if pd.notna(bias) else "n/a"))
            summary_rows.append({"Channel": key[0], "Item": key[1],
                                  "Latest forecast (kg)": round(grp.iloc[-1]["forecast_kg"]),
                                  "Latest actual (kg)": round(grp.iloc[-1]["actual_kg"]),
                                  "Recent 4wk bias": f"{bias*100:+.0f}%" if pd.notna(bias) else "n/a",
                                  "Status": status})
        report_overview_df = pd.DataFrame(summary_rows).sort_values(
            "Status", key=lambda s: s.map({"ALERT": 0, "WATCH": 1, "OK": 2, "n/a": 3}))

        acc = backtest_df.copy()
        acc["abs_variance_pct"] = acc["variance_pct"].abs()
        acc_monthly = acc.copy()
        acc_monthly["month"] = pd.to_datetime(acc_monthly["week_start"]).dt.to_period("M").astype(str)
        report_accuracy_df = acc_monthly.groupby(["channel", "product"], as_index=False).agg(
            MAPE=("abs_variance_pct", "mean"), Bias=("variance_pct", "mean"), weeks_tracked=("week_start", "count"))
        report_accuracy_df["MAPE_%"] = (report_accuracy_df["MAPE"] * 100).round(1)
        report_accuracy_df["Bias_%"] = (report_accuracy_df["Bias"] * 100).round(1)
        report_accuracy_df = report_accuracy_df.rename(columns={"channel": "Channel", "product": "Item"})[
            ["Channel", "Item", "MAPE_%", "Bias_%", "weeks_tracked"]].sort_values("MAPE_%", ascending=False)

        # trend chart data (actual, all history) for embedding
        d_report_trend = sales_df.copy()
        d_report_trend["record_date"] = pd.to_datetime(d_report_trend["record_date"], errors="coerce")
        d_report_trend = d_report_trend.dropna(subset=["record_date"])
        d_report_trend["period"] = (d_report_trend["record_date"] -
                                     pd.to_timedelta(d_report_trend["record_date"].dt.weekday, unit="D")).dt.date.astype(str)
        report_trend_df = d_report_trend.groupby("period", as_index=False)["kg"].sum().sort_values("period")

        report_bar_df = dollar_by_cp.copy() if not dollar_by_cp.empty else pd.DataFrame()

        def build_report_html():
            cap_html = ""
            if cap_status_text:
                color2 = "#b3432f" if cap_shortfall else "#4a7a5c"
                cap_html = f'<p style="color:{color2};font-weight:600">{cap_status_text}</p>'
            # report is intentionally KPI + charts only -- the detail tables were dropped so
            # it reads as a one-glance summary for a meeting rather than a data dump

            # match the Overview chart: a blue actual line (no fill), named series, last 6
            # months only. The filled brown area over 2+ years was unreadable and the legend
            # said "trace 0".
            _rt = report_trend_df.tail(26)
            trend_fig = go.Figure(go.Scatter(x=_rt["period"], y=_rt["kg"], mode="lines",
                                              name="Actual", line=dict(color="#1f77b4", width=2)))
            trend_fig.update_layout(height=340, margin=dict(l=10, r=10, t=40, b=10),
                                     plot_bgcolor="white",
                                     xaxis_title="Week", yaxis_title="Total kg",
                                     yaxis=dict(showgrid=True, gridcolor="rgba(0,0,0,0.06)"),
                                     legend=dict(orientation="h", yanchor="bottom", y=1.02,
                                                 xanchor="left", x=0))
            trend_chart_html = trend_fig.to_html(include_plotlyjs="cdn", full_html=False)

            # overlay the backtested forecast line, exactly as the Overview chart shows it,
            # so the report is the same picture rather than a different-looking summary
            try:
                _rep_bt = walk_forward_all_segments(sales_df, n_periods=26, freq="W")
                if not _rep_bt.empty:
                    _rep_adj = event_adjustment_by_period(all_events_all, _rep_bt["period"].tolist(), freq="W")
                    _rep_bt = _rep_bt.copy()
                    _rep_bt["forecast_kg"] = _rep_bt["forecast_kg"] + _rep_bt["period"].map(_rep_adj).fillna(0)
                    trend_fig.add_trace(go.Scatter(
                        x=_rep_bt[_rep_bt["period"].isin(_rt["period"])]["period"],
                        y=_rep_bt[_rep_bt["period"].isin(_rt["period"])]["forecast_kg"], mode="lines",
                        name="Auto forecast (backtested)", line=dict(color="#8b5a3c", dash="dash")))
                    trend_fig.update_layout(showlegend=True, legend=dict(
                        orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0))
                    trend_chart_html = trend_fig.to_html(include_plotlyjs="cdn", full_html=False)
            except Exception:
                pass  # report should still generate if the backtest can't be computed

            # the three segment charts, same as the Overview page
            seg_charts_html = ""
            try:
                for _lab, _sdf in split_into_segments(sales_df).items():
                    _sagg = aggregate_periods(_sdf, ["product_type"], "W")
                    _sagg = _sagg.groupby("period", as_index=False)["actual_kg"].sum().sort_values("period").tail(26)
                    if len(_sagg) < 2:
                        continue
                    _sfig = go.Figure(go.Scatter(x=_sagg["period"], y=_sagg["actual_kg"], mode="lines",
                                                  name="Actual", line=dict(color="#1f77b4", width=2)))
                    _sproj = project_forward_with_range(_sagg["actual_kg"].tolist(), None,
                                                         n_periods=8, keep_trend=True)
                    if not _sproj.empty:
                        _last = pd.Timestamp(_sagg["period"].iloc[-1])
                        _fx = [(_last + pd.Timedelta(weeks=i + 1)).date().isoformat()
                               for i in range(len(_sproj))]
                        _sfig.add_trace(go.Scatter(x=_fx, y=_sproj["forecast_kg"], mode="lines",
                                                    name="Forecast (ahead)", line=dict(color="#555", width=2)))
                    _sfig.update_layout(height=260, margin=dict(l=10, r=10, t=40, b=10),
                                         plot_bgcolor="white", yaxis_title="kg",
                                         yaxis=dict(showgrid=True, gridcolor="rgba(0,0,0,0.06)"),
                                         legend=dict(orientation="h", yanchor="bottom", y=1.02,
                                                     xanchor="left", x=0))
                    seg_charts_html += f"<h3 style='font-size:14px;margin-top:1.2rem'>{_lab}</h3>"
                    seg_charts_html += _sfig.to_html(include_plotlyjs=False, full_html=False)
            except Exception:
                pass

            # the breakdown exactly as it's filtered on screen, not a fixed top-15 bar chart
            bar_chart_html = ""
            _bd = st.session_state.get("_report_breakdown")
            if _bd is not None and not _bd.empty:
                bar_sorted = _bd.sort_values("forecast_kg", ascending=True).tail(15)
                bar_fig = go.Figure(go.Bar(x=bar_sorted["forecast_kg"], y=bar_sorted["Segment"].astype(str),
                                            orientation="h", marker_color="#2F6F6B"))
                bar_fig.update_layout(height=max(280, 26 * len(bar_sorted)),
                                       margin=dict(l=10, r=10, t=10, b=10), xaxis_title="Forecast kg")
                bar_chart_html = bar_fig.to_html(include_plotlyjs=False, full_html=False)
                bar_chart_html += _bd[["Segment", "forecast_kg"]].rename(
                    columns={"forecast_kg": "Forecast kg"}).to_html(index=False, border=0)

            return f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<title>Demand Planning Report — {cycle}</title>
<style>
body{{font-family:-apple-system,Arial,sans-serif;max-width:900px;margin:2rem auto;color:#2b2622;padding:0 1.5rem}}
h1{{font-size:22px;margin-bottom:4px}} h2{{font-size:16px;margin-top:2rem}}
table{{width:100%;border-collapse:collapse;font-size:13px;margin-top:0.5rem}}
th,td{{text-align:left;padding:6px 10px;border-bottom:1px solid #e3ddd1}}
th{{background:#f3efe8}} .meta{{color:#6b6258;font-size:13px}}
</style></head><body>
<h1>Demand Planning Report</h1>
<p class="meta">Cycle {cycle} — generated {gen_time} — as of week of {latest_wk}</p>
{cap_html}
<h2>Overall trend — actual sales</h2>
{trend_chart_html}
<h2>Forecast by segment</h2>
{seg_charts_html if seg_charts_html else "<p>Not enough history per segment yet.</p>"}
<h2>Forecast breakdown — by {st.session_state.get("_report_breakdown_label", "segment")}</h2>
{bar_chart_html if bar_chart_html else "<p>Open the Dashboard breakdown first, then generate the report.</p>"}
<p class="meta">Generated automatically from 49th Parallel's demand planning app.</p>
</body></html>"""

        def build_report_pdf():
            def safe(t):
                t = str(t)
                for a, b in [("\u2014", "-"), ("\u2013", "-"), ("\u2018", "'"), ("\u2019", "'"),
                             ("\u201c", '"'), ("\u201d", '"'), ("\u2026", "...")]:
                    t = t.replace(a, b)
                return t.encode("latin-1", "replace").decode("latin-1")

            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 16)
            pdf.cell(0, 10, "Demand Planning Report", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 9)
            pdf.cell(0, 6, safe(f"Cycle {cycle} -- generated {gen_time} -- as of week of {latest_wk}"),
                     new_x="LMARGIN", new_y="NEXT")
            if cap_status_text:
                pdf.ln(2)
                pdf.set_text_color(179, 67, 47) if cap_shortfall else pdf.set_text_color(74, 122, 92)
                pdf.set_font("Helvetica", "B", 10)
                pdf.multi_cell(0, 6, safe(cap_status_text))
                pdf.set_text_color(0, 0, 0)

            def draw_table(title, df, col_widths):
                pdf.ln(4)
                pdf.set_font("Helvetica", "B", 12)
                pdf.cell(0, 8, safe(title), new_x="LMARGIN", new_y="NEXT")
                if df.empty:
                    pdf.set_font("Helvetica", "", 9)
                    pdf.cell(0, 6, "No data available.", new_x="LMARGIN", new_y="NEXT")
                    return
                pdf.set_font("Helvetica", "B", 8)
                for w, h in zip(col_widths, df.columns):
                    pdf.cell(w, 7, safe(str(h))[:28], border=1)
                pdf.ln()
                pdf.set_font("Helvetica", "", 8)
                for _, r in df.iterrows():
                    for w, v in zip(col_widths, r):
                        pdf.cell(w, 6, safe(str(v))[:30], border=1)
                    pdf.ln()

            def draw_line_chart(title, series_list, width=180, height=70):
                """Hand-drawn multi-series line chart -- fpdf has no charting and rendering a
                plotly image would need a headless browser, which isn't available here. Same
                content as the HTML report's chart, drawn with primitives."""
                pdf.ln(4)
                pdf.set_font("Helvetica", "B", 12)
                pdf.cell(0, 8, safe(title), new_x="LMARGIN", new_y="NEXT")
                _all = [v for _, ys, _ in series_list for v in ys if v is not None]
                if not _all:
                    pdf.set_font("Helvetica", "", 9)
                    pdf.cell(0, 6, "No data available.", new_x="LMARGIN", new_y="NEXT")
                    return
                lo, hi = min(_all), max(_all)
                span = (hi - lo) or 1
                x0, y0 = pdf.get_x(), pdf.get_y()
                pdf.set_draw_color(210, 210, 210)
                pdf.rect(x0, y0, width, height)
                n_max = max(len(ys) for _, ys, _ in series_list)
                for name, ys, rgb in series_list:
                    pdf.set_draw_color(*rgb)
                    pts = [(x0 + (i / max(n_max - 1, 1)) * width,
                            y0 + height - ((v - lo) / span) * height)
                           for i, v in enumerate(ys) if v is not None]
                    for a, b in zip(pts, pts[1:]):
                        pdf.line(a[0], a[1], b[0], b[1])
                pdf.set_y(y0 + height + 2)
                pdf.set_font("Helvetica", "", 7)
                pdf.set_text_color(90, 90, 90)
                pdf.cell(0, 4, safe("  |  ".join(f"{n}" for n, _, _ in series_list)
                                    + f"   (range {lo:,.0f} - {hi:,.0f} kg)"),
                         new_x="LMARGIN", new_y="NEXT")
                pdf.set_text_color(0, 0, 0)
                pdf.set_draw_color(0, 0, 0)

            def draw_bar_chart(title, labels, values, chart_width=180, bar_height=6, gap=2):
                """Hand-drawn horizontal bar chart -- no image rendering, no Chrome needed."""
                pdf.ln(4)
                pdf.set_font("Helvetica", "B", 12)
                pdf.cell(0, 8, safe(title), new_x="LMARGIN", new_y="NEXT")
                if not len(values):
                    pdf.set_font("Helvetica", "", 9)
                    pdf.cell(0, 6, "No data available.", new_x="LMARGIN", new_y="NEXT")
                    return
                max_val = max(values) if max(values) > 0 else 1
                label_width = 65
                bar_area = chart_width - label_width - 20
                x0 = pdf.get_x()
                for label, val in zip(labels, values):
                    y0 = pdf.get_y()
                    pdf.set_font("Helvetica", "", 7)
                    pdf.cell(label_width, bar_height, safe(str(label))[:38], new_x="LMARGIN", new_y="TOP")
                    bar_len = max(1, (val / max_val) * bar_area)
                    pdf.set_fill_color(139, 90, 60)
                    pdf.rect(x0 + label_width, y0, bar_len, bar_height, style="F")
                    pdf.set_xy(x0 + label_width + bar_len + 2, y0)
                    pdf.set_font("Helvetica", "", 7)
                    pdf.cell(20, bar_height, f"{val:,.0f}", new_x="LMARGIN", new_y="NEXT")
                    pdf.set_x(x0)
                    pdf.set_y(y0 + bar_height + gap)

            # same two charts as the HTML report: actual-vs-backtested trend, then the
            # on-screen breakdown. Tables intentionally omitted -- this report is charts only.
            if not report_trend_df.empty:
                _tr = report_trend_df.tail(26)
                _series = [("Actual", _tr["kg"].tolist(), (31, 119, 180))]
                try:
                    _pbt = walk_forward_all_segments(sales_df, n_periods=26, freq="W")
                    if not _pbt.empty:
                        _padj = event_adjustment_by_period(all_events_all, _pbt["period"].tolist(), freq="W")
                        _pbt = _pbt.copy()
                        _pbt["forecast_kg"] = _pbt["forecast_kg"] + _pbt["period"].map(_padj).fillna(0)
                        _m = _tr[["period"]].merge(_pbt, on="period", how="left")
                        _series.append(("Auto forecast (backtested)",
                                        _m["forecast_kg"].tolist(), (139, 90, 60)))
                except Exception:
                    pass
                draw_line_chart("Overall trend - actual vs forecast", _series)

            _pbd = st.session_state.get("_report_breakdown")
            if _pbd is not None and not _pbd.empty:
                _bt2 = _pbd.sort_values("forecast_kg", ascending=False).head(12)
                draw_bar_chart(
                    f"Forecast breakdown - by {st.session_state.get('_report_breakdown_label', 'segment')}",
                    _bt2["Segment"].astype(str).tolist(), _bt2["forecast_kg"].tolist())

            pdf.add_page()
            draw_table("Forecast accuracy - MAPE and bias", report_accuracy_df, [45, 55, 25, 25, 30])

            pdf.ln(6)
            pdf.set_font("Helvetica", "I", 8)
            pdf.set_text_color(107, 98, 88)
            pdf.multi_cell(0, 5, "ALERT = actuals off by 15%+ from forecast over the last 4 weeks. WATCH = 8-15%. "
                                 "MAPE = average error size regardless of direction. Bias: positive = "
                                 "under-forecasting, negative = over-forecasting. "
                                 "Generated automatically from 49th Parallel's demand planning app.")
            return bytes(pdf.output())

        def build_report_excel():
            from openpyxl.chart import BarChart, LineChart, Reference

            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = "Summary"
            ws1["A1"] = "Demand Planning Report"
            ws1["A1"].font = Font(bold=True, size=14)
            ws1["A2"] = f"Cycle {cycle} — generated {gen_time} — as of week of {latest_wk}"
            if cap_status_text:
                ws1["A4"] = cap_status_text
                ws1["A4"].font = Font(bold=True, color="B3432F" if cap_shortfall else "4A7A5C")

            def write_df(ws, df, start_row=1):
                for j, col in enumerate(df.columns, start=1):
                    c = ws.cell(row=start_row, column=j, value=col)
                    c.font = Font(bold=True, color="FFFFFF")
                    c.fill = PatternFill("solid", fgColor="1F4E78")
                for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
                    for j, val in enumerate(row, start=1):
                        ws.cell(row=i, column=j, value=val)
                for j, col in enumerate(df.columns, start=1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = max(14, len(str(col)) + 2)

            # trend chart -- native Excel line chart, built from real data written into the sheet
            ws_trend = wb.create_sheet("Overall Trend")
            trend_for_excel = report_trend_df.rename(columns={"period": "Week", "kg": "Actual kg"})
            # add the backtested forecast as a second series, so Excel shows the same
            # actual-vs-forecast picture as the HTML and PDF reports
            try:
                _xbt = walk_forward_all_segments(sales_df, n_periods=26, freq="W")
                if not _xbt.empty:
                    _xadj = event_adjustment_by_period(all_events_all, _xbt["period"].tolist(), freq="W")
                    _xbt = _xbt.copy()
                    _xbt["forecast_kg"] = _xbt["forecast_kg"] + _xbt["period"].map(_xadj).fillna(0)
                    trend_for_excel = trend_for_excel.merge(
                        _xbt.rename(columns={"period": "Week", "forecast_kg": "Auto forecast kg"}),
                        on="Week", how="left")
            except Exception:
                pass
            write_df(ws_trend, trend_for_excel)
            if len(trend_for_excel) > 1:
                chart1 = LineChart()
                chart1.title = "Overall trend - actual vs forecast"
                chart1.y_axis.title = "Total kg"
                chart1.x_axis.title = "Week"
                _ncols = len(trend_for_excel.columns)
                data_ref = Reference(ws_trend, min_col=2, max_col=_ncols,
                                      min_row=1, max_row=len(trend_for_excel) + 1)
                cats_ref = Reference(ws_trend, min_col=1, min_row=2, max_row=len(trend_for_excel) + 1)
                chart1.add_data(data_ref, titles_from_data=True)
                chart1.set_categories(cats_ref)
                chart1.width, chart1.height = 26, 11
                ws_trend.add_chart(chart1, f"{openpyxl.utils.get_column_letter(_ncols + 2)}2")

            # breakdown chart -- the same on-screen slice the other two formats show
            _xbd = st.session_state.get("_report_breakdown")
            if _xbd is not None and not _xbd.empty:
                ws_bd = wb.create_sheet("Breakdown")
                _bd_x = _xbd.sort_values("forecast_kg", ascending=False).rename(
                    columns={"forecast_kg": "Forecast kg"})
                write_df(ws_bd, _bd_x)
                chart_bd = BarChart()
                chart_bd.type = "bar"
                chart_bd.title = f"Forecast breakdown - by {st.session_state.get('_report_breakdown_label', 'segment')}"
                chart_bd.x_axis.title = "Forecast kg"
                d_ref = Reference(ws_bd, min_col=2, min_row=1, max_row=len(_bd_x) + 1)
                c_ref = Reference(ws_bd, min_col=1, min_row=2, max_row=len(_bd_x) + 1)
                chart_bd.add_data(d_ref, titles_from_data=True)
                chart_bd.set_categories(c_ref)
                chart_bd.width, chart_bd.height = 24, max(10, 0.5 * len(_bd_x))
                ws_bd.add_chart(chart_bd, "D2")

            ws2 = wb.create_sheet("Translated Forecast")
            if not report_dollar_df.empty:
                bar_data = report_dollar_df.sort_values("Forecast (kg)", ascending=False).copy()
                bar_data.insert(0, "Segment", bar_data["Channel"] + " - " + bar_data["Item"])
                write_df(ws2, bar_data)
                chart2 = BarChart()
                chart2.type = "bar"
                chart2.title = "Forecast by segment (kg)"
                data_ref2 = Reference(ws2, min_col=4, min_row=1, max_row=len(bar_data) + 1)
                cats_ref2 = Reference(ws2, min_col=1, min_row=2, max_row=len(bar_data) + 1)
                chart2.add_data(data_ref2, titles_from_data=True)
                chart2.set_categories(cats_ref2)
                chart2.width, chart2.height = 24, 12
                ws2.add_chart(chart2, "G2")

            ws3 = wb.create_sheet("Accuracy Overview")
            write_df(ws3, report_overview_df)

            ws4 = wb.create_sheet("Accuracy - MAPE and Bias")
            write_df(ws4, report_accuracy_df)

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        # Reports are built ONLY when asked for. st.download_button computes its data
        # eagerly, so having three of these unguarded meant building a full HTML report, a
        # full PDF, AND a full Excel workbook (with charts) on every single page load --
        # even though nobody had clicked anything. Easily one of the most expensive things
        # in the app, and completely invisible as a cost.
        report_kind = st.radio("Report format", ["HTML", "PDF", "Excel"], horizontal=True, key="report_kind")
        if st.button("Generate report"):
            if report_kind == "HTML":
                st.download_button("Download report (HTML)", build_report_html(),
                                    f"demand_report_{cycle}.html", mime="text/html")
            elif report_kind == "PDF":
                st.download_button("Download report (PDF)", build_report_pdf(),
                                    f"demand_report_{cycle}.pdf", mime="application/pdf")
            else:
                st.download_button("Download report (Excel)", build_report_excel(),
                                    f"demand_report_{cycle}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.caption("HTML opens in any browser and can be printed to PDF from there too. "
                   "PDF and Excel are generated directly, ready to attach or print for a meeting.")

# --- TAB 1: Upload ---
with tab_data:
    st.subheader("Upload raw sales records")
    st.caption(
        "One row per sale line: channel, customer (optional), product, size, kg sold, revenue, and date. "
        "Export this from Lightspeed/Acumatica and upload as CSV. Uploading new data automatically "
        "updates the forecast and checks last week's prediction."
    )
    # what's already in the database -- so you can see at a glance what period is covered
    # and spot a missing or short month before it quietly skews a forecast
    if has_data:
        _p = sales_df.copy()
        _p["record_date"] = pd.to_datetime(_p["record_date"], errors="coerce")
        _p = _p.dropna(subset=["record_date"])
        if not _p.empty:
            _lo, _hi = _p["record_date"].min(), _p["record_date"].max()
            m1, m2, m3 = st.columns(3)
            m1.metric("Rows in database", f"{len(_p):,}")
            m2.metric("Earliest record", _lo.strftime("%b %d, %Y"))
            m3.metric("Latest record", _hi.strftime("%b %d, %Y"))
            # surface duplicate batches -- if the same file went in twice, every downstream
            # number is inflated and it is very hard to spot from the dashboard alone
            _dupe_check = _p.groupby("upload_batch", as_index=False).agg(
                rows=("kg", "size"), total_kg=("kg", "sum"),
                first_date=("record_date", "min"), last_date=("record_date", "max"))
            _dupe_check["signature"] = (_dupe_check["rows"].astype(str) + " rows | "
                                        + _dupe_check["first_date"].dt.strftime("%Y-%m-%d") + " to "
                                        + _dupe_check["last_date"].dt.strftime("%Y-%m-%d"))
            _sig_counts = _dupe_check["signature"].value_counts()
            _suspect = _dupe_check[_dupe_check["signature"].isin(_sig_counts[_sig_counts > 1].index)]
            if not _suspect.empty:
                st.error(f"Possible duplicate uploads: {len(_suspect)} batches cover the exact same "
                         "date range with the same row count. If the same file was uploaded twice, "
                         "every forecast built on it is overstated — delete the extra batch below.")
                st.dataframe(_suspect[["upload_batch", "rows", "total_kg", "signature"]]
                             .rename(columns={"upload_batch": "Batch", "rows": "Rows",
                                              "total_kg": "Total kg", "signature": "Covers"}),
                             use_container_width=True, hide_index=True)

            with st.expander("Records by month — check for gaps or short months"):
                _by_month = _p.assign(Month=_p["record_date"].dt.to_period("M").astype(str)) \
                    .groupby("Month", as_index=False).agg(Rows=("kg", "size"), Total_kg=("kg", "sum"))
                _by_month["Total_kg"] = _by_month["Total_kg"].round(0)
                _by_month = _by_month.sort_values("Month", ascending=False)
                st.dataframe(_by_month.rename(columns={"Total_kg": "Total kg"}),
                             use_container_width=True, hide_index=True)
                st.caption("A month with unusually few rows or low kg is often a partial upload — "
                           "worth checking before trusting a forecast built on it.")
        st.divider()

    uploaded = st.file_uploader("CSV or Excel file", type=["csv", "xlsx", "xls"])

    if uploaded is not None:
        if uploaded.name.lower().endswith((".xlsx", ".xls")):
            raw = pd.read_excel(uploaded)
        else:
            raw = pd.read_csv(uploaded)
        st.write("Preview:")
        st.dataframe(raw.head(), use_container_width=True)

        st.markdown("**Map your columns**")
        cols = list(raw.columns)
        col_defaults = load_upload_column_defaults()
        st.caption("Remembers your choices from last time — just confirm they still look right." if col_defaults else "")
        c1, c2, c3 = st.columns(3)
        with c1:
            date_col = st.selectbox("Date column", cols, index=default_index(cols, col_defaults.get("date_col")),
                                     help="Required now — the forecast is built from real dates.")
            channel_col = st.selectbox("Channel column", cols, index=default_index(cols, col_defaults.get("channel_col")))
            pt_options = ["(not available — derive it automatically)"] + cols
            product_type_col = st.selectbox(
                "Single vs Staple column", pt_options,
                index=default_index(pt_options, col_defaults.get("product_type_col")),
                help="Optional — lets you break the forecast down by Single vs Staple, and run each "
                     "with its own forecast (matching your earlier SARIMA coursework: staple ~6.4% "
                     "MAPE, single-origin ~2.9% MAPE, on their own history). Real Acumatica exports "
                     "usually don't have this column directly — pick 'derive it automatically' and "
                     "the app will work it out from the Product column instead.")
        with c2:
            cust_options = ["(not available)"] + cols
            customer_col = st.selectbox(
                "Customer column", cust_options, index=default_index(cust_options, col_defaults.get("customer_col")),
                help="Optional — many Acumatica exports don't include this.")
            product_col = st.selectbox("Product column", cols, index=default_index(cols, col_defaults.get("product_col")))
        with c3:
            size_col = st.selectbox("Size / package column", cols, index=default_index(cols, col_defaults.get("size_col")))
            revenue_col = st.selectbox("Revenue ($) column", cols, index=default_index(cols, col_defaults.get("revenue_col")))

        kg_mode_options = ["I have a direct KG column", "I have Units + Weight-per-unit (kg)"]
        kg_mode = st.radio("How is weight recorded?", kg_mode_options,
                            index=default_index(kg_mode_options, col_defaults.get("kg_mode")))
        if kg_mode == "I have a direct KG column":
            kg_col = st.selectbox("KG column", cols, index=default_index(cols, col_defaults.get("kg_col")))
            qty_options = ["(not available)"] + cols
            quantity_col = st.selectbox(
                "Quantity / bag count column (optional)", qty_options,
                index=default_index(qty_options, col_defaults.get("quantity_col")),
                help="How many units/bags each row represents — used to work out real kg-per-bag "
                     "rates, so forecasts can be converted into 'how many bags to order', not just kg.")
        else:
            units_col = st.selectbox("Units column", cols, index=default_index(cols, col_defaults.get("units_col")))
            weight_col = st.selectbox("Weight per unit (kg) column", cols, index=default_index(cols, col_defaults.get("weight_col")))
            quantity_col = units_col  # already exactly what's needed -- don't ask twice

        needs_auto_classify = product_type_col == "(not available — derive it automatically)"
        manual_overrides_by_item = {}
        classification_hint_col = None
        if needs_auto_classify:
            st.markdown("**Single vs Staple — classify by product, remembered across uploads**")
            st.caption(
                "Classification is keyed to the Product column above (e.g. an Inventory ID like "
                "'EE12') — that's what gets stored and needs a Staple/Single label going forward. "
                "But that field is often just a short code without useful words in it, so pick a "
                "different column below (like Item Class) that actually contains classification "
                "hints — this is only used to help detect the label, not stored as the product itself."
            )
            classification_hint_col = st.selectbox(
                "Which column has classification hints (e.g. Item Class)?",
                ["(use the Product column itself)"] + cols, key="classification_hint_col")
            hint_source = raw[product_col] if classification_hint_col == "(use the Product column itself)" \
                else raw[classification_hint_col]

            known_map = load_known_classifications()
            preview_products = raw[product_col].dropna().unique().tolist()
            audit_rows = []
            newly_learned = []
            for p in preview_products:
                if p in known_map:
                    audit_rows.append({"Product": p, "Classification": known_map[p], "Source": "Remembered from before"})
                else:
                    hint_val = raw.loc[raw[product_col] == p, classification_hint_col
                                        if classification_hint_col != "(use the Product column itself)" else product_col].iloc[0]
                    auto = classify_product_type_auto(hint_val)
                    if auto != "Unknown":
                        audit_rows.append({"Product": p, "Classification": auto, "Source": "Auto-detected (new)"})
                        newly_learned.append((p, auto))
                    else:
                        audit_rows.append({"Product": p, "Classification": "Unknown", "Source": "Needs manual input"})

            audit_df = pd.DataFrame(audit_rows)
            st.dataframe(audit_df, use_container_width=True, hide_index=True)
            unknown_items = audit_df[audit_df["Classification"] == "Unknown"]["Product"].tolist()

            if unknown_items:
                st.warning(f"{len(unknown_items)} product(s) couldn't be automatically classified — "
                           f"assign them once below, and this app will remember your answer for every "
                           f"future upload of the same product.")
                for item in unknown_items:
                    manual_overrides_by_item[item] = st.selectbox(
                        f"'{item}' is:", ["Staple", "Single"], key=f"manual_classify_{item}")
            else:
                st.success("Every product is classified — remembered from before, or auto-detected just now.")

        batch_name = st.text_input("Name this upload batch", value=f"upload_{datetime.now().strftime('%Y%m%d_%H%M')}")

        if st.button("Process and save", type="primary"):
            # Guard against double-inserting the same batch. Real bug: nothing stopped the
            # same file being written twice -- a double-click, or a rerun landing before the
            # first insert finished, would duplicate every row. Two layers: a session flag
            # (catches a fast second click within the same session) and a database check
            # (catches a repeat after a reboot, where session state is gone).
            _inflight = st.session_state.get("_upload_inflight")
            _existing = pd.read_sql(
                "SELECT COUNT(*) AS n FROM sales_records WHERE upload_batch = ?",
                conn, params=(batch_name,))
            _already = int(_existing["n"].iloc[0]) if not _existing.empty else 0

            if _inflight == batch_name:
                st.warning("That upload is already being processed — give it a moment rather than "
                           "clicking again.")
                st.stop()
            if _already > 0:
                st.error(f"A batch named '{batch_name}' already exists with {_already:,} rows. "
                         "Rename this batch if it's genuinely new data, or delete the existing "
                         "one first — importing it again would duplicate every row.")
                st.stop()

            st.session_state["_upload_inflight"] = batch_name
            std = pd.DataFrame()
            std["record_date"] = raw[date_col].astype(str)
            std["channel"] = raw[channel_col].astype(str)
            std["customer"] = raw[customer_col].astype(str) if customer_col != "(not available)" else "(not tracked)"
            std["product"] = raw[product_col].astype(str)
            if needs_auto_classify:
                known_map = load_known_classifications()
                effective_hint_col = product_col if classification_hint_col == "(use the Product column itself)" else classification_hint_col

                def resolve_and_remember(p, hint_val):
                    p = str(p)
                    if p in known_map:
                        return known_map[p]
                    if p in manual_overrides_by_item:
                        save_classification(p, manual_overrides_by_item[p], "manual")
                        return manual_overrides_by_item[p]
                    auto = classify_product_type_auto(hint_val)
                    if auto != "Unknown":
                        save_classification(p, auto, "auto")
                        return auto
                    return "(not tracked)"

                std["product_type"] = [resolve_and_remember(p, h) for p, h in zip(raw[product_col], raw[effective_hint_col])]
            else:
                std["product_type"] = raw[product_type_col].astype(str)
            std["size_label"] = raw[size_col].astype(str)
            std["revenue"] = pd.to_numeric(raw[revenue_col], errors="coerce")
            if kg_mode == "I have a direct KG column":
                std["kg"] = pd.to_numeric(raw[kg_col], errors="coerce")
                std["quantity"] = pd.to_numeric(raw[quantity_col], errors="coerce") if quantity_col != "(not available)" else np.nan
            else:
                std["kg"] = pd.to_numeric(raw[units_col], errors="coerce") * pd.to_numeric(raw[weight_col], errors="coerce")
                std["quantity"] = pd.to_numeric(raw[quantity_col], errors="coerce")
            std = std.dropna(subset=["kg", "revenue"])
            std["upload_batch"] = batch_name
            std["uploaded_at"] = datetime.now().isoformat()
            insert_dataframe("sales_records", std, show_progress=True)

            mapping_to_remember = {
                "date_col": date_col, "channel_col": channel_col, "product_type_col": product_type_col,
                "customer_col": customer_col, "product_col": product_col, "size_col": size_col,
                "revenue_col": revenue_col, "kg_mode": kg_mode,
            }
            if kg_mode == "I have a direct KG column":
                mapping_to_remember.update({"kg_col": kg_col, "quantity_col": quantity_col})
            else:
                mapping_to_remember.update({"units_col": units_col, "weight_col": weight_col})
            save_upload_column_defaults(mapping_to_remember)

            reset_all_derived_state()
            st.session_state.pop("_upload_inflight", None)  # clear the guard once safely written
            st.success(f"Saved {len(std)} records from batch '{batch_name}'. Forecast will update below.")
            st.rerun()

    st.divider()
    st.caption(f"Total records in database: {len(sales_df)}")
    if has_data:
        batches = sales_df["upload_batch"].unique().tolist()
        st.write("Batches uploaded so far:", ", ".join(batches))

        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Delete a specific batch**")
            batch_to_delete = st.selectbox("Choose a batch", batches)
            if st.button("Delete this batch"):
                conn.execute("DELETE FROM sales_records WHERE upload_batch = ?", (batch_to_delete,))
                conn.commit()
                reset_all_derived_state()

                # real cleanup, not just a disclosed limitation -- figure out what the latest
                # actual week is NOW that this batch is gone, and remove any frozen forecasts
                # that were generated assuming data beyond that point existed. Without this,
                # a forecast frozen for the week right after the deleted batch stays sitting
                # in the database, and the app keeps showing/labeling around it as if that
                # batch were still there -- exactly the inconsistency this fixes.
                remaining_weekly = pd.read_sql("SELECT record_date, kg FROM sales_records", conn)
                if not remaining_weekly.empty:
                    remaining_weekly["record_date"] = pd.to_datetime(remaining_weekly["record_date"], errors="coerce")
                    remaining_weekly = remaining_weekly.dropna(subset=["record_date"])
                if remaining_weekly.empty:
                    conn.execute("DELETE FROM auto_forecasts")
                    removed_note = "No sales data remains, so all forecasts were cleared too."
                else:
                    remaining_weekly["week_start"] = (remaining_weekly["record_date"] -
                        pd.to_timedelta(remaining_weekly["record_date"].dt.weekday, unit="D")).dt.date.astype(str)
                    new_latest_week = sorted(remaining_weekly["week_start"].unique())[-1]
                    new_boundary = (pd.Timestamp(new_latest_week) + pd.Timedelta(days=7)).date().isoformat()
                    orphaned = pd.read_sql("SELECT id FROM auto_forecasts WHERE target_week > ?",
                                            conn, params=(new_boundary,))
                    if not orphaned.empty:
                        conn.execute("DELETE FROM auto_forecasts WHERE target_week > ?", (new_boundary,))
                        removed_note = f"Also removed {len(orphaned)} forecast(s) that were generated assuming this batch's data existed."
                    else:
                        removed_note = "No orphaned forecasts needed removing."
                conn.commit()

                st.warning(f"Deleted batch '{batch_to_delete}'. {removed_note}")
                st.rerun()
        with col2:
            st.markdown("**Or clear everything**")
            if st.button("Clear ALL uploaded records (careful)"):
                conn.execute("DELETE FROM sales_records")
                conn.execute("DELETE FROM auto_forecasts")
                conn.commit()
                st.warning("All sales records and forecasts cleared.")
                st.rerun()

# --- TAB 2: Computed rates ---
with tab_rates:
    st.subheader("Rates computed from your uploaded data")
    if not has_data:
        st.warning("No sales data uploaded yet — go to tab 1 first.")
    else:
        st.markdown("**Price per kg** — computed from your actual revenue and kg sold, last 45 days")
        display_price = price_df.copy()
        display_price["$ per kg"] = display_price["price_per_kg"].round(2)
        display_price["kg per $1 CAD"] = (1 / display_price["price_per_kg"].replace(0, np.nan)).round(4)
        if "price_min" in display_price.columns:
            display_price["customer_spread"] = display_price.apply(
                lambda r: f"${r['price_min']:.2f}–${r['price_max']:.2f}" if pd.notna(r.get("price_min")) else "n/a",
                axis=1)

            # search/filter instead of dumping every row -- this table gets long fast, and
            # people come here looking for one specific rate, not the whole list
            st.markdown("**Group the rates by whichever dimensions you care about**")
            _dims = {"Channel": "channel", "Item": "product", "Bag size": "size_label"}
            if "customer" in sales_df.columns and not (sales_df["customer"] == "(not tracked)").all():
                _dims["Customer"] = "customer"
            _picked = st.multiselect(
                "Break down by", list(_dims.keys()), default=["Bag size"],
                help="Pick one to see rates for just that dimension (e.g. Bag size alone gives the "
                     "average $/kg for each size across the whole business). Pick several to split "
                     "them further.")

            if not _picked:
                st.info("Pick at least one dimension above to see rates.")
            else:
                _gcols = [_dims[p] for p in _picked]
                _q = st.text_input(
                    "Search", key="rate_q", placeholder="e.g. OSEE12, Costco, 12oz",
                    help="Searches ALL of your data — channel, item, bag size and customer — not just "
                         "the dimensions you're grouping by. Searching an item while grouped by "
                         "Channel gives you that item's rate per channel.")

                # filter the RAW data first, then compute rates on what's left. Previously the
                # search only looked at the grouped columns, so searching an item while grouped
                # by Channel matched nothing and returned an empty table -- which looks like
                # the item doesn't exist rather than like a search that couldn't apply.
                _src = sales_df
                if _q.strip():
                    _term = _q.strip()
                    _m = pd.Series(False, index=_src.index)
                    for _col in ("channel", "product", "size_label", "customer"):
                        if _col in _src.columns:
                            _m |= _src[_col].astype(str).str.contains(_term, case=False, na=False)
                    _src = _src[_m]
                    if _src.empty:
                        st.warning(f"Nothing in your sales data matches '{_term}' — check the spelling "
                                   "against the item and channel names in your uploads.")

                _rates = compute_rates_by(_src, _gcols) if not _src.empty else pd.DataFrame()
                if _rates.empty:
                    if not _q.strip():
                        st.info("Not enough data for that combination.")
                else:
                    _view = _rates
                    _scope = f" matching '{_q.strip()}'" if _q.strip() else ""
                    st.caption(f"Showing {len(_view):,} row(s){_scope} — "
                               f"grouped by {', '.join(_picked).lower()}, last 45 days.")
                    st.dataframe(
                        _view[_gcols + ["$ per kg", "kg per $1 CAD", "total_kg", "lines"]]
                        .rename(columns={"total_kg": "Kg sold (basis)", "lines": "Sale lines"}),
                        use_container_width=True, hide_index=True)
                    st.caption("**$ per kg** — revenue one kilo brings in. **kg per $1 CAD** — the flip "
                               "side, useful for budgeting. 'Kg sold (basis)' shows how much real volume "
                               "each rate is computed from — a rate built on very little volume is less "
                               "reliable than one built on a lot.")
        st.download_button("Download price_per_kg.csv", price_df.to_csv(index=False), "price_per_kg.csv")

        st.markdown("**Price per kg by customer** — real customer-specific pricing, where there's enough "
                     "of that customer's own history to trust it")
        customer_price_df = compute_customer_price_per_kg(sales_df)
        if customer_price_df.empty:
            st.info("This data source doesn't include customer identity — customer-specific pricing isn't available.")
        else:
            display_cust_price = customer_price_df.copy()
            display_cust_price["Rate used"] = np.where(
                display_cust_price["confident"], "Customer-specific", "Channel average (not enough customer history)")
            display_cust_price = display_cust_price.rename(columns={
                "channel": "Channel", "customer": "Customer", "product": "Item",
                "customer_price_per_kg": "This customer's price/kg", "price_per_kg_used": "Price used in translation",
                "n_transactions": "Transactions"})
            st.dataframe(
                display_cust_price[["Channel", "Customer", "Item", "This customer's price/kg",
                                     "Price used in translation", "Transactions", "Rate used"]]
                .sort_values(["Channel", "Item", "Price used in translation"], ascending=[True, True, False]),
                use_container_width=True)
            n_confident = int(customer_price_df["confident"].sum())
            st.caption(f"{n_confident} of {len(customer_price_df)} customer/item combinations have enough "
                       f"transactions (3+) for a real customer-specific rate — the rest fall back to the "
                       f"channel average, since a price from 1-2 orders is noise, not a rate.")
            st.download_button("Download customer_price_per_kg.csv", customer_price_df.to_csv(index=False),
                                "customer_price_per_kg.csv")

        st.markdown("**Size mix %** (within each channel-product)")
        st.dataframe(size_mix_df, use_container_width=True)

        st.markdown("**Customer mix %** (within each channel-product, highest share first)")
        if customer_mix_df.empty or (customer_mix_df["customer"] == "(not tracked)").all():
            st.info("This data source doesn't include customer identity — customer-level breakdown isn't "
                     "available. Everything else still works fine without it.")
        else:
            st.dataframe(customer_mix_df, use_container_width=True)
            st.download_button("Download customer_mix.csv", customer_mix_df.to_csv(index=False), "customer_mix.csv")

# --- TAB 3: Forecast (auto) ---
with tab_forecast:
    st.subheader("Auto-generated forecast — no manual entry")
    st.caption(
        "This forecast isn't typed in — it's calculated from your own uploaded sales history. When there's "
        "enough history (8+ weeks), it uses ARIMA (tested against this data: ~13.5% average error, vs 15.2% "
        "for the simpler fallback method). With less history, it falls back to a trend method (median of the "
        "last 4 weeks, damped growth vs the prior 4) since ARIMA is unreliable on very short series. Plus "
        "anything logged in Pipeline / known events. It updates automatically every time you upload new data."
    )

    st.markdown("**Check and fix unstable historical forecasts**")
    st.caption(
        "Forecasts are frozen once generated, so a fix to the forecasting method only affects NEW forecasts "
        "going forward — it can't reach back and correct a number already stored. This checks every stored "
        "forecast against what a safe, bounded method would have predicted at the time, and corrects any "
        "that are wildly unstable (e.g. a spike caused by ARIMA overreacting to a short-lived blip)."
    )
    if st.button("Check and fix now", key="audit_forecasts_btn"):
        with st.spinner("Checking every stored forecast..."):
            checked, fixed, examples = audit_and_fix_historical_forecasts(weekly_actual)
        if fixed == 0:
            st.success(f"Checked {checked} stored forecasts — all within a sane range, nothing to fix.")
        else:
            st.warning(f"Checked {checked} stored forecasts, corrected {fixed} that were wildly unstable:")
            example_df = pd.DataFrame(examples, columns=["Channel", "Product", "Week", "Was", "Now"])
            st.dataframe(example_df, use_container_width=True, hide_index=True)
            st.info("Corrected — refresh the page or switch tabs to see the updated numbers everywhere.")

    st.divider()

    if forecast_by_cp.empty:
        st.warning("Not enough history yet — upload a few weeks of sales data in tab 1 first.")
    else:
        display = forecast_by_cp.copy()
        if not active_overrides.empty:
            display["Overridden"] = display.apply(
                lambda r: "Yes" if ((active_overrides["channel"] == r["channel"]) &
                                     (active_overrides["product"] == r["product"])).any() else "", axis=1)
        else:
            display["Overridden"] = ""
        display = display.rename(columns={"forecast_kg": "Forecast_kg (incl. pipeline + overrides)", "pipeline_kg": "Pipeline_adjustment_kg"})
        st.dataframe(display[["channel", "product", "target_week", "Pipeline_adjustment_kg",
                               "Forecast_kg (incl. pipeline + overrides)", "Overridden"]],
                     use_container_width=True)
        total_kg = forecast_by_cp["forecast_kg"].sum()
        st.metric("Total forecast kg (next unforecasted week, all channels/products)", f"{total_kg:,.0f} kg")

        # ---- model tuning: searched once, on demand, never on a page load ----
        with st.expander("Forecast model settings"):
            st.caption(
                "Each segment is forecast with an ARIMA model. By default it uses a sensible "
                "general-purpose configuration, ARIMA(1,1,1). Running a search tests many "
                "configurations against your own history and keeps whichever fits best. It takes "
                "a minute or two, so it runs only when you ask — the result is stored and reused "
                "until you search again. Worth re-running if accuracy starts drifting."
            )
            _segs = split_into_segments(sales_df) if has_data else {}
            if _segs:
                _rows = []
                for _lab in _segs:
                    _s = get_stored_order(_lab, "W")
                    _rows.append({
                        "Segment": _lab,
                        "Model in use": f"ARIMA{_s[0]}" if _s else "ARIMA(1, 1, 1) — default",
                        "Seasonal": (f"every {_s[1][3]} weeks" if _s and _s[1][3] else "none"),
                        "Last searched": _s[2] if _s else "never",
                    })
                st.dataframe(pd.DataFrame(_rows), use_container_width=True, hide_index=True)

                if st.button("Find the best model for each segment", key="run_model_search"):
                    _prog = st.progress(0.0, text="Searching…")
                    _out = []
                    for _i, (_lab, _sdf) in enumerate(_segs.items()):
                        _prog.progress(_i / len(_segs), text=f"Searching {_lab}…")
                        _ag = aggregate_periods(_sdf, ["product_type"], "W")
                        _ser = _ag.groupby("period", as_index=False)["actual_kg"].sum() \
                            .sort_values("period")["actual_kg"].tolist()
                        if len(_ser) < 12:
                            _out.append(f"{_lab}: not enough history to search — keeping the default")
                            continue
                        _o, _so = search_best_order(_lab, _ser, "W")
                        _out.append(f"{_lab}: ARIMA{_o}" + (f" seasonal every {_so[3]} weeks" if _so[3] else ""))
                    _prog.empty()
                    reset_adjustment_state()
                    st.success("Search complete — these models are now in use:\n\n" +
                               "\n\n".join(f"- {line}" for line in _out))
                    st.rerun()

        # visible consensus check -- these per-item numbers are reconciled to the same
        # segment-based total the Dashboard shows, so the two should agree. An active
        # override deliberately breaks that tie (it replaces a number by design), so the
        # check names that explicitly rather than looking like a bug.
        _dash_seg = compute_segment_forecast(sales_df, freq="W")
        _dash_total = (sum(_dash_seg.values()) if _dash_seg else 0) + \
            (pipeline_by_cp["pipeline_kg"].sum() if not pipeline_by_cp.empty else 0)
        if _dash_total > 0:
            _gap = abs(total_kg - _dash_total)
            if _gap < max(1.0, _dash_total * 0.005):
                st.success(f"Matches the Dashboard total ({_dash_total:,.0f} kg) — every tab is using the same forecast.")
            elif not active_overrides.empty:
                st.info(f"Dashboard total is {_dash_total:,.0f} kg. The {_gap:,.0f} kg difference is from "
                        f"{len(active_overrides)} active manual override(s), which deliberately replace the "
                        "calculated number here.")
            else:
                st.warning(f"This total ({total_kg:,.0f} kg) doesn't match the Dashboard ({_dash_total:,.0f} kg) — "
                           f"a {_gap:,.0f} kg gap with no active overrides to explain it. Worth a look.")

        if not translated.empty:
            st.markdown("**Broken down by size — with bag counts to order**")
            has_bags = "forecast_bags" in translated.columns and translated["forecast_bags"].notna().any()
            if has_bags:
                # the number Operations actually orders with: total bags per size, across
                # every channel and item. Rounded UP -- you can't order a partial bag.
                bag_summary = translated.dropna(subset=["forecast_bags"]).groupby(
                    "size_label", as_index=False).agg(
                    forecast_kg=("forecast_kg", "sum"), bags_to_order=("forecast_bags", "sum"))
                bag_summary["forecast_kg"] = bag_summary["forecast_kg"].round(0)
                bag_summary["bags_to_order"] = np.ceil(
                    pd.to_numeric(bag_summary["bags_to_order"], errors="coerce").astype("float64")
                ).fillna(0).astype("int64")  # fillna before astype(int) -- NaN can't cast to int
                st.dataframe(bag_summary.rename(columns={
                    "size_label": "Bag size", "forecast_kg": "Forecast (kg)",
                    "bags_to_order": "Bags to order"}), use_container_width=True, hide_index=True)
                st.caption("Bag counts come from a real kg-per-bag rate computed from your own sales history "
                           "(total kg ÷ total units for that size), rounded up — you can't order a partial bag. "
                           "The full channel/item detail is below.")
                with st.expander("Full detail by channel and item"):
                    st.dataframe(translated, use_container_width=True)
            else:
                st.info("Showing kg only — no quantity/bag-count column was mapped on upload, so a real "
                        "kg-per-bag rate can't be computed. Re-upload with a Quantity column mapped (tab 1) "
                        "to get bag counts here.")
                st.dataframe(translated, use_container_width=True)
            st.download_button("Download forecast_by_size.csv", translated.to_csv(index=False), "forecast_by_size.csv")

with tab_salesplan:
    st.subheader("Sales plan — the top-down half of the S&OP bridge")
    st.caption(
        "This is Sales' own forward-looking plan — entered by month, channel, and item — translated "
        "into kg using the same price/kg rates used everywhere else in this app. It's compared against "
        "the app's own demand-sensing forecast below, so a gap between 'what Sales planned' and "
        "'what's actually happening' is visible early, not discovered at year-end."
    )

    plan_year = st.text_input("Plan year", value=str(date.today().year), key="plan_year")

    st.markdown("### Upload a plan (bulk)")
    plan_layout = st.radio(
        "How is your plan laid out?",
        ["Wide — months across the top (our standard template)", "Long — one row per channel/month"],
        help="The company revenue template has channels down the side and one column per month. "
             "Pick 'Wide' for that; the app will unpivot it for you.")
    plan_file = st.file_uploader("CSV or Excel", type=["csv", "xlsx", "xls"], key="plan_upload")

    if plan_file is not None and plan_layout.startswith("Wide"):
        _raw = pd.read_excel(plan_file) if plan_file.name.lower().endswith((".xlsx", ".xls")) else pd.read_csv(plan_file)
        st.dataframe(_raw.head(10), use_container_width=True)

        _label_col = _raw.columns[0]
        _date_cols = [col for col in _raw.columns if isinstance(col, datetime)]
        if not _date_cols:
            st.warning("No date-style column headers found in this file.")
        else:
            # This template encodes the plan year in the DAY part of each month header
            # (…-01-26 = Jan of the 2026 plan, …-01-27 = Jan of the 2027 plan), with the two
            # year blocks separated by a Notes column. Grouping by the day component is what
            # keeps the 2026 and 2027 sections apart -- reading them all as one year would
            # silently merge two plans on top of each other.
            _by_year = {}
            for col in _date_cols:
                _by_year.setdefault(col.day, []).append(col)
            _year_opts = {}
            for _d, _cols in _by_year.items():
                _yr = 2000 + _d if 20 <= _d <= 99 else pd.Timestamp(_cols[0]).year
                _year_opts[f"{_yr} ({len(_cols)} months)"] = (_yr, _cols)

            _pick_years = st.multiselect(
                "Which plan year(s) to import?", list(_year_opts.keys()),
                default=list(_year_opts.keys()),
                help="This template holds two years side by side. Importing both gives you a "
                     "continuous plan from January of the first year through December of the second.")
            if not _pick_years:
                st.info("Pick at least one plan year.")
                st.stop()
            _year_jobs = [_year_opts[k] for k in _pick_years]
            _year, _month_cols = _year_jobs[0]
            _all_month_cols = [mc for _, cols in _year_jobs for mc in cols]

            _unit = st.radio("These values are:", ["Revenue ($)", "Volume (kg)"], horizontal=True, key="wide_unit")

            _row_labels = _raw[_label_col].astype(str).fillna("")
            _has_numbers = _raw[_all_month_cols].apply(pd.to_numeric, errors="coerce").notna().any(axis=1)
            _usable = _raw[_has_numbers]
            if _usable.empty:
                st.warning("None of the rows in this file have numeric values in the month columns "
                           "— the sheet may still be an empty template.")
            else:
                st.caption(f"{len(_usable)} row(s) have plan numbers: "
                           + ", ".join(_usable[_label_col].astype(str).tolist()))
                _which = st.multiselect(
                    "Import which rows?", _usable[_label_col].astype(str).tolist(),
                    default=_usable[_label_col].astype(str).tolist(),
                    help="If only the Total row is filled in, import that — the app will compare "
                         "it against the whole-company forecast.")

                _replace = st.checkbox(
                    f"Replace any existing {_year} plan rows (recommended)", value=True,
                    help="Leave this on unless you're deliberately adding to an existing plan. "
                         "Re-importing without it appends a second copy of every row.")
                if st.button("Import this plan", type="primary"):
                    _rows_written = 0
                    for _yr, _cols in _year_jobs:
                        if _replace:
                            conn.execute("DELETE FROM sales_plan WHERE plan_year = ?", (str(_yr),))
                        for _, _r in _usable.iterrows():
                            _name = str(_r[_label_col]).strip()
                            if _name not in _which:
                                continue
                            # a Total row is stored as a company-wide plan so it can be compared
                            # against the total forecast rather than one channel's
                            _is_total = "total" in _name.lower()
                            _chan = "(total)" if _is_total else _name
                            for _mc in _cols:
                                _val = pd.to_numeric(_r[_mc], errors="coerce")
                                if pd.isna(_val):
                                    continue
                                _kg = float(_val) if _unit.startswith("Volume") else None
                                _rev = float(_val) if _unit.startswith("Revenue") else None
                                conn.execute("""INSERT INTO sales_plan
                                    (plan_year, month, channel, product, planned_kg, planned_dollars,
                                     updated_by, updated_at, note)
                                    VALUES (?,?,?,?,?,?,?,?,?)""",
                                    (str(_yr), f"{_mc.month:02d}", _chan, "(all)", _kg, _rev,
                                     "template import", datetime.now().isoformat(),
                                     "company total" if _is_total else "from template"))
                                _rows_written += 1
                    conn.commit()
                    st.success(f"Imported {_rows_written} plan rows across "
                               f"{', '.join(str(y) for y, _ in _year_jobs)}.")
                    st.rerun()
        plan_file = None  # handled above; skip the long-format path below

    if plan_file is not None:
        plan_raw = pd.read_excel(plan_file) if plan_file.name.lower().endswith((".xlsx", ".xls")) else pd.read_csv(plan_file)
        st.dataframe(plan_raw.head(), use_container_width=True)
        pcols = list(plan_raw.columns)
        pc1, pc2, pc3 = st.columns(3)
        with pc1:
            plan_channel_col = st.selectbox("Channel column", pcols, key="plan_channel_col")
            plan_month_col = st.selectbox("Month column", pcols, key="plan_month_col")
        with pc2:
            plan_product_col = st.selectbox("Item column", pcols, key="plan_product_col")
            plan_amount_col = st.selectbox("Planned amount column", pcols, key="plan_amount_col")
        with pc3:
            plan_amount_type = st.radio("This amount is in", ["Dollars ($)", "Kilograms (kg)"], key="plan_amount_type")

        if st.button("Save this plan", type="primary"):
            std_plan = pd.DataFrame()
            std_plan["channel"] = plan_raw[plan_channel_col].astype(str)
            std_plan["product"] = plan_raw[plan_product_col].astype(str)
            std_plan["month"] = plan_raw[plan_month_col].astype(str)
            if plan_amount_type == "Dollars ($)":
                std_plan["planned_dollars"] = pd.to_numeric(plan_raw[plan_amount_col], errors="coerce")
                std_plan["planned_kg"] = np.nan
            else:
                std_plan["planned_kg"] = pd.to_numeric(plan_raw[plan_amount_col], errors="coerce")
                std_plan["planned_dollars"] = np.nan
            std_plan = std_plan.dropna(subset=["channel", "product", "month"])

            if not price_df.empty:
                rate_lookup = price_df.groupby(["channel", "product"], as_index=False)["price_per_kg"].mean()
                std_plan = std_plan.merge(rate_lookup, on=["channel", "product"], how="left")
                std_plan["planned_kg"] = np.where(
                    std_plan["planned_kg"].isna() & std_plan["planned_dollars"].notna() & std_plan["price_per_kg"].notna(),
                    std_plan["planned_dollars"] / std_plan["price_per_kg"], std_plan["planned_kg"])
                std_plan = std_plan.drop(columns=["price_per_kg"])

            std_plan["plan_year"] = plan_year
            std_plan["updated_at"] = datetime.now().isoformat()
            std_plan["updated_by"] = ""
            std_plan["note"] = ""
            insert_dataframe("sales_plan", std_plan)
            st.success(f"Saved {len(std_plan)} plan rows for {plan_year}.")
            st.rerun()

    st.divider()
    st.markdown("### Edit the current plan")
    st.caption("Add, edit, or delete rows directly — this is how Sales keeps the plan updated over time.")
    existing_plan = pd.read_sql("SELECT * FROM sales_plan WHERE plan_year = ? ORDER BY month, channel, product",
                                 conn, params=(plan_year,))

    # De-duplicate on the real key. Re-importing the same file (easy to do while getting the
    # options right) previously appended a second full set of rows, so every month appeared
    # twice with no way to tell which was current. Keeping the highest id keeps the most
    # recent import and quietly discards the superseded one.
    if not existing_plan.empty:
        _before = len(existing_plan)
        existing_plan = existing_plan.sort_values("id").drop_duplicates(
            subset=["plan_year", "month", "channel", "product"], keep="last")
        if len(existing_plan) < _before:
            st.info(f"{_before - len(existing_plan)} superseded row(s) hidden — showing the most "
                    "recent import for each channel and month. Use 'Clear this plan year' below "
                    "if you want to start clean.")

    if not existing_plan.empty:
        # show the year alongside the month so Jan 2026 and Jan 2027 are distinguishable --
        # "01" on its own made two different years look like duplicate rows
        existing_plan = existing_plan.copy()
        existing_plan["period"] = existing_plan["plan_year"].astype(str) + "-" + \
            existing_plan["month"].astype(str).str.zfill(2)
    edit_base = existing_plan[["channel", "product", "period", "month", "planned_dollars", "planned_kg", "note"]] \
        if not existing_plan.empty else pd.DataFrame(columns=["channel", "product", "period", "month", "planned_dollars", "planned_kg", "note"])

    if st.button("Clear this plan year and start over", key="clear_plan_year"):
        conn.execute("DELETE FROM sales_plan WHERE plan_year = ?", (plan_year,))
        conn.commit()
        st.warning(f"Cleared all plan rows for {plan_year}.")
        st.rerun()

    # flag rows that would be silently dropped on save, and rows imported without an item --
    # a wide-template import fills product with "(all)", which is fine, but a blank channel or
    # month makes the row unusable and was previously just discarded without explanation
    _bad = edit_base[edit_base[["channel", "month"]].isna().any(axis=1)] if not edit_base.empty else pd.DataFrame()
    if not _bad.empty:
        st.warning(f"{len(_bad)} row(s) are missing a channel or month and will be skipped when you "
                   "save. Fill them in below or delete them.")
    if not edit_base.empty:
        _no_val = edit_base[edit_base["planned_dollars"].isna() & edit_base["planned_kg"].isna()]
        if not _no_val.empty:
            st.info(f"{len(_no_val)} row(s) have no dollar or kg value yet — they'll save but won't "
                    "contribute to the plan-vs-reality comparison until you fill one in.")

    edited = st.data_editor(edit_base, num_rows="dynamic", use_container_width=True, key="plan_editor",
                             column_config={
                                 "planned_dollars": st.column_config.NumberColumn("Planned $", format="%.0f"),
                                 "planned_kg": st.column_config.NumberColumn("Planned kg", format="%.1f"),
                             })
    updated_by = st.text_input("Your name (for the record)", key="plan_updated_by_input")

    if st.button("Save changes to plan", type="primary"):
        edited_clean = edited.dropna(subset=["channel", "product", "month"], how="any").copy()
        if not edited_clean.empty and not price_df.empty:
            rate_lookup = price_df.groupby(["channel", "product"], as_index=False)["price_per_kg"].mean()
            edited_clean = edited_clean.merge(rate_lookup, on=["channel", "product"], how="left")
            edited_clean["planned_kg"] = np.where(
                edited_clean["planned_kg"].isna() & edited_clean["planned_dollars"].notna() & edited_clean["price_per_kg"].notna(),
                edited_clean["planned_dollars"] / edited_clean["price_per_kg"], edited_clean["planned_kg"])
            edited_clean = edited_clean.drop(columns=["price_per_kg"])
        edited_clean["plan_year"] = plan_year
        edited_clean["updated_at"] = datetime.now().isoformat()
        edited_clean["updated_by"] = updated_by

        conn.execute("DELETE FROM sales_plan WHERE plan_year = ?", (plan_year,))
        conn.commit()
        if not edited_clean.empty:
            # Keep only columns that actually exist in the table. The editor carries
            # display-only columns (a combined year-month label, and any helper column added
            # during rate conversion), and passing one of those to an INSERT fails with an
            # unhelpful "undefined column" error. Filtering against the real schema means a
            # new display column can never break saving again.
            _cols = pd.read_sql("SELECT * FROM sales_plan LIMIT 0", conn).columns.tolist()
            _keep = [col for col in edited_clean.columns if col in _cols and col != "id"]
            insert_dataframe("sales_plan", edited_clean[_keep])
        st.success(f"Plan for {plan_year} updated ({len(edited_clean)} rows).")
        st.rerun()

    st.divider()
    st.markdown("## Reconciliation — plan vs. reality")
    st.caption(
        "For months that already happened, this compares the plan to real actuals. For months still "
        "ahead, it compares the plan to the app's own demand-sensing projection — which gets less "
        "certain the further out it looks, unlike the plan (which usually assumes similar confidence "
        "across the whole year). Treat far-future gaps as a rough signal, not a precise miss."
    )

    # Compare in whatever unit the plan was entered in. The company template is in revenue
    # dollars, and converting that to kg just to compare against kg introduces avoidable
    # error -- comparing dollars to dollars is both simpler and more faithful to the plan.
    _plan_unit = "kg"
    if not existing_plan.empty:
        _has_kg = existing_plan["planned_kg"].notna().any()
        _has_dollars = existing_plan["planned_dollars"].notna().any()
        if _has_dollars and not _has_kg:
            _plan_unit = "dollars"
        elif _has_dollars and _has_kg:
            _plan_unit = st.radio("Compare the plan in:", ["kg", "dollars"], horizontal=True, key="recon_unit")

    _plan_src = existing_plan.copy() if not existing_plan.empty else pd.DataFrame()

    # Drop the "(total)" row when individual channels are also present. Importing both the
    # Total row AND the channel rows counts every dollar twice -- which is why the plan showed
    # roughly double the template's real monthly figure.
    if not _plan_src.empty and "channel" in _plan_src.columns:
        _has_channels = (_plan_src["channel"] != "(total)").any()
        _has_total = (_plan_src["channel"] == "(total)").any()
        if _has_channels and _has_total:
            _plan_src = _plan_src[_plan_src["channel"] != "(total)"]
            st.info("Both a company Total row and individual channel rows were imported. Using the "
                    "channel rows and ignoring the Total, so nothing is counted twice.")

    # Let the user compare in kg even when the plan is in dollars, by converting through the
    # real computed rate. Without this a revenue plan can never be compared against volume.
    _plan_unit_choice = _plan_unit
    _converted_note = ""
    if _plan_unit == "dollars" and not price_df.empty:
        _plan_unit_choice = st.radio(
            "Compare in:", ["dollars (as planned)", "kg (converted from $)"], horizontal=True,
            key="recon_convert",
            help="Your plan is in revenue. Converting to kg uses the volume-weighted rate from "
                 "tab 2, so you can compare against actual kg and drive supply decisions.")
        if _plan_unit_choice.startswith("kg"):
            _wk = float(sales_df["kg"].sum()) if has_data else 0
            _wr = float(sales_df["revenue"].sum()) if has_data else 0
            _rate = (_wr / _wk) if _wk > 0 else None
            if _rate:
                _plan_src = _plan_src.copy()
                _plan_src["planned_kg"] = _plan_src["planned_dollars"] / _rate
                _plan_unit = "kg"
                _converted_note = f" — converted from $ at ${_rate:,.2f}/kg (your blended actual rate)"
            else:
                st.warning("Can't convert to kg — no revenue/kg history to derive a rate from.")

    _plan_col = "planned_kg" if _plan_unit == "kg" else "planned_dollars"

    # Key the plan on YYYY-MM, not a bare month number. Real bug: the plan stored month as
    # "01" while actuals are keyed "2026-01", so the merge never matched and every Reality
    # cell came back blank.
    if not _plan_src.empty:
        _plan_src = _plan_src.copy()
        _plan_src["month_key"] = _plan_src["plan_year"].astype(str) + "-" + \
            _plan_src["month"].astype(str).str.zfill(2)
    plan_monthly = _plan_src.groupby("month_key", as_index=False)[_plan_col].sum() \
        .rename(columns={_plan_col: "planned_kg", "month_key": "month"}) if not _plan_src.empty else pd.DataFrame()
    if not plan_monthly.empty:
        plan_monthly = plan_monthly[plan_monthly["planned_kg"].notna() & (plan_monthly["planned_kg"] != 0)]

    _unit_label = "kg" if _plan_unit == "kg" else "$"

    if plan_monthly.empty:
        st.info("Enter a plan above to see the reconciliation view.")
    elif not has_data:
        st.info("Upload sales history in tab 1 to compare the plan against.")
    else:
        st.caption(f"Comparing in **{_unit_label}**{_converted_note}.")
        actuals_monthly = sales_df.copy()
        actuals_monthly["record_date"] = pd.to_datetime(actuals_monthly["record_date"], errors="coerce")
        actuals_monthly["month"] = actuals_monthly["record_date"].dt.to_period("M").astype(str)
        _acol = "kg" if _plan_unit == "kg" else "revenue"
        actuals_monthly = actuals_monthly.groupby("month", as_index=False)[_acol].sum() \
            .rename(columns={_acol: "actual_kg"})

        recon = plan_monthly.merge(actuals_monthly, on="month", how="left").sort_values("month")
        missing_months = recon[recon["actual_kg"].isna()]["month"].tolist()

        recon["demand_sensing_kg"] = np.nan
        if missing_months:
            company_monthly = sales_df.copy()
            company_monthly["record_date"] = pd.to_datetime(company_monthly["record_date"], errors="coerce")
            company_monthly["month"] = company_monthly["record_date"].dt.to_period("M").astype(str)
            # project in the SAME unit the comparison is using. Real bug: this always
            # projected kg, so when comparing in dollars the line switched from ~1,300,000
            # (revenue actuals) to ~30,000 (kg projection) the moment actuals ran out --
            # which looked like demand collapsing to nothing.
            company_monthly_agg = company_monthly.groupby("month", as_index=False)[_acol].sum() \
                .rename(columns={_acol: "kg"}).sort_values("month")
            # same partial-month exclusion as the dashboard -- projecting the rest of the year
            # from a month that only has one week of data produces a wildly wrong plan
            # comparison, which is worse than no comparison
            _cl = company_monthly["record_date"].max()
            if pd.notna(_cl) and _cl < _cl.to_period("M").to_timestamp("M"):
                company_monthly_agg = company_monthly_agg[
                    company_monthly_agg["month"] != str(_cl.to_period("M"))]
            if len(company_monthly_agg) >= 2:
                with st.spinner("Computing demand-sensing projection for remaining months..."):
                    proj_recon = project_forward_with_range(company_monthly_agg["kg"].tolist(), None,
                                                              n_periods=min(len(missing_months) + 3, 12),
                                                              keep_trend=True)
                # apply events + overrides here too, so plan-vs-reality compares the plan
                # against the SAME forward number the rest of the app shows -- otherwise a
                # signed contract would look like a plan gap rather than expected volume
                _monthly_adj = 0.0
                if type_level_forecasts:
                    for _lab, _base in type_level_forecasts.items():
                        _monthly_adj += (type_level_forecasts_with_pipeline.get(_lab, _base) - _base)
                _monthly_adj *= 4.345
                if abs(_monthly_adj) > 0.5:
                    proj_recon = proj_recon.copy()
                    proj_recon["forecast_kg"] = (proj_recon["forecast_kg"] + _monthly_adj).clip(lower=0)
                    st.caption(f"Projection includes logged events and overrides ({_monthly_adj:+,.0f} kg/mo).")
                last_month = pd.Period(company_monthly_agg["month"].iloc[-1], freq="M")
                proj_months = [(last_month + i + 1).strftime("%Y-%m") for i in range(len(proj_recon))]
                proj_lookup = dict(zip(proj_months, proj_recon["forecast_kg"]))
                recon["demand_sensing_kg"] = recon["month"].map(proj_lookup)

                # Also fill in what the system WOULD have forecast for months that already
                # happened, using only the data available before each one. Without this the
                # system line only exists in the future, so you can never check it against
                # actuals -- which is the one comparison that tells you if the model is any good.
                _hist = company_monthly_agg.reset_index(drop=True)
                _back = {}
                for _i in range(2, len(_hist)):
                    _f = trend_forecast(_hist["kg"].iloc[:_i].tolist())
                    if _f is not None:
                        _back[_hist["month"].iloc[_i]] = _f
                recon["demand_sensing_kg"] = recon["demand_sensing_kg"].fillna(
                    recon["month"].map(_back))

        recon["Status"] = np.where(recon["actual_kg"].notna(), "Actual", "Forecast (projected)")
        recon["Reality (kg)"] = recon["actual_kg"].fillna(recon["demand_sensing_kg"])
        recon["Gap vs plan (kg)"] = recon["Reality (kg)"] - recon["planned_kg"]
        recon["Gap %"] = np.where(recon["planned_kg"] != 0, recon["Gap vs plan (kg)"] / recon["planned_kg"] * 100, np.nan)

        # label the columns with the unit actually being compared -- they previously always
        # said "kg" even when the comparison was in dollars, which is quietly misleading
        _u = _unit_label
        display_recon = recon.rename(columns={
            "month": "Month", "planned_kg": f"Plan ({_u})",
            "Reality (kg)": f"Reality ({_u})", "Gap vs plan (kg)": f"Gap vs plan ({_u})"})[
            ["Month", f"Plan ({_u})", f"Reality ({_u})", f"Gap vs plan ({_u})", "Gap %", "Status"]].round(1)
        st.dataframe(display_recon, use_container_width=True, hide_index=True)

        # Three separate lines rather than two. Previously actuals and the system's forecast
        # were merged into one "Actual / demand-sensing" series, so you couldn't see where
        # real data ended and prediction began, or compare the plan against the system's own
        # view for months that have already happened.
        fig_recon = go.Figure()
        fig_recon.add_trace(go.Scatter(
            x=recon["month"], y=recon["planned_kg"], mode="lines+markers",
            name="Sales plan (what we said we'd do)",
            line=dict(color="rgb(217,119,6)", width=2)))
        fig_recon.add_trace(go.Scatter(
            x=recon["month"], y=recon["actual_kg"], mode="lines+markers",
            name="Actual (what really happened)",
            line=dict(color="rgb(31,119,180)", width=3),
            connectgaps=False))
        if "demand_sensing_kg" in recon.columns and recon["demand_sensing_kg"].notna().any():
            fig_recon.add_trace(go.Scatter(
                x=recon["month"], y=recon["demand_sensing_kg"], mode="lines+markers",
                name="System forecast (what the data predicts)",
                line=dict(color="#555", width=2, dash="dash"),
                connectgaps=False))
        fig_recon.update_layout(height=380, margin=dict(l=10, r=10, t=46, b=10), plot_bgcolor="white",
                                 yaxis_title=_unit_label, hovermode="x unified",
                                 yaxis=dict(showgrid=True, gridcolor="rgba(0,0,0,0.06)"),
                                 legend=dict(orientation="h", yanchor="bottom", y=1.02,
                                             xanchor="left", x=0))
        st.plotly_chart(fig_recon, use_container_width=True)
        st.caption(
            "**Plan vs Actual** for months that have happened tells you how well Sales planned. "
            "**Plan vs System forecast** for months ahead tells you whether the plan still looks "
            "achievable. **Actual vs System forecast** on past months is the accuracy check on the "
            f"model itself. All three are in {_unit_label}."
        )

# --- TAB: Pipeline / known events ---
with tab_pipeline:
    st.subheader("Log what's happening right now — before it shows up in sales data")
    st.caption(
        "Signed a new contract? Lost an account? Know volume is about to change? Log it here. "
        "A trend forecast can't see a deal that hasn't shipped yet — but you already know about it. "
        "Once logged, it adjusts the forecast automatically, on top of the auto-generated baseline."
    )

    with st.form("pipeline_form"):
        c1, c2 = st.columns(2)
        with c1:
            event_type = st.selectbox("What happened", [
                "New contract signed", "Account lost / churned", "Expected volume change", "Other"])
            customer = st.text_input("Customer / account name")
            channel = st.selectbox("Channel", sorted(sales_df["channel"].unique().tolist())
                                    if has_data else ["Wholesale", "Retail"])
            product = st.selectbox("Product", sorted(sales_df["product"].unique().tolist())
                                    if has_data else ["Product"])
        with c2:
            rate_unit = st.radio(
                "How is this volume expressed?", ["Per month", "Per week"], horizontal=True,
                help="A new account quoted as 'X kg a month' vs 'X kg a week' are very different "
                     "amounts — pick whichever way you were told it, and the app converts.")
            expected_kg_raw = st.number_input(
                "Expected kg impact", step=1.0, value=0.0,
                help="Positive for new/growing business, negative for a lost account or a "
                     "downgrade. Negative values are allowed.")
            start_date = st.date_input(
                "Starts on", value=datetime.now().date(),
                help="The date this actually begins — e.g. the day a new account's first order "
                     "ships. The event counts from this date forward; it won't affect any week "
                     "before it.")
            # stored as the month label the rest of the app already keys on, so the date is a
            # friendlier way to say the same thing rather than a schema change
            starting_cycle = start_date.strftime("%Y-%m")
            ongoing = st.checkbox("Ongoing (keeps applying to future cycles)", value=True,
                                   help="Uncheck for a one-time event that only affects this one cycle.")
            submitted_by = st.text_input("Logged by")
        note = st.text_area("Note (context for whoever reads this later)")

        if st.form_submit_button("Log this event", type="primary"):
            # store everything as a monthly figure regardless of how it was entered, so there
            # is exactly one unit in the database and no ambiguity downstream
            expected_kg = expected_kg_raw * 4.345 if rate_unit == "Per week" else expected_kg_raw
            conn.execute("""INSERT INTO pipeline_events
                (timestamp, submitted_by, event_type, customer, channel, product,
                 expected_kg_per_month, starting_cycle, ongoing, note)
                VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (datetime.now().isoformat(), submitted_by, event_type, customer, channel, product,
                 expected_kg, starting_cycle, int(ongoing), note))
            conn.commit()
            st.success("Logged — the forecast now includes this.")
            st.rerun()

    st.divider()
    if all_events_all.empty:
        st.info("No events logged yet.")
    else:
        st.markdown(f"**Currently applying to cycle {cycle}:**")
        if applicable.empty:
            st.caption("None of the logged events apply to this cycle yet.")
        else:
            st.dataframe(
                applicable[["event_type", "customer", "channel", "product",
                            "expected_kg_per_month", "starting_cycle", "ongoing", "note"]],
                use_container_width=True)
            total_pipeline_kg = applicable["expected_kg_per_month"].sum()
            st.metric("Total pipeline adjustment this cycle", f"{total_pipeline_kg:+,.0f} kg")

            # An event is a stand-in for volume the model can't see YET. Once real sales for
            # that scope start arriving, the model learns the new level on its own -- and
            # leaving the event on then double-counts it. This flags when that's happened so
            # you know it's safe to turn off, rather than having to remember.
            if has_data and not weekly_actual.empty:
                _ready = []
                for _, ev in applicable.iterrows():
                    _scope = weekly_actual.copy()
                    for dim in ("channel", "product"):
                        if dim in ev and pd.notna(ev.get(dim)) and str(ev[dim]) not in ("", "(all)"):
                            _scope = _scope[_scope[dim] == ev[dim]]
                    if _scope.empty:
                        continue
                    _start_month = str(ev.get("starting_cycle", ""))
                    _after = _scope[_scope["week_start"].astype(str) >= f"{_start_month}-01"]
                    _weeks_of_actuals = _after["week_start"].nunique()
                    if _weeks_of_actuals >= 4:
                        _expected_weekly = float(ev["expected_kg_per_month"]) / 4.345
                        _before = _scope[_scope["week_start"].astype(str) < f"{_start_month}-01"]
                        if _before.empty:
                            continue
                        _base = _before.groupby("week_start")["actual_kg"].sum().tail(8).mean()
                        _now = _after.groupby("week_start")["actual_kg"].sum().mean()
                        _observed_lift = _now - _base
                        # has the real lift caught up to at least half what the event predicted?
                        if _expected_weekly != 0 and _observed_lift / _expected_weekly >= 0.5:
                            _ready.append({
                                "Event": f"{ev.get('customer') or ''} {ev.get('channel') or ''} "
                                         f"{ev.get('product') or ''}".strip(),
                                "Weeks of real data since start": _weeks_of_actuals,
                                "Event predicted (kg/wk)": round(_expected_weekly, 1),
                                "Actually observed (kg/wk)": round(_observed_lift, 1),
                            })
                if _ready:
                    st.warning(f"{len(_ready)} event(s) now show up in real sales data. The forecast has "
                               "learned this volume on its own, so leaving the event on double-counts it "
                               "— worth turning these off.")
                    st.dataframe(pd.DataFrame(_ready), use_container_width=True, hide_index=True)

            st.markdown("**Is this actually changing the forecast? Here's the before/after.**")
            if not live_forecast.empty:
                impact = live_forecast[["channel", "product", "forecast_kg"]].rename(
                    columns={"forecast_kg": "Baseline forecast (kg) — trend only, no events"})
                impact = impact.merge(pipeline_by_cp, on=["channel", "product"], how="inner")
                if impact.empty:
                    st.warning("None of the logged events match a channel/product the trend model currently "
                               "has a forecast for — double check the Channel and Product you selected when "
                               "logging the event match what's actually in your uploaded sales data.")
                else:
                    impact["Adjusted forecast (kg) — with events"] = (
                        impact["Baseline forecast (kg) — trend only, no events"] + impact["pipeline_kg"])
                    impact = impact.rename(columns={"channel": "Channel", "product": "Item",
                                                     "pipeline_kg": "Event adjustment (kg)"})
                    st.dataframe(impact[["Channel", "Item", "Baseline forecast (kg) — trend only, no events",
                                          "Event adjustment (kg)", "Adjusted forecast (kg) — with events"]],
                                 use_container_width=True)
                    st.caption("If the baseline and adjusted columns are identical, the event isn't actually "
                               "reaching this channel/product — check the Channel/Product spelling matches "
                               "your sales data exactly.")
            else:
                st.info("No baseline forecast to compare against yet — upload more sales history first.")

        st.markdown("**Manage events**")
        st.caption("**Stop applying** ends an event going forward but keeps it in the record, so past "
                   "forecasts and accuracy stay exactly as they were reported at the time. **Delete** "
                   "erases it completely and will retroactively change past forecast numbers — use it "
                   "only for something logged in error.")

        def _ev_label(r):
            _st = "" if int(r.get("active", 1)) == 1 else "  [inactive]"
            return (f"#{r['id']} — {r['event_type']} — {r['customer']} — {r['channel']}/{r['product']} "
                    f"({r['expected_kg_per_month']:+.0f} kg/mo, starts {r['starting_cycle']}){_st}")

        _act = all_events_all[all_events_all["active"] == 1]
        _inact = all_events_all[all_events_all["active"] == 0]

        if not _act.empty:
            _al = _act.apply(_ev_label, axis=1).tolist()
            _am = dict(zip(_al, _act["id"]))
            _pick_off = st.selectbox("Active events", _al, key="ev_stop_pick")
            cstop, cdel = st.columns(2)
            with cstop:
                if st.button("Stop applying", key="ev_stop"):
                    conn.execute("UPDATE pipeline_events SET active = 0, deactivated_at = ? WHERE id = ?",
                                 (datetime.now().strftime("%Y-%m-%d"), int(_am[_pick_off]),))
                    conn.commit()
                    reset_adjustment_state()
                    st.warning("Event stopped — it no longer affects the forecast, but stays in the record.")
                    st.rerun()
            with cdel:
                if st.button("Delete permanently", key="ev_del"):
                    conn.execute("DELETE FROM pipeline_events WHERE id = ?", (int(_am[_pick_off]),))
                    conn.commit()
                    reset_adjustment_state()
                    st.warning("Deleted. Past forecast numbers that included this event have changed.")
                    st.rerun()

        if not _inact.empty:
            _il = _inact.apply(_ev_label, axis=1).tolist()
            _im = dict(zip(_il, _inact["id"]))
            _pick_on = st.selectbox("Inactive events", _il, key="ev_on_pick")
            if st.button("Reactivate this event", key="ev_on"):
                conn.execute("UPDATE pipeline_events SET active = 1, deactivated_at = NULL WHERE id = ?",
                             (int(_im[_pick_on]),))
                conn.commit()
                reset_adjustment_state()
                st.success("Event reactivated — it applies to the forecast again.")
                st.rerun()

        with st.expander("All logged events (all cycles)"):
            _h = all_events_all.copy()
            _h["status"] = _h["active"].map({1: "Active", 0: "Inactive"})
            st.dataframe(_h, use_container_width=True)

    st.divider()
    st.subheader("Manual override — replace the number directly")
    st.caption("Both tools live here: log an **event** above to ADD volume with a reason, or set an "
               "**override** below to REPLACE a number outright.")
    st.caption(
        "Pipeline events add a reason-backed adjustment on top of the auto forecast. This is different: "
        "it directly replaces the auto+pipeline number for a channel/item with whatever you type — for "
        "when you just know better than the trend, without a specific logged event to point to. "
        "An active override wins over everything else, and shows up in every table and KPI above and "
        "throughout the app (Dashboard, Ops capacity, translated $) — this is the one number everything else builds on."
    )

    current_target_week = None
    with st.form("override_form"):
        st.caption(f"Set any dimension to **{OVERRIDE_ANY}** to make it apply broadly — e.g. channel="
                   f"Specialty Retail with item={OVERRIDE_ANY} overrides that whole channel. If several "
                   "overrides could apply, the most specific one wins (an item-level rule beats a "
                   "channel-wide one).")
        oc1, oc2 = st.columns(2)
        with oc1:
            _channels = [OVERRIDE_ANY] + (sorted(sales_df["channel"].unique().tolist()) if has_data else [])
            _products = [OVERRIDE_ANY] + (sorted(sales_df["product"].unique().tolist()) if has_data else [])
            _has_cust = has_data and "customer" in sales_df.columns and \
                not (sales_df["customer"] == "(not tracked)").all()
            _customers = [OVERRIDE_ANY] + (sorted(
                sales_df.loc[sales_df["customer"] != "(not tracked)", "customer"].unique().tolist())
                if _has_cust else [])
            ov_channel = st.selectbox("Channel", _channels, key="ov_channel")
            ov_product = st.selectbox("Item", _products, key="ov_product")
            ov_customer = st.selectbox("Customer", _customers, key="ov_customer",
                                        help="Leave as (all) unless this override is for one specific customer."
                                        if _has_cust else "This data source doesn't include customer identity.",
                                        disabled=not _has_cust)
        with oc2:
            # no min_value: a negative override is legitimate (e.g. modelling a return or a
            # correction), and blocking it silently forced people to work around the tool
            ov_kg = st.number_input("Override forecast (kg)", step=10.0, value=0.0,
                                     help="The number this scope should forecast at. Negative is allowed.")
            ov_by = st.text_input("Your name")
        ov_period = st.radio(
            "How long should this apply?",
            ["One-time (just the current forecast period)", "Ongoing (until I turn it off)"],
            help="One-time automatically expires and reverts to the auto forecast once new data moves the "
                 "forecast to the next period. Ongoing keeps applying every time until you manually turn it off.")
        ov_note = st.text_area("Why are you overriding this? (kept for the record, doesn't gate the override)")

        if st.form_submit_button("Set override", type="primary"):
            if ov_channel == OVERRIDE_ANY and ov_product == OVERRIDE_ANY and ov_customer == OVERRIDE_ANY:
                st.error("At least one of Channel, Item, or Customer must be a specific value — "
                         "an override with everything set to (all) would replace the entire company forecast.")
            else:
                if not live_forecast.empty and ov_channel != OVERRIDE_ANY and ov_product != OVERRIDE_ANY:
                    match_row = live_forecast[(live_forecast["channel"] == ov_channel) & (live_forecast["product"] == ov_product)]
                    current_target_week = match_row.iloc[0]["target_week"] if not match_row.empty else None
                elif not live_forecast.empty:
                    # a broad override isn't tied to one channel/item pair -- anchor its
                    # one-time expiry to the current forecast week overall
                    current_target_week = live_forecast["target_week"].max()
                period_type = "One-time" if ov_period.startswith("One-time") else "Ongoing"
                # deactivate any existing override with the EXACT same scope, so setting one
                # twice replaces it rather than leaving two competing rules at equal specificity
                conn.execute("""UPDATE manual_overrides SET active = 0
                    WHERE channel = ? AND product = ? AND COALESCE(customer, ?) = ?""",
                    (ov_channel, ov_product, OVERRIDE_ANY, ov_customer))
                conn.execute("""INSERT INTO manual_overrides
                    (timestamp, submitted_by, channel, product, customer, override_kg, note, active, period_type, target_week)
                    VALUES (?,?,?,?,?,?,?,1,?,?)""",
                    (datetime.now().isoformat(), ov_by, ov_channel, ov_product, ov_customer, ov_kg,
                     ov_note, period_type, current_target_week))
                conn.commit()
                _scope = " / ".join(p for p in [ov_channel, ov_product, ov_customer] if p != OVERRIDE_ANY)
                st.success(f"Override set — {_scope} now forecasts at {ov_kg:,.0f} kg "
                           f"({period_type.lower()}), replacing the auto+pipeline number.")
                st.rerun()

    active_overrides_display = pd.read_sql("SELECT * FROM manual_overrides WHERE active = 1 ORDER BY id DESC", conn)
    if active_overrides_display.empty:
        st.info("No active overrides — every forecast is currently coming from the auto method + pipeline events.")
    else:
        st.markdown("**Active overrides**")
        _disp = active_overrides_display.copy()
        if "customer" not in _disp.columns:
            _disp["customer"] = OVERRIDE_ANY
        _disp["customer"] = _disp["customer"].fillna(OVERRIDE_ANY)
        # show how broad each rule is, so it's obvious at a glance which one would win
        _disp["scope"] = _disp.apply(
            lambda r: "Most specific" if override_specificity(r) == 3
            else ("Specific" if override_specificity(r) == 2 else "Broad"), axis=1)
        st.dataframe(_disp[["channel", "product", "customer", "override_kg", "scope", "period_type",
                            "submitted_by", "note", "timestamp"]],
                     use_container_width=True)
        st.caption(f"A dimension showing {OVERRIDE_ANY} is a wildcard — it applies to everything in that "
                   "dimension. When more than one override could apply to the same forecast, the most "
                   "specific one wins.")

        _labels = (_disp["channel"].astype(str) + " — " + _disp["product"].astype(str)
                   + " — " + _disp["customer"].astype(str)).tolist()
        _ids = _disp["id"].tolist()
        turn_off_label = st.selectbox(
            "Turn off an override (reverts that scope back to the auto forecast)", _labels)
        if st.button("Turn off this override"):
            # turn off by row id, not by re-parsing the label -- a channel or item containing
            # an em dash would have broken the old split-based matching
            _sel_id = _ids[_labels.index(turn_off_label)]
            conn.execute("UPDATE manual_overrides SET active = 0 WHERE id = ?", (int(_sel_id),))
            conn.commit()
            st.warning("Override turned off — reverting to the auto forecast.")
            st.rerun()

    # surface ongoing overrides that real actuals have consistently contradicted -- an
    # ongoing override never expires on its own (by design), so without this a stale one
    # could quietly distort numbers indefinitely with nobody noticing
    stale_flags = check_stale_ongoing_overrides(weekly_actual, active_overrides)
    if stale_flags:
        st.warning(f"{len(stale_flags)} ongoing override(s) look out of step with recent actuals — worth a check.")
        stale_df = pd.DataFrame(stale_flags).rename(columns={
            "override_kg": "Override (kg)", "recent_avg_actual_kg": "Recent avg actual (kg)",
            "off_by_pct": "Off by %", "weeks_checked": "Weeks checked"})
        st.dataframe(stale_df, use_container_width=True, hide_index=True)
        st.caption("These aren't removed automatically — an ongoing override may still be correct "
                   "(e.g. a real, permanent change in how an account orders). This is a prompt to "
                   "confirm it's still right, not an instruction to delete it.")

    # real history, not just what's currently active -- turning an override off updates its
    # status but never deletes the record, yet nowhere in the app previously showed that
    # history. This fixes a real gap: another team should be able to see what was overridden
    # in the past, by whom, and why, not just what's active right now.
    all_overrides_history = pd.read_sql("SELECT * FROM manual_overrides ORDER BY id DESC", conn)
    if not all_overrides_history.empty:
        with st.expander("Override history (active and turned-off)"):
            history_display = all_overrides_history.copy()
            if "customer" not in history_display.columns:
                history_display["customer"] = OVERRIDE_ANY
            history_display["customer"] = history_display["customer"].fillna(OVERRIDE_ANY)
            history_display["status"] = history_display["active"].map({1: "Active", 0: "Turned off"})
            st.dataframe(history_display[["channel", "product", "customer", "override_kg", "period_type",
                                           "submitted_by", "note", "timestamp", "status"]],
                         use_container_width=True)

            # turning one back on -- previously the only way to reinstate a turned-off
            # override was to retype it in the form above, which loses the original note,
            # who set it, and when
            _off = history_display[history_display["active"] == 0]
            if not _off.empty:
                st.markdown("**Turn one back on**")
                _off_labels = (_off["channel"].astype(str) + " — " + _off["product"].astype(str)
                               + " — " + _off["customer"].astype(str)
                               + "  ·  " + _off["override_kg"].round(0).astype(int).astype(str) + " kg"
                               + "  ·  set by " + _off["submitted_by"].fillna("?").astype(str)).tolist()
                _off_ids = _off["id"].tolist()
                _pick = st.selectbox("Turned-off overrides", _off_labels, key="reactivate_pick")
                if st.button("Turn this override back on"):
                    _rid = _off_ids[_off_labels.index(_pick)]
                    _row = all_overrides_history[all_overrides_history["id"] == _rid].iloc[0]
                    _cust = _row.get("customer") or OVERRIDE_ANY
                    # deactivate anything with the same scope first, so reinstating this one
                    # doesn't leave two competing rules at identical specificity
                    conn.execute("""UPDATE manual_overrides SET active = 0
                        WHERE channel = ? AND product = ? AND COALESCE(customer, ?) = ? AND active = 1""",
                        (_row["channel"], _row["product"], OVERRIDE_ANY, _cust))
                    conn.execute("UPDATE manual_overrides SET active = 1 WHERE id = ?", (int(_rid),))
                    conn.commit()
                    st.success("Override reinstated — it applies again everywhere in the app.")
                    st.rerun()

# --- TAB: Sales plan (S&OP) ---
# --- TAB: Ops capacity check ---
with tab_ops:
    st.subheader("Operations: enter capacity and check against the plan")
    if forecast_by_cp.empty:
        st.warning("No forecast yet — upload sales data in tab 1 first.")
    else:
        existing_cap = pd.read_sql("SELECT * FROM ops_capacity WHERE cycle_label = ? ORDER BY id DESC LIMIT 1",
                                    conn, params=(cycle,))
        existing_cap = existing_cap.iloc[0] if not existing_cap.empty else None
        with st.form("ops_form"):
            ops_name = st.text_input("Your name", value=existing_cap["submitted_by"] if existing_cap is not None else "")
            monthly_capacity = st.number_input(
                "Monthly production capacity (kg)",
                value=float(existing_cap["monthly_capacity_kg"]) if existing_cap is not None else 4000.0)
            ops_notes = st.text_area("Notes", value=existing_cap["notes"] if existing_cap is not None else "")
            if st.form_submit_button("Save capacity", type="primary"):
                conn.execute("""INSERT INTO ops_capacity (timestamp, submitted_by, cycle_label, monthly_capacity_kg, notes)
                    VALUES (?,?,?,?,?)""", (datetime.now().isoformat(), ops_name, cycle, monthly_capacity, ops_notes))
                conn.commit()
                st.success("Capacity saved.")
                st.rerun()

        cap_row = pd.read_sql("SELECT * FROM ops_capacity WHERE cycle_label = ? ORDER BY id DESC LIMIT 1",
                               conn, params=(cycle,))
        if not cap_row.empty:
            cap = cap_row.iloc[0]["monthly_capacity_kg"]
            weekly_planned = forecast_by_cp["forecast_kg"].sum()
            monthly_planned = weekly_planned * 4.345
            status = "SHORTFALL" if cap < monthly_planned else "OK"
            st.metric("Monthly-equivalent planned kg vs capacity",
                       f"{monthly_planned:,.0f} kg planned / {cap:,.0f} kg capacity", status)
            if status == "SHORTFALL":
                st.error(f"Capacity shortfall of {monthly_planned - cap:,.0f} kg/month — flag for sign-off discussion.")
            else:
                st.success("Capacity covers this plan.")

# --- TAB 5: Sign-off ---
with tab_signoff:
    st.subheader("Consensus sign-off")
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Sales sign-off**")
        sales_signer = st.text_input("Sales rep name", key="sales_signer")
        if st.button("Sign off — Sales"):
            conn.execute("INSERT INTO signoffs (timestamp, cycle_label, role, name) VALUES (?,?,?,?)",
                         (datetime.now().isoformat(), cycle, "Sales", sales_signer))
            conn.commit()
            st.success("Sales sign-off recorded.")
    with col2:
        st.markdown("**Ops sign-off**")
        ops_signer = st.text_input("Ops rep name", key="ops_signer")
        if st.button("Sign off — Ops"):
            conn.execute("INSERT INTO signoffs (timestamp, cycle_label, role, name) VALUES (?,?,?,?)",
                         (datetime.now().isoformat(), cycle, "Ops", ops_signer))
            conn.commit()
            st.success("Ops sign-off recorded.")

    signoffs = pd.read_sql("SELECT * FROM signoffs WHERE cycle_label = ? ORDER BY id DESC", conn, params=(cycle,))
    has_sales = (signoffs["role"] == "Sales").any() if not signoffs.empty else False
    has_ops = (signoffs["role"] == "Ops").any() if not signoffs.empty else False
    if has_sales and has_ops:
        st.success(f"Cycle {cycle} APPROVED by both sides.")
    else:
        st.info(f"Pending: {'Sales OK' if has_sales else 'Sales -'} / {'Ops OK' if has_ops else 'Ops -'}")
    st.dataframe(signoffs, use_container_width=True)

# --- TAB: Ask AI ---
with tab_history:
    st.subheader("Business summary")
    if not has_data:
        st.info("Upload sales data to see the business summary.")
    else:
        _h = sales_df.copy()
        _h["record_date"] = pd.to_datetime(_h["record_date"], errors="coerce")
        _h = _h.dropna(subset=["record_date"])
        _win = st.radio("Period", ["Last 90 days", "Last 12 months", "All history"],
                        horizontal=True, key="hist_win")
        if _win != "All history" and not _h.empty:
            _days = 90 if _win.startswith("Last 90") else 365
            _h = _h[_h["record_date"] >= _h["record_date"].max() - pd.Timedelta(days=_days)]
        _unit = st.radio("Measure by", ["Volume (kg)", "Revenue ($)"], horizontal=True, key="hist_unit")
        _col = "kg" if _unit.startswith("Volume") else "revenue"
        _fmt = "{:,.0f} kg" if _col == "kg" else "${:,.0f}"

        if _h.empty:
            st.info("No records in that period.")
        else:
            k1, k2, k3, k4 = st.columns(4)
            k1.metric(f"Total {'kg' if _col=='kg' else 'revenue'}", _fmt.format(_h[_col].sum()))
            k2.metric("Distinct items", f"{_h['product'].nunique():,}")
            k3.metric("Channels", f"{_h['channel'].nunique():,}")
            _ncust = _h.loc[_h["customer"] != "(not tracked)", "customer"].nunique() \
                if "customer" in _h.columns else 0
            k4.metric("Customers", f"{_ncust:,}" if _ncust else "not tracked")

            def _bar(df, dim, title, n=10):
                g = df.groupby(dim, as_index=False)[_col].sum().sort_values(_col, ascending=True).tail(n)
                if g.empty:
                    return
                _tot = df[_col].sum()
                g["_share"] = g[_col] / _tot * 100 if _tot else 0
                fig = go.Figure(go.Bar(
                    x=g[_col], y=g[dim].astype(str), orientation="h", marker_color="#2F6F6B",
                    text=[f"{v:,.0f} ({s:.0f}%)" for v, s in zip(g[_col], g["_share"])],
                    textposition="auto"))
                fig.update_layout(height=max(240, 30 * len(g)), margin=dict(l=10, r=10, t=40, b=10),
                                   plot_bgcolor="white", title=title,
                                   xaxis_title=f"{'kg sold' if _col=='kg' else 'revenue ($)'} — {_win.lower()}",
                                   xaxis=dict(showgrid=True, gridcolor="rgba(0,0,0,0.06)"))
                st.plotly_chart(fig, use_container_width=True)

            c1, c2 = st.columns(2)
            with c1:
                _bar(_h, "product", "Top 10 items", 10)
                _bar(_h, "channel", "By channel", 10)
            with c2:
                if _ncust:
                    _bar(_h[_h["customer"] != "(not tracked)"], "customer", "Top 5 customers", 5)
                if "size_label" in _h.columns:
                    _bar(_h, "size_label", "By bag size", 10)

            if "product_type" in _h.columns:
                _mix = _h.groupby("product_type", as_index=False)[_col].sum()
                _mix["share"] = (_mix[_col] / _mix[_col].sum() * 100).round(1)
                st.markdown("**Staple vs Single mix**")
                st.dataframe(_mix.rename(columns={
                    "product_type": "Type", _col: f"Total {'kg' if _col=='kg' else '$'}",
                    "share": "% of total"}), use_container_width=True, hide_index=True)

            # concentration -- a real risk signal for stakeholders
            _top = _h.groupby("product", as_index=False)[_col].sum().sort_values(_col, ascending=False)
            if len(_top) >= 5:
                _c5 = _top.head(5)[_col].sum() / _top[_col].sum() * 100
                st.caption(f"Top 5 items account for **{_c5:.0f}%** of "
                           f"{'volume' if _col=='kg' else 'revenue'} in this period — "
                           "a useful concentration check when planning supply.")

    st.divider()
    st.subheader("Backup / export everything")
    st.markdown("**Sales records**")
    # show a capped preview, and make the full CSV opt-in. st.download_button computes its
    # data EAGERLY -- so sales_df.to_csv() was converting the entire sales history into a
    # CSV string on every single page load, whether or not anyone ever clicked download.
    st.caption(f"{len(sales_df):,} rows total — showing the most recent 500.")
    st.dataframe(sales_df.tail(500), use_container_width=True)
    if st.button("Prepare sales records download"):
        st.download_button("Download sales_records.csv", sales_df.to_csv(index=False), "sales_records.csv")
    st.markdown("**Auto-generated forecasts (frozen predictions, for accuracy tracking)**")
    # opt-in: Streamlit renders EVERY tab on every rerun, so this was downloading 500 rows
    # from the database on every single page load even if nobody opened this tab. That is
    # billed egress, and it added up fast.
    if st.button("Show recent forecasts"):
        af = pd.read_sql("SELECT * FROM auto_forecasts ORDER BY id DESC LIMIT 500", conn)
        st.caption("Showing the 500 most recent.")
        st.dataframe(af, use_container_width=True)
    if st.button("Prepare full forecast history download"):
        af_full = pd.read_sql("SELECT * FROM auto_forecasts ORDER BY id DESC", conn)
        st.download_button("Download auto_forecasts.csv", af_full.to_csv(index=False), "auto_forecasts.csv")
    st.markdown("**Pipeline events**")
    _hh = all_events_all.copy()
    _hh["status"] = _hh["active"].map({1: "Active", 0: "Inactive"})
    st.dataframe(_hh, use_container_width=True)
    st.markdown("**Sign-offs**")
    st.dataframe(pd.read_sql("SELECT * FROM signoffs ORDER BY id DESC", conn), use_container_width=True)
    st.caption("Download buttons above back up all data as CSV — recommended periodically, "
               "since free-tier hosting can reset local storage on redeploy.")
